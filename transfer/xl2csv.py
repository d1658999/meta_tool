import os
import csv
import pathlib
from openpyxl import load_workbook

# Specify the folder containing Excel files
folder_name = 'transfer'
folder_path = pathlib.Path.cwd()

# Create a list to store data from all tabs in all Excel files
all_data = []

# Loop through all files in the specified folder
header = None
for file_name in os.listdir(folder_path):
    # Check if the file is an Excel file
    if file_name.endswith('.xlsx'):
        # Load the Excel file
        file_path = os.path.join(folder_path, file_name)
        wb = load_workbook(filename=file_path)
        # Loop through all tabs in the Excel file
        for sheet_name in wb.sheetnames:
            # Load the sheet and append its data to the all_data list
            if 'Raw_Data' in sheet_name and header is None:
                sheet = wb[sheet_name]
                rows = sheet.values
                header = next(rows)
                all_data.extend([header] + [row for row in rows])
                flag = 1
            elif 'Raw_Data' in sheet_name and header is not None:
                sheet = wb[sheet_name]
                rows = sheet.values
                next(rows)
                all_data.extend([row for row in rows])

# Write the all_data list to a CSV file
csv_file_name = 'output.csv'
csv_path = pathlib.Path.cwd() / pathlib.Path(csv_file_name)

with open(csv_path, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerows(all_data)
