import time
import datetime
import openpyxl
from openpyxl.chart import LineChart, Reference, BarChart, Series
import pathlib

from utils.loss_handler import get_loss
from utils.log_init import log_set
from connection_interface.connection_serial import ModemComport
from connection_interface.connection_visa import VisaComport
# import fcc
# import ce
# import test_scripts
# # from varname import nameof
# import math

# import common_parameters_ftm as cm_pmt_ftm
# import want_test_band as wt
# from loss_table import loss_table
# from adb_control import get_odpm_current

logger = log_set()

class Cmw100:
    def __init__(self, psu_object=None):
        self.ser = ModemComport()
        self.cmw100 = VisaComport('cmw100')

        # self.parameters_init()
        # self.psu = psu_object
        pass

    def parameters_init(self):
        self.asw_on_off = 0  # 1: AS ON, 0: AS OFF
        self.asw_tech_dict = {
            'GSM': 0,
            'WCDMA': 1,
            'LTE': 2,
            'FR1': 6,
        }
        self.asw_tech = None
        self.asw_path = wt.asw_path
        self.srs_path = wt.srs_path
        self.srs_path_enable = wt.srs_path_enable
        self.rx_path_lte = None
        self.rx_path_fr1 = None
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = None
        self.port_tx = None
        self.port_rx = None
        self.chan = None
        self.sync_path = wt.sync_path  # 0: Main, 1: CA#1, 2: CA#2, 3: CA#3
        self.sync_mode = 0  # 0: MAIN , 1: 4RX, 2: 6RX
        self.sa_nsa_mode = None  # SA: 0, NSA: 1
        self.filename = None
        self.tech = None
        self.band_lte = None
        self.band_fr1 = None
        self.bw_lte = None
        self.bw_fr1 = None
        self.band_segment = wt.band_segment
        self.band_segment_fr1 = wt.band_segment_fr1
        self.tx_freq_lte = None
        self.rx_freq_lte = None
        self.tx_freq_fr1 = None
        self.rx_freq_fr1 = None
        self.mcs_lte = None
        self.mcs_fr1 = None
        self.tx_path = None
        self.scs = None
        self.type_fr1 = None
        self.rb_state = None
        self.rb_size_lte = None
        self.rb_size_fr1 = None
        self.rb_start_lte = None
        self.rb_start_fr1 = None
        self.loss_tx = None
        self.loss_rx = None
        self.uldl_period = None
        self.dl_slot = None
        self.dl_symbol = None
        self.ul_slot = None
        self.ul_symbol = None
        self.mod_gsm = None
        self.tx_1rb_filename_judge = None
        self.tx_path_dict = {
            'TX1': 0,
            'TX2': 1,
            'MIMO': 20,
        }
        self.bw_lte_dict = {
            1.4: 0,
            3: 1,
            5: 2,
            10: 3,
            15: 4,
            20: 5,
        }
        self.bw_fr1_dict = {
            5: 0,
            10: 1,
            15: 2,
            20: 3,
            25: 4,
            30: 5,
            40: 6,
            50: 7,
            60: 8,
            80: 9,
            90: 10,
            100: 11,
            70: 12,
        }
        self.mcs_lte_dict = {
            'QPSK': 0,
            'Q16': 11,
            'Q64': 25,
            'Q256': 27,
        }
        self.mcs_fr1_dict = {
            'BPSK': 1,
            'QPSK': 2,
            'Q16': 4,
            'Q64': 6,
            'Q256': 8,
        }
        self.rb_select_lte_dict = {
            'PRB': 0,
            'FRB': 1,
        }
        self.type_dict = {
            'DFTS': 0,
            'CP': 1,
        }
        self.rb_alloc_fr1_dict = {
            'EDGE_FULL_LEFT': 0,
            'EDGE_FULL_RIGHT': 1,
            'EDGE_1RB_LEFT': 2,
            'EDGE_1RB_RIGHT': 3,
            'OUTER_FULL': 4,
            'INNER_FULL': 5,
            'INNER_1RB_LEFT': 6,
            'INNER_1RB_RIGHT': 7,
        }
        self.scs_dict = {
            15: 0,
            30: 1,
            60: 2,
        }
        self.rx_path_gsm_dict = {
            2: 'RX0',
            1: 'RX1',
            4: 'RX2',
            8: 'RX3',
            3: 'RX0+RX1',
            12: 'RX2+RX3',
            15: 'ALL PATH',
        }
        self.rx_path_wcdma_dict = {
            2: 'RX0',
            1: 'RX1',
            4: 'RX2',
            8: 'RX3',
            3: 'RX0+RX1',
            12: 'RX2+RX3',
            15: 'ALL PATH',
        }
        self.rx_path_lte_dict = {
            2: 'RX0',
            1: 'RX1',
            4: 'RX2',
            8: 'RX3',
            3: 'RX0+RX1',
            12: 'RX2+RX3',
            15: 'ALL PATH',
        }
        self.rx_path_fr1_dict = {
            2: 'RX0',
            1: 'RX1',
            4: 'RX2',
            8: 'RX3',
            3: 'RX0+RX1',
            12: 'RX2+RX3',
            15: 'ALL PATH',
        }
        self.duty_cycle_dict = {
            100: (6, 0, 0, 10, 0),
            50: (6, 5, 0, 5, 0),
        }
        self.sync_path_dict = {
            'Main': 0,
            'CA#1': 1,
            'CA#2': 2,
            'CA#3': 3,
        }
        self.band_tx_set_dict_gsm = {
            900: 0,
            1800: 1,
            1900: 2,
            850: 3,
        }

        self.band_tx_meas_dict_gsm = {
            900: 'G09',
            1800: 'G18',
            1900: 'G19',
            850: 'G085',
        }
        self.mod_dict_gsm = {
            'GMSK': 0,
            'EPSK': 1,
        }

    def get_temperature(self):
        """
        for P22, AT+GOOGTHERMISTOR=1,1 for MHB LPAMid/ MHB Rx1 LFEM, AT+GOOGTHERMISTOR=0,1
        for LB LPAMid, MHB ENDC LPAMid, UHB(n77/n79 LPAF)
        :return:
        """
        res0 = self.command('AT+GOOGTHERMISTOR=0,1')
        res1 = self.command('AT+GOOGTHERMISTOR=1,1')
        res_list = [res0, res1]
        therm_list = []
        for res in res_list:
            for r in res:
                if 'TEMPERATURE' in r.decode().strip():
                    try:
                        temp = eval(r.decode().strip().split(':')[1]) / 1000
                        therm_list.append(temp)
                    except Exception as err:
                        logger.debug(err)
                        therm_list.appned(None)
        logger.info(f'thermistor0 get temp: {therm_list[0]}')
        logger.info(f'thermistor1 get temp: {therm_list[1]}')
        return therm_list

    def measure_current(self):
        if wt.odpm_enable:
            return get_odpm_current()
        elif wt.psu_enable:
            return self.psu.current_average()
        else:
            return None

    def preset_instrument(self):
        logger.info('----------Preset CMW100----------')
        self.command_cmw100_write('SYSTem:PRESet:ALL')
        self.command_cmw100_query('SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write('CONFigure:FDCorrection:DEACtivate:ALL')
        self.command_cmw100_write('CONFigure:BASE:FDCorrection:CTABle:DELete:ALL')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write('*RST')
        self.command_cmw100_query('*OPC?')

    def test_reset_gsm(self):
        print("----------EDGE Test Reset----------")
        self.command(f'AT+TESTRESET')

    @staticmethod
    def chan_judge_gsm(band_gsm, rx_freq_gsm):
        if rx_freq_gsm < cm_pmt_ftm.dl_freq_selected('GSM', band_gsm, 0)[1]:
            return 'L'
        elif rx_freq_gsm == cm_pmt_ftm.dl_freq_selected('GSM', band_gsm, 0)[1]:
            return 'M'
        elif rx_freq_gsm > cm_pmt_ftm.dl_freq_selected('GSM', band_gsm, 0)[1]:
            return 'H'

    @staticmethod
    def chan_judge_wcdma(band_wcdma, tx_freq_wcdma):
        rx_freq_wcdma = cm_pmt_ftm.transfer_freq_tx2rx_wcdma(band_wcdma, tx_freq_wcdma)
        if rx_freq_wcdma < cm_pmt_ftm.dl_freq_selected('WCDMA', band_wcdma, 5)[1]:
            return 'L'
        elif rx_freq_wcdma == cm_pmt_ftm.dl_freq_selected('WCDMA', band_wcdma, 5)[1]:
            return 'M'
        elif rx_freq_wcdma > cm_pmt_ftm.dl_freq_selected('WCDMA', band_wcdma, 5)[1]:
            return 'H'

    @staticmethod
    def chan_judge_lte(band_lte, bw_lte, tx_freq_lte):
        rx_freq_lte = cm_pmt_ftm.transfer_freq_tx2rx_lte(band_lte, tx_freq_lte)
        if rx_freq_lte < cm_pmt_ftm.dl_freq_selected('LTE', band_lte, bw_lte)[1]:
            return 'L'
        elif rx_freq_lte == cm_pmt_ftm.dl_freq_selected('LTE', band_lte, bw_lte)[1]:
            return 'M'
        elif rx_freq_lte > cm_pmt_ftm.dl_freq_selected('LTE', band_lte, bw_lte)[1]:
            return 'H'

    @staticmethod
    def chan_judge_fr1(band_fr1, bw_fr1, tx_freq_fr1):
        rx_freq_fr1 = cm_pmt_ftm.transfer_freq_tx2rx_fr1(band_fr1, tx_freq_fr1)
        if rx_freq_fr1 < cm_pmt_ftm.dl_freq_selected('FR1', band_fr1, bw_fr1)[1]:
            return 'L'
        elif rx_freq_fr1 == cm_pmt_ftm.dl_freq_selected('FR1', band_fr1, bw_fr1)[1]:
            return 'M'
        elif rx_freq_fr1 > cm_pmt_ftm.dl_freq_selected('FR1', band_fr1, bw_fr1)[1]:
            return 'H'

    def set_test_mode_gsm(self):
        logger.info('----------Set Test Mode----------')
        self.command(f'AT+HNSSTOP')
        self.command(f'AT+TESTSTR')

    def set_test_mode_wcdma(self):
        logger.info('----------Set Test Mode----------')
        self.command(f'AT+HSPATMSTART')

    def set_test_mode_lte(self):
        logger.info('----------Set Test Mode----------')
        self.command(f'AT+LRFFINALSTART=1,{self.band_lte}')
        self.command(f'AT+LMODETEST')
        # self.command_cmw100_query('*OPC?')

    def set_test_mode_fr1(self):  # SA: 0, NSA: 1
        """
        SA: 0, NSA: 1
        """
        logger.info('----------Set Test Mode----------')
        self.command(f'AT+NRFFINALSTART={self.band_fr1},{self.sa_nsa_mode}')
        # self.command_cmw100_query('*OPC?')

    def set_test_end_gsm(self, delay=0.2):
        logger.info('----------Set End----------')
        self.command(f'AT+TESTEND', delay)
        # self.command_cmw100_query('*OPC?')

    def set_test_end_wcdma(self, delay=0.2):
        logger.info('----------Set End----------')
        self.command(f'AT+HSPATMEND', delay)
        # self.command_cmw100_query('*OPC?')

    def set_test_end_lte(self, delay=0.2):
        logger.info('----------Set End----------')
        self.command(f'AT+LRFFINALFINISH', delay)
        # self.command_cmw100_query('*OPC?')

    def set_test_end_fr1(self, delay=0.2):
        logger.info('----------Set End----------')
        self.command(f'AT+NRFFINALFINISH', delay)
        # self.command_cmw100_query('*OPC?')

    def set_gprf_measurement(self):
        logger.info('----------set GPRF Measurement----------')
        self.command_cmw100_write('CONF:GPRF:MEAS:POW:FILT:TYPE BAND')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:FILT:BAND:BWID {self.bw_fr1}MHz')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:SCO 2')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:REP SING')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:LIST OFF')
        self.command_cmw100_write(f"TRIGger:GPRF:MEAS:POWer:SOURce 'Free Run'")
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:TRIG:SLOP REDG')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:SLEN 5.0e-3')
        # self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:MLEN 8.0e-4')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:MLEN 5.0e-3')
        # self.command_cmw100_write(f'TRIGger:GPRF:MEAS:POWer:OFFSet 2.1E-3')
        # self.command_cmw100_write(f'TRIGger:GPRF:MEAS:POWer:OFFSet 5E-4')
        self.command_cmw100_write(f'TRIGger:GPRF:MEAS:POWer:OFFSet 0')
        self.command_cmw100_write(f'TRIG:GPRF:MEAS:POW:MODE ONCE')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:ENP {self.tx_level}')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:EATT {self.loss_tx}')

    def set_gprf_tx_freq(self):
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:FREQ {self.tx_freq_fr1}KHz')

    def get_gprf_power(self):
        self.command_cmw100_write('INIT:GPRF:MEAS:POW')
        self.command_cmw100_query('*OPC?')
        f_state = self.command_cmw100_query('FETC:GPRF:MEAS:POW:STAT?')
        while f_state != 'RDY':
            f_state = self.command_cmw100_query('FETC:GPRF:MEAS:POW:STAT?')
            self.command_cmw100_query('*OPC?')
        power_average = round(eval(self.command_cmw100_query('FETC:GPRF:MEAS:POWer:AVER?'))[1], 2)
        logger.info(f'Get the GPRF power: {power_average}')
        return power_average

    def sig_gen_gsm(self):
        logger.info('----------Sig Gen----------')
        self.command_cmw100_query('SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write('ROUT:GPRF:GEN:SCEN:SAL R118, TX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:LIST OFF')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:EATT {self.loss_rx}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:BBM ARB')
        waveform_2g = '2G_FINAL.wv'
        self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\{waveform_2g}'")
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SOUR:GPRF:GEN1:ARB:FILE?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:FREQ {self.rx_freq_gsm}KHz')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:LEV {self.rx_level}')
        gprf_gen = self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')
        self.command_cmw100_query('*OPC?')
        if gprf_gen == 'OFF':
            self.command_cmw100_write('SOUR:GPRF:GEN1:STAT ON')
            self.command_cmw100_query('*OPC?')

    def sig_gen_wcdma(self):
        logger.info('----------Sig Gen----------')
        self.command_cmw100_query('SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write('ROUT:GPRF:GEN:SCEN:SAL R118, TX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:LIST OFF')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:EATT {self.loss_rx}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:BBM ARB')
        waveform_3g = '3G_CAL_FINAL.wv'
        self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\{waveform_3g}'")
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SOUR:GPRF:GEN1:ARB:FILE?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:FREQ {self.rx_freq_wcdma}KHz')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:LEV {self.rx_level}')
        gprf_gen = self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')
        self.command_cmw100_query('*OPC?')
        if gprf_gen == 'OFF':
            self.command_cmw100_write('SOUR:GPRF:GEN1:STAT ON')
            self.command_cmw100_query('*OPC?')

    def sig_gen_lte(self):
        logger.info('----------Sig Gen----------')
        self.command_cmw100_query('SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write('ROUT:GPRF:GEN:SCEN:SAL R118, TX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:LIST OFF')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:EATT {self.loss_rx}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:BBM ARB')
        self.command_cmw100_query('*OPC?')
        self.band_lte = int(self.band_lte)
        if self.band_lte in [34, 38, 39, 40, 41, 42, 48]:
            self.command_cmw100_write(
                f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_Channel_CC0_RxAnt0_RF_Verification_10M_SIMO_01.wv'")
        else:
            # self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_Ant0_FRC_10MHz.wv'")
            bw_lte = '1p4' if self.bw_lte == 1.4 else '03' if self.bw_lte == 3 else self.bw_lte
            self.command_cmw100_write(f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_Ant0_FRC_{bw_lte}MHz.wv'")
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SOUR:GPRF:GEN1:ARB:FILE?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:FREQ {self.rx_freq_lte}KHz')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:LEV {self.rx_level}')
        gprf_gen = self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')
        self.command_cmw100_query('*OPC?')
        if gprf_gen == 'OFF':
            self.command_cmw100_write('SOUR:GPRF:GEN1:STAT ON')
            self.command_cmw100_query('*OPC?')

    def sig_gen_fr1(self):
        """
        scs: FDD is forced to 15KHz and TDD is to be 30KHz
        """
        logger.info('----------Sig Gen----------')
        self.band_fr1 = int(self.band_fr1)
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 77, 78,
                                     79] else 0  # for now FDD is forced to 15KHz and TDD is to be 30KHz
        scs = 15 * (2 ** scs)
        mcs_fr1_wv = 4
        self.scs = scs
        self.command_cmw100_query('SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write('ROUT:GPRF:GEN:SCEN:SAL R118, TX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:GPRF:GEN:CMWS:USAGe:TX:ALL R118, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:LIST OFF')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:EATT {self.loss_rx}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('SOUR:GPRF:GEN1:BBM ARB')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write('CONFigure:NRSub:MEAS:ULDL:PERiodicity MS10')
        self.command_cmw100_query('*OPC?')
        if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 77, 78, 79]:
            self.command_cmw100_write(
                f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_NR_Ant0_NR_{self.bw_fr1}MHz_SCS{scs}_TDD_Sens_MCS{mcs_fr1_wv}_rescale.wv'")
        else:
            self.command_cmw100_write(
                f"SOUR:GPRF:GEN1:ARB:FILE 'C:\CMW100_WV\SMU_NodeB_NR_Ant0_LTE_NR_{self.bw_fr1}MHz_SCS{scs}_FDD_Sens_MCS_{mcs_fr1_wv}.wv'")
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SOUR:GPRF:GEN1:ARB:FILE?')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:FREQ {self.rx_freq_fr1}KHz')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:LEV {self.rx_level}')
        gprf_gen = self.command_cmw100_query('SOUR:GPRF:GEN1:STAT?')
        self.command_cmw100_query('*OPC?')
        if gprf_gen == 'OFF':
            self.command_cmw100_write('SOUR:GPRF:GEN1:STAT ON')
            self.command_cmw100_query('*OPC?')

    def sync_gsm(self):
        logger.info('---------Sync----------')
        self.command(f'AT+TESTRESET', delay=0.2)
        self.command(
            f'AT+TESTSYNC={self.band_tx_set_dict_gsm[self.band_gsm]},0,{self.rx_chan_gsm},{-1 * int(round(self.rx_level, 0))}',
            delay=0.5)

    def sync_wcdma(self):
        logger.info('---------Sync----------')
        self.command(f'AT+HDLSYNC={self.rx_chan_wcdma}', delay=0.5)

    def sync_lte(self):
        logger.info('---------Sync----------')
        response = self.command(f'AT+LSYNC={self.sync_path_dict[self.sync_path]},{self.sync_mode},{self.rx_freq_lte}',
                                delay=1.2)
        while b'+LSYNC:1\r\n' not in response:
            logger.info('**********Sync repeat**********')
            time.sleep(1)
            response = self.command(
                f'AT+LSYNC={self.sync_path_dict[self.sync_path]},{self.sync_mode},{self.rx_freq_lte}', delay=2)

    def sync_fr1(self):
        logger.info('---------Sync----------')
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else 0
        response = self.command(
            f'AT+NRFSYNC={self.sync_path_dict[self.sync_path]},{self.sync_mode},{scs},{self.bw_fr1_dict[self.bw_fr1]},0,{self.rx_freq_fr1}',
            delay=1)
        while b'+NRFSYNC:1\r\n' not in response:
            logger.info('**********Sync repeat**********')
            time.sleep(1)
            response = self.command(
                f'AT+NRFSYNC={self.sync_path_dict[self.sync_path]},{self.sync_mode},{scs},{self.bw_fr1_dict[self.bw_fr1]},0,{self.rx_freq_fr1}',
                delay=2)

    def tx_set_gsm(self):
        logger.info('---------Tx Set----------')
        self.command(
            f'AT+TESTTX={self.band_tx_set_dict_gsm[self.band_gsm]},{self.mod_dict_gsm[self.mod_gsm]},{self.rx_chan_gsm},1,1')
        self.command(f'AT+TESTPWR=0,{self.pcl},{self.pcl},{self.pcl},{self.pcl}')
        logger.info(f'Band: {self.band_gsm}, Modulation: {self.mod_gsm}, Chan: {self.rx_chan_gsm}, PCL: {self.pcl}')

    def tx_set_wcdma(self):
        logger.info('---------Tx Set----------')
        self.command(f'AT+HDELULCHAN')
        self.command(f'AT+HTXPERSTART={self.tx_chan_wcdma}')
        self.command(f'AT+HSETMAXPOWER={self.tx_level * 10}')
        logger.info(f'Tx_chan: {self.tx_chan_wcdma}, Tx_level: {self.tx_level}')

    def tx_set_lte(self):
        """
        tx_path: TX1: 0 (main path)| TX2: 1 (sub path)
        bw_lte: 1.4: 0 | 3: 1 | 5: 2 | 10: 3 | 15: 4 | 20: 5
        tx_freq_lte:
        rb_num:
        rb_start:
        mcs: "QPSK": 0 | "Q16": 11 | "Q64": 25 | "Q256": 27
        pwr:

        """
        logger.info('---------Tx Set----------')
        self.command(
            f'AT+LTXSENDREQ={self.tx_path_dict[self.tx_path]},{self.bw_lte_dict[self.bw_lte]},{self.tx_freq_lte},{self.rb_size_lte},{self.rb_start_lte},{self.mcs_lte_dict[self.mcs_lte]},2,1,{self.tx_level}')
        logger.info(
            f'TX_PATH: {self.tx_path}, BW: {self.bw_lte}, TX_FREQ: {self.tx_freq_lte}, RB_SIZE: {self.rb_size_lte}, RB_OFFSET: {self.rb_start_lte}, MCS: {self.mcs_lte}, TX_LEVEL: {self.tx_level}')
        # self.command_cmw100_query('*OPC?')

    def tx_set_fr1(self):
        logger.info('---------Tx Set----------')
        self.command(
            f'AT+NTXSENDREQ={self.tx_path_dict[self.tx_path]},{self.tx_freq_fr1},{self.bw_fr1_dict[self.bw_fr1]},{self.scs_dict[self.scs]},{self.rb_size_fr1},{self.rb_start_fr1},{self.mcs_fr1_dict[self.mcs_fr1]},{self.type_dict[self.type_fr1]},{self.tx_level}')
        logger.info(
            f'TX_PATH: {self.tx_path}, BW: {self.bw_fr1}, TX_FREQ: {self.tx_freq_fr1}, RB_SIZE: {self.rb_size_fr1}, RB_OFFSET: {self.rb_start_fr1}, MCS: {self.mcs_fr1}, TX_LEVEL: {self.tx_level}')
        # self.command_cmw100_query('*OPC?')

    def set_duty_cycle(self):
        logger.info(f'----------Set duty cycle: {wt.duty_cycle}----------')
        if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79]:
            self.uldl_period = self.duty_cycle_dict[wt.duty_cycle][0]
            self.dl_slot = self.duty_cycle_dict[wt.duty_cycle][1]
            self.dl_symbol = self.duty_cycle_dict[wt.duty_cycle][2]
            self.ul_slot = self.duty_cycle_dict[wt.duty_cycle][3]
            self.ul_symbol = self.duty_cycle_dict[wt.duty_cycle][4]
            logger.info('---TDD, so need to set the duty cycle')
            logger.debug(
                f'Duty Cycle setting: {self.uldl_period}, {self.dl_slot}, {self.dl_symbol}, {self.ul_slot}, {self.ul_symbol}')
        else:
            self.uldl_period = 0
            self.dl_slot = 0
            self.dl_symbol = 0
            self.ul_slot = 0
            self.ul_symbol = 0
            logger.info("---FDD, so don't need to set the duty cycle")
            logger.debug(
                f'Duty Cycle setting: {self.uldl_period}, {self.dl_slot}, {self.dl_symbol}, {self.ul_slot}, {self.ul_symbol}')

    def tx_set_no_sync_fr1(self):
        logger.info('---------Tx No Set----------')
        self.scs = 30 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else 15
        self.command(
            f'AT+NRFACTREQ={self.tx_path_dict[self.tx_path]},{self.tx_freq_fr1},{self.bw_fr1_dict[self.bw_fr1]},{self.scs_dict[self.scs]},{self.rb_size_fr1},{self.rb_start_fr1},{self.mcs_fr1_dict[self.mcs_fr1]},{self.type_dict[self.type_fr1]},{self.tx_level},{self.uldl_period},{self.dl_slot},{self.dl_symbol},{self.ul_slot},{self.ul_symbol}')
        logger.info(
            f'TX_PATH: {self.tx_path}, BW: {self.bw_fr1}, TX_FREQ: {self.tx_freq_fr1}, RB_SIZE: {self.rb_size_fr1}, RB_OFFSET: {self.rb_start_fr1}, MCS: {self.mcs_fr1}, TX_LEVEL: {self.tx_level}, Duty cycle: {wt.duty_cycle} %')

    def antenna_switch(self):  # 1: AS ON, 0: AS OFF
        logger.info('---------Antenna Switch----------')
        self.command(f'AT+LTXASTUNESET={self.asw_on_off}')
        if self.asw_on_off == 0:
            logger.info('Antenna Switch OFF')
        elif self.asw_on_off == 1:
            logger.info('Antenna Switch ON')

    def antenna_switch_v2(self):
        """
        this is to place on the first to activate
        AT+ANTSWSEL=P0,P1	//Set Tx DPDT switch
        P0: RAT (0=GSM, 1=WCDMA, 2=LTE, 4=CDMA, 6=NR)
        P1: ANT path (0=default, 1=switched, 4=dynamic mode),
        P1 (P0=NR): ANT Path (0=Tx-Ant1, 1=Tx-Ant2, 2=Tx-Ant3, 3=Tx-Ant4, 4=dynamic switch mode)
        tech:
        ant_path:
        :return:
        """
        self.asw_tech = self.tech
        logger.info('---------Antenna Switch----------')
        self.command(f'AT+ANTSWSEL={self.asw_tech_dict[self.asw_tech]},{self.asw_path}')
        logger.info(f'RAT: {self.asw_tech}, ANT_PATH: {self.asw_path}')
        # self.command_cmw100_query('*OPC?')

    def srs_switch(self):
        logger.info('---------SRS Switch----------')
        self.command(f'AT+NTXSRSSWPATHSET={self.srs_path}')
        logger.info(f'SRS_PATH: {self.srs_path}')

    def rx_path_setting_gsm(self):
        """
        0: PRX, 1: DRX
        :return:
        """
        logger.info('----------Rx path setting----------')
        logger.info(f'---------Now is {self.rx_path_gsm_dict[self.rx_path_gsm]}---------')
        rx_path_gsm = None
        if self.rx_path_gsm == 2:
            rx_path_gsm = 0
        elif self.rx_path_gsm == 1:
            rx_path_gsm = 1
        self.command(f'AT+ERXSEL={rx_path_gsm}')
        # self.command_cmw100_query('*OPC?')

    def rx_path_setting_wcdma(self):
        """
        2: PRX, 1: DRX, 3: PRX+DRX 15: ALL PATH
        :return:
        """
        logger.info('----------Rx path setting----------')
        logger.info(f'----------Now is {self.rx_path_wcdma_dict[self.rx_path_wcdma]}---------')
        self.command(f'AT+HRXMODESET={self.rx_path_wcdma}')

        # self.command_cmw100_query('*OPC?')

    def rx_path_setting_lte(self):
        """
        2: PRX, 1: DRX, 4: RX2, 8:RX3, 3: PRX+DRX, 12: RX2+RX3, 15: ALL PATH
        :return:
        """
        if self.rx_path_lte is None:
            self.rx_path_lte = 15
        logger.info('----------Rx path setting----------')
        logger.info(f'----------Now is {self.rx_path_lte_dict[self.rx_path_lte]}---------')
        self.command(f'AT+LRXMODESET={self.rx_path_lte}')
        # self.command_cmw100_query('*OPC?')

    def rx_path_setting_fr1(self):
        """
        2: PRX, 1: DRX, 4: RX2, 8:RX3, 3: PRX+DRX, 12: RX2+RX3, 15: ALL PATH
        :return:
        """
        if self.rx_path_fr1 is None:
            self.rx_path_fr1 = 15
        logger.info('----------Rx path setting----------')
        logger.info(f'----------Now is {self.rx_path_fr1_dict[self.rx_path_fr1]}---------')
        self.command(f'AT+NRXMODESET={self.rx_path_fr1}')
        # self.command_cmw100_query('*OPC?')

    def tx_monitor_lte(self):
        logger.info('---------Tx Monitor----------')
        # self.sig_gen_lte()
        self.command_cmw100_write(f'CONFigure:LTE:MEAS:MEV:RES:PMONitor ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f"TRIG:LTE:MEAS:MEV:SOUR 'GPRF Gen1: Restart Marker'")
        self.command_cmw100_write(f'CONFigure:LTE:MEAS:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'TRIG:LTE:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOEX ON')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CPR NORM')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MSUB 2, 10, 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO ON')
        mode = 'TDD' if self.band_lte in [38, 39, 40, 41, 42, 48] else 'FDD'
        self.command_cmw100_write(f'CONF:LTE:MEAS:DMODe {mode}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:BAND OB{self.band_lte}')
        rb = f'0{self.bw_lte * 10}' if self.bw_lte < 10 else f'{self.bw_lte * 10}'
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CBAN B{rb}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOD:MSCH {self.mcs_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:FREQ {self.tx_freq_lte}KHz')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'ROUT:LTE:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:ENP {self.tx_level}.00')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'INIT:LTE:MEAS:MEV')
        f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
        while f_state != 'RDY':
            f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
            self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('*OPC?')
        power_results = self.command_cmw100_query(f'FETCh:LTE:MEAS:MEV:PMON:AVER?')
        power = power_results.strip().split(',')[2]
        logger.info(f'LTE power by Tx monitor: {round(eval(power), 2)}')
        return round(eval(power), 2)

    def tx_measure_gsm(self):
        logger.info('---------Tx Measure----------')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:PVT 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:SMOD 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:SSW 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:SCEN:ACT STAN')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'ROUT:GSM:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:GSM:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_write(f'CONF:GSM:MEAS:RFS:UMAR 10.00')
        self.command_cmw100_write(f"TRIG:GSM:MEAS:MEV:SOUR 'Power'")
        self.command_cmw100_write(f'TRIG:GSM:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON, OFF, ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(
            f'CONF:GSM:MEAS:MEV:SMOD:OFR 100KHZ,200KHZ,250KHZ,400KHZ,600KHZ,800KHZ,1600KHZ,1800KHZ,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SMOD:EAR ON,6,45,ON,90,129')
        self.command_cmw100_write(
            f'CONF:GSM:MEAS:MEV:SSW:OFR 400KHZ,600KHZ,1200KHZ,1800KHZ,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONF:GSM:MEAS:BAND {self.band_tx_meas_dict_gsm[self.band_gsm]}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:GSM:MEAS:CHAN {self.rx_chan_gsm}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:MOEX ON')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:TSEQ TSC{self.tsc}')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:MVI {",".join([self.mod_gsm] * 8)}')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:MSL 0,1,0')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.power_init_gsm()
        self.command_cmw100_write(f'CONF:GSM:MEAS:RFS:ENP {self.pwr_init_gsm}')
        mod_results = self.command_cmw100_query(
            f'READ:GSM:MEAS:MEV:MOD:AVER?')  # P12 is Power, P6 is phase_err_rms, P2 is EVM_rms, P10 is ferr
        mod_results = mod_results.split(',')
        mod_results = [mod_results[12], mod_results[6], mod_results[7],
                       mod_results[10]]  # power, phase_err_rms, phase_peak, ferr
        mod_results = [round(eval(m), 2) for m in mod_results]
        logger.info(
            f'Power: {mod_results[0]:.2f}, Phase_err_rms: {mod_results[1]:.2f}, Phase_peak: {mod_results[2]:.2f}, Ferr: {mod_results[3]:.2f}')
        self.command_cmw100_write(f'INIT:GSM:MEAS:MEV')
        self.command_cmw100_write(f'*OPC?')
        self.command_cmw100_write(f'CONF:GSM:MEAS:RFS:ENP {mod_results[0]:.2f}')
        f_state = self.command_cmw100_query(f'FETC:GSM:MEAS:MEV:STAT?')
        while f_state != 'RDY':
            time.sleep(0.2)
            f_state = self.command_cmw100_query('FETC:GSM:MEAS:MEV:STAT?')
            self.command_cmw100_query('*OPC?')
        pvt = self.command_cmw100_query(f'FETC:GSM:MEAS:MEV:PVT:AVER:SVEC?')  # PVT, but it is of no use
        orfs_mod = self.command_cmw100_query(f'FETC:GSM:MEAS:MEV:SMOD:FREQ?')  # MOD_ORFS
        orfs_mod = [round(eval(orfs_mod), 2) for orfs_mod in orfs_mod.split(',')[13:29]]
        orfs_mod = [
            orfs_mod[6],  # -200 KHz
            orfs_mod[10],  # 200 KHz
            orfs_mod[4],  # -400 KHz
            orfs_mod[12],  # 400 KHz
            orfs_mod[3],  # -600 KHz
            orfs_mod[13],  # 600 KHz
        ]
        logger.info(f'ORFS_MOD_-200KHz: {orfs_mod[0]}, ORFS_MOD_200KHz: {orfs_mod[1]}')
        logger.info(f'ORFS_MOD_-400KHz: {orfs_mod[2]}, ORFS_MOD_400KHz: {orfs_mod[3]}')
        logger.info(f'ORFS_MOD_-600KHz: {orfs_mod[4]}, ORFS_MOD_600KHz: {orfs_mod[5]}')
        orfs_sw = self.command_cmw100_query(f'FETC:GSM:MEAS:MEV:SSW:FREQ?')  # SW_ORFS
        orfs_sw = [round(eval(orfs_sw), 2) for orfs_sw in orfs_sw.split(',')[17:25]]
        orfs_sw = [
            orfs_sw[3],  # -400 KHz
            orfs_sw[5],  # 400 KHz
            orfs_sw[2],  # -600 KHz
            orfs_sw[6],  # 600 KHz
            orfs_sw[1],  # -1200 KHz
            orfs_sw[7],  # 1200 KHz
        ]
        logger.info(f'ORFS_SW_-400KHz: {orfs_sw[0]}, ORFS_SW_400KHz: {orfs_sw[1]}')
        logger.info(f'ORFS_SW_-600KHz: {orfs_sw[2]}, ORFS_SW_600KHz: {orfs_sw[3]}')
        logger.info(f'ORFS_SW_-1200KHz: {orfs_sw[4]}, ORFS_SW_1200KHz: {orfs_sw[5]}')
        self.command_cmw100_write(f'STOP:WCDM:MEAS:MEV')
        self.command_cmw100_query('*OPC?')

        return mod_results + orfs_mod + orfs_sw  # [0~3] + [4~10] + [11~17]

    def tx_measure_wcdma(self):
        logger.info('---------Tx Measure----------')
        self.command_cmw100_write(f'*CLS')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:SCO:SPEC 5')
        self.command_cmw100_write(f'ROUT:WCDMA:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:UMAR 10.00')
        self.command_cmw100_write(f"TRIG:WCDM:MEAS:MEV:SOUR 'Free Run (Fast sync)'")
        self.command_cmw100_write(f'TRIG:WCDM:MEAS:MEV:THR -30')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:RES:ALL ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:BAND OB{self.band_wcdma}')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:RFS:FREQ {self.tx_chan_wcdma} CH')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:DPDC ON')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:SFOR 0')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:SCOD 13496235')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:ULC WCDM')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:UMAR 10.00')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:RFS:ENP {self.tx_level + 5}')
        mod_results = self.command_cmw100_query(
            f'READ:WCDM:MEAS:MEV:MOD:AVER?')  # P1 is EVM, P4 is Ferr, P8 is IQ Offset
        mod_results = mod_results.split(',')
        mod_results = [mod_results[1], mod_results[4], mod_results[8]]
        mod_results = [eval(m) for m in mod_results]
        logger.info(f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
        self.command_cmw100_write(f'INIT:WCDM:MEAS:MEV')
        self.command_cmw100_write(f'*OPC?')
        f_state = self.command_cmw100_query(f'FETC:WCDM:MEAS:MEV:STAT?')
        while f_state != 'RDY':
            time.sleep(0.2)
            f_state = self.command_cmw100_query('FETC:WCDM:MEAS:MEV:STAT?')
            self.command_cmw100_query('*OPC?')
        spectrum_results = self.command_cmw100_query(
            f'FETC:WCDM:MEAS:MEV:SPEC:AVER?')  # P1: Power, P2: ACLR_-2, P3: ACLR_-1, P4:ACLR_+1, P5:ACLR_+2, P6:OBW
        spectrum_results = spectrum_results.split(',')
        spectrum_results = [
            round(eval(spectrum_results[1]), 2),
            round(eval(spectrum_results[3]) - eval(spectrum_results[1]), 2),
            round(eval(spectrum_results[4]) - eval(spectrum_results[1]), 2),
            round(eval(spectrum_results[2]) - eval(spectrum_results[1]), 2),
            round(eval(spectrum_results[5]) - eval(spectrum_results[1]), 2),
            round(eval(spectrum_results[6]) / 1000000, 2)
        ]
        logger.info(
            f'Power: {spectrum_results[0]:.2f}, ACLR_-1: {spectrum_results[2]:.2f}, ACLR_1: {spectrum_results[3]:.2f}, ACLR_-2: {spectrum_results[1]:.2f}, ACLR_+2: {spectrum_results[4]:.2f}, OBW: {spectrum_results[5]:.2f}MHz')
        self.command_cmw100_write(f'STOP:WCDM:MEAS:MEV')
        self.command_cmw100_query('*OPC?')

        return spectrum_results + mod_results

    def tx_measure_lte(self):
        logger.info('---------Tx Measure----------')
        mode = 'TDD' if self.band_lte in [38, 39, 40, 41, 42, 48] else 'FDD'
        self.command_cmw100_write(f'CONF:LTE:MEAS:DMODe {mode}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:BAND OB{self.band_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:FREQ {self.tx_freq_lte}KHz')
        self.command_cmw100_query('*OPC?')
        rb = f'0{self.bw_lte * 10}' if self.bw_lte < 10 else f'{self.bw_lte * 10}'
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CBAN B{rb}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOD:MSCH {self.mcs_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:NRB {self.rb_size_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:ORB {self.rb_start_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CPR NORM')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:DSSP 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO OFF')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOEX ON')
        lim1 = -10 if self.bw_lte == 1.4 else -13 if self.bw_lte == 3 else -15 if self.bw_lte == 5 else -18 if self.bw_lte == 10 else -20 if self.bw_lte == 15 else -21
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN{self.bw_lte * 10} ON,0MHz,1MHz,{lim1},K030')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN{self.bw_lte * 10} ON,1MHz,2.5MHz,-10,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN{self.bw_lte * 10} ON,2.5MHz,2.8MHz,-25,M1') if self.bw_lte < 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN{self.bw_lte * 10} ON,2.5MHz,2.8MHz,-10,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN{self.bw_lte * 10} ON,2.8MHz,5MHz,-10,M1') if self.bw_lte >= 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN{self.bw_lte * 10} OFF,2.8MHz,5MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN{self.bw_lte * 10} ON,5MHz,6MHz,-13,M1') if self.bw_lte > 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN{self.bw_lte * 10} OFF,5MHz,6MHz,-25,M1') if self.bw_lte < 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN{self.bw_lte * 10} ON,5MHz,6MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN{self.bw_lte * 10} ON,6MHz,10MHz,-13,M1') if self.bw_lte > 5 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN{self.bw_lte * 10} OFF,6MHz,10MHz,-25,M1') if self.bw_lte < 5 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN{self.bw_lte * 10} ON,6MHz,10MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN{self.bw_lte * 10} ON,10MHz,15MHz,-13,M1') if self.bw_lte > 10 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN{self.bw_lte * 10} OFF,10MHz,15MHz,-25,M1') if self.bw_lte < 10 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN{self.bw_lte * 10} ON,10MHz,15MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN{self.bw_lte * 10} ON,15MHz,20MHz,-13,M1') if self.bw_lte > 15 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN{self.bw_lte * 10} OFF,15MHz,20MHz,-25,M1') if self.bw_lte < 15 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN{self.bw_lte * 10} ON,15MHz,20MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN{self.bw_lte * 10} ON,20MHz,25MHz,-25,M1') if self.bw_lte == 20 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN{self.bw_lte * 10} OFF,20MHz,25MHz,-25,M1')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONFigure:LTE:MEAS:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:ENP {self.tx_level + 5}.00')
        self.command_cmw100_write(f'ROUT:LTE:MEAS:SCEN:SAL R11, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO ON')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f"TRIG:LTE:MEAS:MEV:SOUR 'GPRF Gen1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:LTE:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MSUB 2, 10, 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:SCEN:ACT SAL')
        self.command_cmw100_query('SYST:ERR:ALL?')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'ROUT:LTE:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_query('*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query('*OPC?')
        time.sleep(0.2)
        mod_results = self.command_cmw100_query(
            'READ:LTE:MEAS:MEV:MOD:AVER?')  # P3 is EVM, P15 is Ferr, P14 is IQ Offset
        mod_results = mod_results.split(',')
        mod_results = [mod_results[3], mod_results[15], mod_results[14]]
        mod_results = [eval(m) for m in mod_results]
        logger.info(f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
        self.command_cmw100_write(f'INIT:LTE:MEAS:MEV')
        self.command_cmw100_query('*OPC?')
        f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
        while f_state != 'RDY':
            f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
            self.command_cmw100_query('*OPC?')
        aclr_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:ACLR:AVER?')
        aclr_results = aclr_results.split(',')[1:]
        aclr_results = [eval(aclr) * -1 if eval(aclr) > 30 else eval(aclr) for aclr in
                        aclr_results]  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2
        logger.info(
            f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
        iem_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:IEM:MARG?')
        iem_results = iem_results.split(',')
        logger.info(f'InBandEmissions Margin: {eval(iem_results[2]):.2f}dB')
        # logger.info(f'IEM_MARG results: {iem_results}')
        esfl_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:ESFL:EXTR?')
        esfl_results = esfl_results.split(',')
        ripple1 = round(eval(esfl_results[2]), 2) if esfl_results[2] != 'NCAP' else esfl_results[2]
        ripple2 = round(eval(esfl_results[3]), 2) if esfl_results[3] != 'NCAP' else esfl_results[3]
        logger.info(f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
        time.sleep(0.2)
        # logger.info(f'ESFL results: {esfl_results}')
        sem_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:MARG?')
        logger.info(f'SEM_MARG results: {sem_results}')
        sem_avg_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:AVER?')
        sem_avg_results = sem_avg_results.split(',')
        logger.info(
            f'OBW: {eval(sem_avg_results[2]) / 1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
        # logger.info(f'SEM_AVER results: {sem_avg_results}')
        self.command_cmw100_write(f'STOP:LTE:MEAS:MEV')
        self.command_cmw100_query('*OPC?')

        logger.debug(aclr_results + mod_results)
        return aclr_results + mod_results  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET

    def tx_measure_fr1(self):
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78,
                                     79] else 0  # for now FDD is forced to 15KHz and TDD is to be 30KHz
        scs = 15 * (2 ** scs)  # for now TDD only use 30KHz, FDD only use 15KHz
        logger.info('---------Tx Measure----------')
        mode = "TDD" if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else "FDD"
        self.command_cmw100_query(f'SYSTem:BASE:OPTion:VERSion?  "CMW_NRSub6G_Meas"')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:DMODe {mode}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:BAND OB{self.band_fr1}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:FREQ {self.tx_freq_fr1}KHz')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:MOEX ON')
        bw = f'00{self.bw_fr1}' if self.bw_fr1 < 10 else f'0{self.bw_fr1}' if 10 <= self.bw_fr1 < 100 else self.bw_fr1
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:BWC S{scs}K, B{bw}')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN{self.bw_fr1} ON,0.015MHz,0.0985MHz,{round(-13.5 - 10 * math.log10(self.bw_fr1 / 5), 1)},K030')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN{self.bw_fr1} ON,1.5MHz,4.5MHz,-8.5,M1')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN{self.bw_fr1} ON,5.5MHz,{round(-0.5 + self.bw_fr1, 1)}MHz,-11.5,M1')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN{self.bw_fr1} ON,{round(0.5 + self.bw_fr1, 1)}MHz,{round(4.5 + self.bw_fr1, 1)}MHz,-23.5,M1')
        _256Q_flag = 2 if self.mcs_fr1 == 'Q256' else 0
        self.command_cmw100_write(
            f'CONFigure:NRSub:MEASurement:MEValuation:PUSChconfig {self.mcs_fr1},A,OFF,{self.rb_size_fr1},{self.rb_start_fr1},14,0,T1,SING,{_256Q_flag},2')
        type_ = 'ON' if self.type_fr1 == 'DFTS' else 'OFF'  # DFTS: ON, CP: OFF
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:DFTPrecoding {type_}')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PCOMp OFF, 6000E+6')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:REPetition SING')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PLCid 0')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:CTYPe PUSC')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PER MS25')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PATT S{scs}K, 3,0,1,14 ')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:ENP {self.tx_level + 5}.00')
        self.command_cmw100_write(f'ROUT:NRS:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_write(f"TRIG:NRS:MEAS:MEV:SOUR 'GPRF GEN1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:NRS:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:NSUB 10')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:SCEN:ACT SAL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'ROUT:NRS:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'INIT:NRS:MEAS:MEV')
        self.command_cmw100_query(f'*OPC?')
        f_state = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:STAT?')
        while f_state != 'RDY':
            f_state = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:STAT?')
            self.command_cmw100_query('*OPC?')
        mod_results = self.command_cmw100_query(
            'FETC:NRS:MEAS:MEV:MOD:AVER?')  # P3 is EVM, P15 is Ferr, P14 is IQ Offset
        mod_results = mod_results.split(',')
        mod_results = [mod_results[3], mod_results[15], mod_results[14]]
        mod_results = [eval(m) for m in mod_results]
        logger.info(f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
        aclr_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:ACLR:AVER?')
        aclr_results = aclr_results.split(',')[1:]
        aclr_results = [eval(aclr) * -1 if eval(aclr) > 30 else eval(aclr) for aclr in
                        aclr_results]  # UTRA2(-), UTRA1(-), NR(-), TxP, NR(+), UTRA1(+), UTRA2(+)
        logger.info(
            f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
        iem_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:IEM:MARG:AVER?')
        iem_results = iem_results.split(',')
        iem = f'{eval(iem_results[2]):.2f}' if iem_results[2] != 'INV' else 'INV'
        logger.info(f'InBandEmissions Margin: {iem}dB')
        # logger.info(f'IEM_MARG results: {iem_results}')
        esfl_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:ESFL:EXTR?')
        esfl_results = esfl_results.split(',')
        ripple1 = round(eval(esfl_results[2]), 2) if esfl_results[2] != 'NCAP' else esfl_results[2]
        ripple2 = round(eval(esfl_results[3]), 2) if esfl_results[3] != 'NCAP' else esfl_results[3]
        logger.info(f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
        time.sleep(0.2)
        # logger.info(f'ESFL results: {esfl_results}')
        sem_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:MARG:ALL?')
        logger.info(f'SEM_MARG results: {sem_results}')
        sem_avg_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:AVERage?')
        sem_avg_results = sem_avg_results.split(',')
        logger.info(
            f'OBW: {eval(sem_avg_results[2]) / 1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
        # logger.info(f'SEM_AVER results: {sem_avg_results}')
        self.command_cmw100_write(f'STOP:NRS:MEAS:MEV')
        self.command_cmw100_query('*OPC?')

        logger.debug(aclr_results + mod_results)
        return aclr_results + mod_results  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET

    def set_rx_level(self):
        logger.info(f'==========Search: {self.rx_level} dBm==========')
        self.command_cmw100_write(f'SOUR:GPRF:GEN1:RFS:LEV {self.rx_level}')
        # self.command_cmw100_query('*OPC?')

    def query_rsrp_cinr_wcdma(self):
        res = self.command(f'AT+LRXMEAS={self.rx_path_lte},20')
        for line in res:
            if '+LRXMEAS:' in line.decode():
                self.rsrp_list = line.decode().split(':')[1].strip().split(',')[:4]
                self.cinr_list = line.decode().split(':')[1].strip().split(',')[4:]
                self.rsrp_list = [eval(rsrp) / 100 for rsrp in self.rsrp_list]
                self.cinr_list = [eval(cinr) / 100 for cinr in self.cinr_list]
                logger.info(f'**** RSRP: {self.rsrp_list} ****')
                logger.info(f'**** CINR: {self.cinr_list} ****')

    def query_rsrp_cinr_lte(self):
        res = self.command(f'AT+LRXMEAS={self.rx_path_lte},20')
        for line in res:
            if '+LRXMEAS:' in line.decode():
                self.rsrp_list = line.decode().split(':')[1].strip().split(',')[:4]
                self.cinr_list = line.decode().split(':')[1].strip().split(',')[4:]
                self.rsrp_list = [eval(rsrp) / 100 for rsrp in self.rsrp_list]
                self.cinr_list = [eval(cinr) / 100 for cinr in self.cinr_list]
                logger.info(f'**** RSRP: {self.rsrp_list} ****')
                logger.info(f'**** CINR: {self.cinr_list} ****')

    def query_rsrp_cinr_fr1(self):
        res = self.command(f'AT+NRXMEAS={self.rx_path_fr1},20')
        for line in res:
            if '+NRXMEAS:' in line.decode():
                self.rsrp_list = line.decode().split(':')[1].strip().split(',')[:4]
                self.cinr_list = line.decode().split(':')[1].strip().split(',')[4:]
                self.rsrp_list = [eval(rsrp) / 100 for rsrp in self.rsrp_list]
                self.cinr_list = [eval(cinr) / 100 for cinr in self.cinr_list]
                logger.info(f'**** RSRP: {self.rsrp_list} ****')
                logger.info(f'**** CINR: {self.cinr_list} ****')

    def query_agc_wcdma(self):
        res = self.command(f'AT+LRX1RX2AGCIDXRD')
        for line in res:
            if '+LRX1RX2AGCIDXRD:' in line.decode():
                self.agc_list = line.decode().split(':')[1].strip().split(',')
                self.agc_list = [eval(agc) for agc in self.agc_list]
                logger.info(f'**** AGC: {self.agc_list} ****')

    def query_agc_lte(self):
        res = self.command(f'AT+LRX1RX2AGCIDXRD')
        for line in res:
            if '+LRX1RX2AGCIDXRD:' in line.decode():
                self.agc_list = line.decode().split(':')[1].strip().split(',')
                self.agc_list = [eval(agc) for agc in self.agc_list]
                logger.info(f'**** AGC: {self.agc_list} ****')

    def query_agc_fr1(self):
        res = self.command(f'AT+NAGCIDXRD')
        for line in res:
            if '+NRX1RX2AGCIDXRD:' in line.decode():
                self.agc_list = line.decode().split(':')[1].strip().split(',')
                self.agc_list = [eval(agc) for agc in self.agc_list]
                logger.info(f'**** AGC: {self.agc_list} ****')

    def get_esens_wcdma(self):
        self.esens_list = [round(self.rx_level - c - 1, 2) for c in self.cinr_list]
        logger.info(f'**** ESENS: {self.esens_list} ****')

    def get_esens_lte(self):
        self.esens_list = [round(self.rx_level - c - 1, 2) for c in self.cinr_list]
        logger.info(f'**** ESENS: {self.esens_list} ****')

    def get_esens_fr1(self):
        self.esens_list = [round(self.rx_level - c - 1, 2) for c in self.cinr_list]
        logger.info(f'**** ESENS: {self.esens_list} ****')

    def query_rx_measure_wcdma(self):
        self.query_rsrp_cinr_wcdma()
        self.query_agc_wcdma()
        self.get_esens_wcdma()

    def query_rx_measure_lte(self):
        self.query_rsrp_cinr_lte()
        self.query_agc_lte()
        self.get_esens_lte()

    def query_rx_measure_fr1(self):
        self.query_rsrp_cinr_fr1()
        self.query_agc_fr1()
        self.get_esens_fr1()

    def query_fer_measure_gsm(self):
        logger.info('========== FER measure ==========')
        self.sig_gen_gsm()
        self.sync_gsm()
        res = self.command(
            f'AT+TESTBER={self.band_tx_set_dict_gsm[self.band_gsm]},{self.mod_dict_gsm[self.mod_gsm]},0,1,{self.rx_chan_gsm},{-1 * int(round(self.rx_level, 0))},7,2',
            delay=0.5)
        for line in res:
            if '+TESTBER: ' in line.decode():
                results = eval(line.decode().split(': ')[1])
                self.rssi, self.fer = [round(r / 100, 2) for r in results]
                logger.info(f'****RSSI: {self.rssi} ****')
                logger.info(f'****FER: {self.fer} %****')

    def query_fer_measure_wcdma(self):
        logger.info('========== FER measure ==========')
        res = self.command('AT+HGETSENSE=100', delay=2)
        for line in res:
            if '+GETSENSE:' in line.decode():
                self.fer = eval(line.decode().split(':')[1])
                logger.info(f'****FER: {self.fer / 1000} %****')

    def query_fer_measure_lte(self):
        logger.info('========== FER measure ==========')
        res = self.command('AT+LFERMEASURE=500', delay=0.5)
        for line in res:
            if '+LFERMEASURE:' in line.decode():
                self.fer = eval(line.decode().split(':')[1])
                logger.info(f'****FER: {self.fer / 100} %****')

    def query_fer_measure_fr1(self):
        logger.info('========== FER measure ==========')
        res = self.command('AT+NFERMEASURE=500', delay=0.5)
        for line in res:
            if '+NFERMEASURE:' in line.decode():
                self.fer = eval(line.decode().split(':')[1])
                logger.info(f'****FER: {self.fer / 100} %****')

    def search_window_gsm(self):
        rssi = None
        self.query_fer_measure_gsm()
        while self.fer < 2:
            rssi = self.rssi
            self.rx_level = round(self.rx_level - self.window, 1)  # to reduce a unit
            # self.set_rx_level()
            self.query_fer_measure_gsm()
        return rssi
        # self.command_cmw100_query('*OPC?')

    def search_window_wcdma(self):
        self.query_fer_measure_wcdma()
        while self.fer < 100:
            self.rx_level = round(self.rx_level - self.window, 1)  # to reduce a unit
            self.set_rx_level()
            self.query_fer_measure_wcdma()
            # self.command_cmw100_query('*OPC?')

    def search_window_lte(self):
        self.query_fer_measure_lte()
        while self.fer < 500:
            self.rx_level = round(self.rx_level - self.window, 1)  # to reduce a unit
            self.set_rx_level()
            self.query_fer_measure_lte()
            # self.command_cmw100_query('*OPC?')

    def search_window_fr1(self):
        self.query_fer_measure_fr1()
        while self.fer < 500:
            self.rx_level = round(self.rx_level - self.window, 1)  # to reduce a unit
            self.set_rx_level()
            self.query_fer_measure_fr1()
            # self.command_cmw100_query('*OPC?')

    def search_sensitivity_gsm(self):
        reset_rx_level = -104
        self.rx_level = reset_rx_level
        coarse_1 = 2
        coarse_2 = 1
        fine = 0.2
        logger.info('----------Search RX Level----------')
        self.window = coarse_1
        self.search_window_gsm()  # first time by coarse_1
        logger.info('Second time to search')
        self.rx_level += coarse_1 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        # self.set_rx_level()
        self.window = coarse_2
        self.search_window_gsm()  # second time by coarse_2
        logger.info('Third time to search')
        self.rx_level += coarse_2 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        # self.set_rx_level()
        self.window = fine
        rssi = self.search_window_gsm()  # second time by fine
        self.rx_level = round(self.rx_level + fine, 1)
        self.rssi = rssi
        logger.info(f'Final Rx Level: {self.rx_level}')
        logger.info(f'Final RSSI: {self.rssi}')
        self.command('AT+TESTRESET')

    def search_sensitivity_wcdma(self):
        reset_rx_level = -106
        self.rx_level = reset_rx_level
        coarse_1 = 2
        coarse_2 = 1
        fine = 0.2
        logger.info('----------Search RX Level----------')
        self.window = coarse_1
        self.search_window_wcdma()  # first time by coarse_1
        logger.info('Second time to search')
        self.rx_level += coarse_1 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        self.set_rx_level()
        self.window = coarse_2
        self.search_window_wcdma()  # second time by coarse_2
        logger.info('Third time to search')
        self.rx_level += coarse_2 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        self.set_rx_level()
        self.window = fine
        self.search_window_wcdma()  # second time by fine
        self.rx_level = round(self.rx_level + fine, 1)
        logger.info(f'Final Rx Level: {self.rx_level}')

    def search_sensitivity_lte(self):
        reset_rx_level = -80
        self.rx_level = reset_rx_level
        coarse_1 = 2
        coarse_2 = 1
        fine = 0.2
        logger.info('----------Search RX Level----------')
        self.window = coarse_1
        self.search_window_lte()  # first time by coarse_1
        logger.info('Second time to search')
        self.rx_level += coarse_1 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        self.set_rx_level()
        self.window = coarse_2
        self.search_window_lte()  # second time by coarse_2
        logger.info('Third time to search')
        self.rx_level += coarse_2 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        self.set_rx_level()
        self.window = fine
        self.search_window_lte()  # second time by fine
        self.rx_level = round(self.rx_level + fine, 1)
        logger.info(f'Final Rx Level: {self.rx_level}')

    def search_sensitivity_fr1(self):
        reset_rx_level = -80
        self.rx_level = reset_rx_level
        coarse_1 = 2
        coarse_2 = 1
        fine = 0.2
        logger.info('----------Search RX Level----------')
        self.window = coarse_1
        self.search_window_fr1()  # first time by coarse_1
        logger.info('Second time to search')
        self.rx_level += coarse_1 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        self.set_rx_level()
        self.window = coarse_2
        self.search_window_fr1()  # second time by coarse_2
        logger.info('Third time to search')
        self.rx_level += coarse_2 * 2
        logger.info(f'==========Back to Search: {self.rx_level} dBm==========')
        self.set_rx_level()
        self.window = fine
        self.search_window_fr1()  # second time by fine
        self.rx_level = round(self.rx_level + fine, 1)
        logger.info(f'Final Rx Level: {self.rx_level}')

    def search_sensitivity_pipline_gsm(self):
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.mod_gsm = wt.mod_gsm
        self.script = 'GENERAL'
        items = [
            (tech, band)
            for tech in wt.tech
            for band in wt.gsm_bands
        ]
        for item in items:
            if item[0] == 'GSM' and wt.gsm_bands != []:
                self.tech = item[0]
                self.band_gsm = item[1]
                self.pcl = wt.tx_pcl_lb if self.band_gsm in [850, 900] else wt.tx_pcl_mb
                self.search_sensitivity_lmh_progress_gsm()
        self.rxs_relative_plot(self.filename, mode=1)  # mode=1: LMH mode

    def search_sensitivity_pipline_wcdma(self):
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.mcs_wcdma = 'QPSK'
        self.script = 'GENERAL'
        items = [
            (tech, tx_path, band)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for band in wt.wcdma_bands
        ]
        for item in items:
            if item[0] == 'WCDMA' and wt.wcdma_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                # for ue_power_bool in wt.tx_max_pwr_sensitivity:
                # self.tx_level = wt.tx_level if ue_power_bool == 1 else -10
                self.band_wcdma = item[2]
                self.search_sensitivity_lmh_progress_wcdma()
                # self.rx_desense_process()
        self.rxs_relative_plot(self.filename, mode=1)  # mode=1: LMH mode

    def search_sensitivity_pipline_lte(self):
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.mcs_lte = 'QPSK'
        self.script = 'GENERAL'
        items = [
            (tech, tx_path, bw, ue_power_bool, band)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.lte_bandwidths
            for ue_power_bool in wt.tx_max_pwr_sensitivity
            for band in wt.lte_bands
        ]
        for item in items:
            if item[0] == 'LTE' and wt.lte_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_lte = item[2]
                self.ue_power_bool = item[3]
                self.tx_level = wt.tx_level if self.ue_power_bool == 1 else -10
                self.band_lte = item[4]
                if self.bw_lte in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                    self.search_sensitivity_lmh_progress_lte()
                else:
                    logger.info(f'B{self.band_lte} does not have BW {self.bw_lte}MHZ')
        for bw in wt.lte_bandwidths:
            try:
                self.bw_lte = bw
                self.filename = f'Sensitivty_{self.bw_lte}MHZ_{self.tech}_LMH.xlsx'
                self.rx_desense_progress()
                self.rxs_relative_plot(self.filename, mode=1)  # mode=1: LMH mode
            except TypeError as err:
                logger.debug(err)
                logger.info(
                    'It might not have the Bw in this Band, so it cannot to be calculated for desens')
            except KeyError as err:
                logger.debug(err)
                logger.info(
                    f"{self.band_lte} doesn't have this {self.bw_lte}, so desens progress cannot run")
            except FileNotFoundError as err:
                logger.debug(err)
                logger.info(f"There is not file to plot BW{bw}")

    def search_sensitivity_pipline_fr1(self):
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.type_fr1 = 'DFTS'
        self.sa_nsa_mode = wt.sa_nsa
        self.script = 'GENERAL'
        self.mcs_fr1 = 'QPSK'
        items = [
            (tech, tx_path, bw, ue_power_bool, band)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.fr1_bandwidths
            for ue_power_bool in wt.tx_max_pwr_sensitivity
            for band in wt.fr1_bands
        ]
        for item in items:
            if item[0] == 'FR1' and wt.fr1_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_fr1 = item[2]
                self.ue_power_bool = item[3]
                self.tx_level = wt.tx_level if self.ue_power_bool == 1 else -10
                self.band_fr1 = item[4]
                if self.bw_fr1 in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                    self.search_sensitivity_lmh_progress_fr1()
                else:
                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')
        for bw in wt.fr1_bandwidths:
            try:
                self.bw_fr1 = bw
                self.filename = f'Sensitivty_{self.bw_fr1}MHZ_{self.tech}_LMH.xlsx'
                self.rx_desense_progress()
                self.rxs_relative_plot(self.filename, mode=1)  # mode=1: LMH mode
            except TypeError as err:
                logger.debug(err)
                logger.info(
                    'It might not have the Bw in this Band, so it cannot to be calculated for desens')
            except KeyError as err:
                logger.debug(err)
                logger.info(
                    f"{self.band_fr1} doesn't have this {self.bw_fr1}, so desens progress cannot run")
            except FileNotFoundError as err:
                logger.debug(err)
                logger.info(f"There is not file to plot BW{bw}")

    def sensitivity_pipline_endc(self):
        self.tx_level_endc_lte = wt.tx_level_endc_lte
        self.tx_level_endc_fr1 = wt.tx_level_endc_fr1
        self.port_tx_lte = wt.port_tx_lte
        self.port_tx_fr1 = wt.port_tx_fr1
        self.sa_nsa_mode = wt.sa_nsa
        self.type_fr1 = 'DFTS'
        self.mcs_lte = self.mcs_fr1 = 'QPSK'
        self.tx_path = 'TX1'
        self.rx_path_lte = 15
        self.rx_path_fr1 = 15
        items = [
            (tech, band_combo, bw_lte, bw_fr1, chan_rb, ue_power_bool)
            for tech in wt.tech
            for band_combo in wt.endc_bands
            for bw_lte in scripts.ENDC[band_combo]
            for bw_fr1 in scripts.ENDC[band_combo][bw_lte]
            for chan_rb in scripts.ENDC[band_combo][bw_lte][bw_fr1]
            for ue_power_bool in wt.tx_max_pwr_sensitivity
        ]
        data = []
        for item in items:
            self.tech = item[0]
            self.band_combo = item[1]
            self.bw_lte = item[2]
            self.bw_fr1 = item[3]
            self.chan_rb = item[4]
            self.ue_power_bool = item[5]
            [self.band_lte, self.band_fr1] = self.band_combo.split('_')
            (self.tx_freq_lte, self.tx_freq_fr1) = self.chan_rb[0]
            (self.rb_size_lte, self.rb_start_lte) = self.chan_rb[1]
            (self.rb_size_fr1, self.rb_start_fr1) = self.chan_rb[2]
            self.tx_freq_lte = int(self.tx_freq_lte * 1000)
            self.tx_freq_fr1 = int(self.tx_freq_fr1 * 1000)
            loss_tx_lte = get_loss(self.tx_freq_lte)
            loss_tx_fr1 = get_loss(self.tx_freq_fr1)
            self.rx_freq_fr1 = cm_pmt_ftm.transfer_freq_tx2rx_fr1(self.band_fr1, self.tx_freq_fr1)
            self.rx_freq_lte = cm_pmt_ftm.transfer_freq_tx2rx_lte(self.band_lte, self.tx_freq_lte)
            self.loss_rx = get_loss(self.rx_freq_fr1)
            self.preset_instrument()
            self.set_test_end_lte(delay=0.5)
            self.set_test_end_fr1(delay=0.5)
            self.set_test_mode_lte()
            self.rx_level = -70
            self.sig_gen_lte()
            self.sync_lte()
            self.set_test_mode_fr1()
            self.sig_gen_fr1()
            self.sync_fr1()
            # set LTE power
            self.tx_level = wt.tx_level_endc_lte if self.ue_power_bool == 1 else -10
            self.loss_tx = loss_tx_lte
            self.port_tx = self.port_tx_lte
            self.tx_set_lte()
            self.rx_path_setting_lte()
            # set FR1 power
            self.tx_level = self.tx_level_endc_fr1
            self.loss_tx = loss_tx_fr1
            self.port_tx = self.port_tx_fr1
            self.tx_set_fr1()
            self.rx_path_setting_fr1()
            aclr_mod_results_fr1 = self.tx_measure_fr1()
            logger.debug(aclr_mod_results_fr1)
            logger.info(f'FR1 Power: {aclr_mod_results_fr1[3]}')
            self.power_endc_fr1 = round(aclr_mod_results_fr1[3], 2)
            self.search_sensitivity_fr1()
            # set LTE power and get ENDC power for LTE
            self.tx_level = wt.tx_level_endc_lte if self.ue_power_bool == 1 else -10
            self.loss_tx = loss_tx_lte
            self.port_tx = self.port_tx_lte
            self.power_monitor_endc_lte = self.tx_monitor_lte()
            data.append([int(self.band_lte), int(self.band_fr1), self.power_monitor_endc_lte,
                         self.power_endc_fr1, self.rx_level, self.bw_lte, self.bw_fr1,
                         self.tx_freq_lte,
                         self.tx_freq_fr1, self.tx_level, self.tx_level_endc_fr1,
                         self.rb_size_lte,
                         self.rb_start_lte, self.rb_size_fr1, self.rb_start_fr1])
            self.set_test_end_fr1(delay=0.5)
            self.set_test_end_lte(delay=0.5)
        self.endc_relative_power_senstivity_export_excel(data)
        self.rx_desense_endc_progress()
        self.rxs_endc_plot('Sensitivty_ENDC.xlsx')

        # for tech in wt.tech:
        #     if tech == 'FR1' and wt.endc_bands != []:
        #         self.tech = 'FR1'
        #         for band_combo in wt.endc_bands:
        #             data = []
        #             [self.band_lte, self.band_fr1] = band_combo.split('_')
        #             for bw_lte in test_scripts.ENDC[band_combo]:
        #                 self.bw_lte = bw_lte
        #                 for bw_fr1 in test_scripts.ENDC[band_combo][bw_lte]:
        #                     self.bw_fr1 = bw_fr1
        #                     for chan_rb in test_scripts.ENDC[band_combo][bw_lte][bw_fr1]:
        #                         (self.tx_freq_lte, self.tx_freq_fr1) = chan_rb[0]
        #                         (self.rb_size_lte, self.rb_start_lte) = chan_rb[1]
        #                         (self.rb_size_fr1, self.rb_start_fr1) = chan_rb[2]
        #                         self.tx_freq_lte = int(self.tx_freq_lte * 1000)
        #                         self.tx_freq_fr1 = int(self.tx_freq_fr1 * 1000)
        #                         loss_tx_lte = get_loss(self.tx_freq_lte)
        #                         loss_tx_fr1 = get_loss(self.tx_freq_fr1)
        #                         self.rx_freq_fr1 = cm_pmt_ftm.transfer_freq_tx2rx_fr1(self.band_fr1, self.tx_freq_fr1)
        #                         self.rx_freq_lte = cm_pmt_ftm.transfer_freq_tx2rx_lte(self.band_lte, self.tx_freq_lte)
        #                         self.loss_rx = get_loss(self.rx_freq_fr1)
        #                         self.preset_instrument()
        #                         self.set_test_end_lte(delay=0.5)
        #                         self.set_test_end_fr1(delay=0.5)
        #                         self.set_test_mode_lte()
        #                         self.sig_gen_lte()
        #                         self.sync_lte()
        #                         self.set_test_mode_fr1()
        #                         self.sig_gen_fr1()
        #                         self.sync_fr1()
        #                         for ue_power_bool in wt.tx_max_pwr_sensitivity:
        #                             # set LTE power
        #                             self.tx_level = wt.tx_level_endc_lte if ue_power_bool == 1 else -10
        #                             self.loss_tx = loss_tx_lte
        #                             self.port_tx = self.port_tx_lte
        #                             self.tx_set_lte()
        #                             self.rx_path_setting_lte()
        #                             # set FR1 power
        #                             self.tx_level = self.tx_level_endc_fr1
        #                             self.loss_tx = loss_tx_fr1
        #                             self.port_tx = self.port_tx_fr1
        #                             self.tx_set_fr1()
        #                             self.rx_path_setting_fr1()
        #                             aclr_mod_results_fr1 = self.tx_measure_fr1()
        #                             logger.debug(aclr_mod_results_fr1)
        #                             logger.info(f'FR1 Power: {aclr_mod_results_fr1[3]}')
        #                             self.power_endc_fr1 = round(aclr_mod_results_fr1[3], 2)
        #                             self.search_sensitivity_fr1()
        #                             # set LTE power and get ENDC power for LTE
        #                             self.tx_level = wt.tx_level_endc_lte if ue_power_bool == 1 else -10
        #                             self.loss_tx = loss_tx_lte
        #                             self.port_tx = self.port_tx_lte
        #                             self.power_monitor_endc_lte = self.tx_monitor_lte()
        #                             data.append([int(self.band_lte), int(self.band_fr1), self.power_monitor_endc_lte,
        #                                          self.power_endc_fr1, self.rx_level, self.bw_lte, self.bw_fr1,
        #                                          self.tx_freq_lte,
        #                                          self.tx_freq_fr1, self.tx_level, self.tx_level_endc_fr1,
        #                                          self.rb_size_lte,
        #                                          self.rb_start_lte, self.rb_size_fr1, self.rb_start_fr1])
        #             self.set_test_end_fr1(delay=0.5)
        #             self.set_test_end_lte(delay=0.5)
        #             self.endc_relative_power_senstivity_export_excel(data)
        #         self.rx_desense_endc_process()
        # self.rxs_endc_plot('Sensitivty_ENDC.xlsx')

    def search_sensitivity_pipline_fast_lte(
            self):  # this is for that RSRP and CINR without issue because this is calculated method
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.mcs_lte = 'QPSK'
        for tech in wt.tech:
            if tech == 'LTE' and wt.lte_bands != []:
                self.tech = 'LTE'
                for tx_path in wt.tx_paths:
                    self.tx_path = tx_path
                    for bw in wt.lte_bandwidths:
                        self.bw_lte = bw
                        try:
                            for band in wt.lte_bands:
                                self.band_lte = band
                                if bw in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                                    self.search_sensitivity_lmh_fast_progress_lte()
                                else:
                                    logger.info(f'B{self.band_lte} does not have BW {self.bw_lte}MHZ')
                            # self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
                        except TypeError as err:
                            logger.debug(err)
                            logger.info(f'there is no data to plot because the band does not have this BW ')

    def search_sensitivity_lmh_process_gsm(self):
        rx_chan_list = cm_pmt_ftm.dl_chan_select_gsm(self.band_gsm)

        rx_chan_select_list = []
        for chan in self.chan:
            if chan == 'L':
                rx_chan_select_list.append(rx_chan_list[0])
            elif chan == 'M':
                rx_chan_select_list.append(rx_chan_list[1])
            elif chan == 'H':
                rx_chan_select_list.append(rx_chan_list[2])

        self.preset_instrument()
        self.set_test_mode_gsm()
        self.set_test_end_gsm()

        for rx_path in wt.rx_paths:
            self.rx_path_gsm = rx_path
            if self.rx_path_gsm in [1, 2]:
                data = {}
                for rx_chan_gsm in rx_chan_select_list:
                    self.rx_level = -70
                    logger.info('----------Test LMH progress---------')
                    self.rx_chan_gsm = rx_chan_gsm
                    self.rx_freq_gsm = cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, self.rx_chan_gsm, 'rx')
                    self.tx_freq_gsm = cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, self.rx_chan_gsm, 'tx')
                    self.loss_rx = get_loss(self.rx_freq_gsm)
                    self.loss_tx = get_loss(self.tx_freq_gsm)
                    self.set_test_mode_gsm()
                    self.rx_path_setting_gsm()
                    self.sig_gen_gsm()
                    self.sync_gsm()
                    # self.antenna_switch_v2()
                    self.search_sensitivity_gsm()
                    data[self.rx_chan_gsm] = [self.rx_level, self.rssi]
                    # logger.info(f'Power: {aclr_mod_results[3]:.1f}, Sensitivity: {self.rx_level}')
                    # data[self.tx_freq_lte] = [aclr_mod_results[3], self.rx_level, self.rsrp_list, self.cinr_list,
                    #                           self.agc_list]  # measured_power, measured_rx_level, rsrp_list, cinr_list, agc_list
                    self.set_test_end_gsm()
                self.filename = self.rx_power_relative_test_export_excel(data, self.band_gsm, 0, -30,
                                                                         mode=1)  # mode=1: LMH mode
            else:
                logger.info(f"GSM doesn't have this RX path {self.rx_path_gsm_dict[self.rx_path_gsm]}")

    def search_sensitivity_lmh_progress_wcdma(self):
        rx_chan_list = cm_pmt_ftm.dl_chan_select_wcdma(self.band_wcdma)
        tx_chan_list = [cm_pmt_ftm.transfer_chan_rx2tx_wcdma(self.band_wcdma, rx_chan) for rx_chan in rx_chan_list]
        tx_rx_chan_list = list(zip(tx_chan_list, rx_chan_list))  # [(tx_chan, rx_chan),...]

        tx_rx_chan_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_rx_chan_select_list.append(tx_rx_chan_list[0])
            elif chan == 'M':
                tx_rx_chan_select_list.append(tx_rx_chan_list[1])
            elif chan == 'H':
                tx_rx_chan_select_list.append(tx_rx_chan_list[2])

        self.preset_instrument()
        for rx_path in wt.rx_paths:
            self.rx_path_wcdma = rx_path
            if self.rx_path_wcdma in [1, 2, 3, 15]:
                data = {}
                for tx_rx_chan_wcdma in tx_rx_chan_select_list:
                    self.rx_level = -70
                    logger.info('----------Test LMH progress---------')
                    self.rx_chan_wcdma = tx_rx_chan_wcdma[1]
                    self.tx_chan_wcdma = tx_rx_chan_wcdma[0]
                    self.rx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.rx_chan_wcdma, 'rx')
                    self.tx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.tx_chan_wcdma, 'tx')
                    self.loss_rx = get_loss(self.rx_freq_wcdma)
                    self.loss_tx = get_loss(self.tx_freq_wcdma)
                    self.set_test_end_wcdma()
                    self.set_test_mode_wcdma()
                    self.command_cmw100_query('*OPC?')
                    self.rx_path_setting_wcdma()
                    self.sig_gen_wcdma()
                    self.sync_wcdma()
                    # self.antenna_switch_v2()
                    # self.tx_chan_wcdma = tx_rx_chan_wcdma[0]
                    # self.tx_set_wcdma()
                    # aclr_mod_results = self.tx_measure_wcdma()
                    # logger.debug(aclr_mod_results)
                    # data[self.tx_chan_wcdma] = aclr_mod_results
                    self.search_sensitivity_wcdma()
                    data[self.tx_chan_wcdma] = self.rx_level
                    # self.query_rx_measure_wcdma()
                    # logger.info(f'Power: {aclr_mod_results[3]:.1f}, Sensitivity: {self.rx_level}')
                    # data[self.tx_freq_lte] = [aclr_mod_results[3], self.rx_level, self.rsrp_list, self.cinr_list,
                    #                           self.agc_list]  # measured_power, measured_rx_level, rsrp_list, cinr_list, agc_list
                    self.set_test_end_wcdma()
                self.filename = self.rx_power_relative_test_export_excel(data, self.band_wcdma, 5, -30,
                                                                         mode=1)  # mode=1: LMH mode
            else:
                logger.info(f"WCDMA doesn't have this RX path {self.rx_path_wcdma_dict[self.rx_path_wcdma]}")
                continue

    def search_sensitivity_lmh_progress_lte(self):
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte,
                                                   self.bw_lte)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        rx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                rx_freq_select_list.append(rx_freq_list[0])
            elif chan == 'M':
                rx_freq_select_list.append(rx_freq_list[1])
            elif chan == 'H':
                rx_freq_select_list.append(rx_freq_list[2])
        for rx_path in wt.rx_paths:
            self.rx_path_lte = rx_path
            data = {}
            for rx_freq in rx_freq_select_list:
                self.rx_level = -70
                self.rx_freq_lte = rx_freq
                self.tx_freq_lte = cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, self.rx_freq_lte)
                self.loss_rx = get_loss(self.rx_freq_lte)
                self.loss_tx = get_loss(self.tx_freq_lte)
                logger.info('----------Test LMH progress---------')
                self.preset_instrument()
                self.set_test_end_lte()
                self.set_test_mode_lte()
                # self.command_cmw100_query('*OPC?')
                self.sig_gen_lte()
                self.sync_lte()
                self.rx_path_setting_lte()
                self.rb_size_lte, self.rb_start_lte = cm_pmt_ftm.special_uplink_config_sensitivity_lte(self.band_lte,
                                                                                                       self.bw_lte)  # for RB set
                self.antenna_switch_v2()
                self.tx_set_lte()
                aclr_mod_results = self.tx_measure_lte()  # aclr_results + mod_results  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET
                # self.command_cmw100_query('*OPC?')
                self.search_sensitivity_lte()
                self.query_rx_measure_lte()
                logger.info(f'Power: {aclr_mod_results[3]:.1f}, Sensitivity: {self.rx_level}')
                data[self.tx_freq_lte] = [aclr_mod_results[3], self.rx_level, self.rsrp_list, self.cinr_list,
                                          self.agc_list]  # measured_power, measured_rx_level, rsrp_list, cinr_list, agc_list
                self.set_test_end_lte()
            self.filename = self.rx_power_relative_test_export_excel(data, self.band_lte, self.bw_lte, self.tx_level,
                                                                     mode=1)  # mode=1: LMH mode

    def search_sensitivity_lmh_process_fr1(self):
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1,
                                                   self.bw_fr1)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        rx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                rx_freq_select_list.append(rx_freq_list[0])
            elif chan == 'M':
                rx_freq_select_list.append(rx_freq_list[1])
            elif chan == 'H':
                rx_freq_select_list.append(rx_freq_list[2])
        for rx_path in wt.rx_paths:
            self.rx_path_fr1 = rx_path
            data = {}
            for rx_freq in rx_freq_select_list:
                self.rx_level = -70
                self.rx_freq_fr1 = rx_freq
                self.tx_freq_fr1 = cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, self.rx_freq_fr1)
                self.loss_rx = get_loss(self.rx_freq_fr1)
                self.loss_tx = get_loss(self.tx_freq_fr1)
                logger.info('----------Test LMH progress---------')
                self.preset_instrument()
                self.set_test_end_fr1()
                self.set_test_mode_fr1()
                # self.command_cmw100_query('*OPC?')
                self.sig_gen_fr1()
                self.sync_fr1()
                self.rx_path_setting_fr1()
                self.rb_size_fr1, self.rb_start_fr1 = cm_pmt_ftm.special_uplink_config_sensitivity_fr1(self.band_fr1,
                                                                                                       self.scs,
                                                                                                       self.bw_fr1)  # for RB set(including special tx setting)
                self.antenna_switch_v2()
                self.tx_set_fr1()
                aclr_mod_results = self.tx_measure_fr1()  # aclr_results + mod_results  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET
                # self.command_cmw100_query('*OPC?')
                self.search_sensitivity_fr1()
                self.query_rx_measure_fr1()
                logger.info(f'Power: {aclr_mod_results[3]:.1f}, Sensitivity: {self.rx_level}')
                data[self.tx_freq_fr1] = [aclr_mod_results[3], self.rx_level, self.rsrp_list, self.cinr_list,
                                          self.agc_list]  # measured_power, measured_rx_level, rsrp_list, cinr_list, agc_list
                self.set_test_end_fr1()
            self.filename = self.rx_power_relative_test_export_excel(data, self.band_fr1, self.bw_fr1, self.tx_level,
                                                                     mode=1)  # mode=1: LMH mode

    def search_sensitivity_lmh_fast_process_lte(self):  # this is not yet used
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte,
                                                   self.bw_lte)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        rx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                rx_freq_select_list.append(rx_freq_list[0])
            elif chan == 'M':
                rx_freq_select_list.append(rx_freq_list[1])
            elif chan == 'H':
                rx_freq_select_list.append(rx_freq_list[2])
        for rx_freq in rx_freq_select_list:
            self.rx_level = -70
            self.rx_freq_lte = rx_freq
            self.tx_freq_lte = cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, self.rx_freq_lte)
            self.loss_rx = get_loss(self.rx_freq_lte)
            self.loss_tx = get_loss(self.tx_freq_lte)
            logger.info('----------Test LMH progress---------')
            self.preset_instrument()
            self.set_test_end_lte()
            self.antenna_switch_v2()
            self.set_test_mode_lte()
            self.command_cmw100_query('*OPC?')
            self.sig_gen_lte()
            self.sync_lte()
            self.rb_size_lte, self.rb_start_lte = cm_pmt_ftm.special_uplink_config_sensitivity_lte(self.band_lte,
                                                                                                   self.bw_lte)  # for RB set
            self.tx_set_lte()
            self.tx_measure_lte()
            aclr_mod_results = self.tx_measure_lte()  # aclr_results + mod_results  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET
            self.rx_path_setting_lte()
            self.query_rx_measure_lte()
            logger.info(f'Power: {aclr_mod_results[3]:.1f}, Sensitivity: {self.esens_list}')
            rsrp_max = max(self.rsrp_list)
            rsrp_max_index = self.rsrp_list.index(rsrp_max)
            rx_level = self.esens_list[rsrp_max_index]
            self.filename = self.rx_power_relative_test_export_excel(data_freq, self.band_lte, self.bw_lte, rx_level,
                                                                     mode=1)  # mode=1: LMH mode

    @staticmethod
    def endc_relative_power_senstivity_export_excel(data):
        """
        :param data:  data = [int(self.band_lte), int(self.band_fr1), self.power_monitor_endc_lte, self.power_endc_fr1, self.rx_level, self.bw_lte, self.bw_fr1, self.tx_freq_lte, self.tx_freq_fr1, self.tx_level_endc_lte, self.tx_level_endc_fr1, self.rb_size_lte, self.rb_start_lte, self.rb_size_fr1, self.rb_start_fr1]
        :return:
        """
        logger.info('----------save to excel----------')

        filename = f'Sensitivty_ENDC.xlsx'

        if pathlib.Path(filename).exists() is False:
            logger.info('----------file does not exist----------')
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            # to create sheet
            wb.create_sheet(f'Raw_Data_ENDC_FR1_TxMax')
            wb.create_sheet(f'Raw_Data_ENDC_FR1_-10dBm')
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                ws['A1'] = 'Band_LTE'
                ws['B1'] = 'Band_FR1'
                ws['C1'] = 'Power_LTE_measured'
                ws['D1'] = 'Power_FR1_measured'
                ws['E1'] = 'Sensitivity_FR1'
                ws['F1'] = 'BW_LTE'
                ws['G1'] = 'BW_FR1'
                ws['H1'] = 'Freq_tx_LTE'
                ws['I1'] = 'Freq_tx_FR1'
                ws['J1'] = 'Tx_level_LTE'
                ws['K1'] = 'Tx_level_FR1'
                ws['L1'] = 'rb_size_LTE'
                ws['M1'] = 'rb_start_LTE'
                ws['N1'] = 'rb_size_FR1'
                ws['O1'] = 'rb_start_FR1'

            wb.create_sheet(f'Desens_ENDC')
            ws = wb[f'Desens_ENDC']
            ws['A1'] = 'Band_LTE'
            ws['B1'] = 'Band_FR1'
            ws['C1'] = 'BW_LTE'
            ws['D1'] = 'BW_FR1'
            ws['E1'] = 'Freq_tx_LTE'
            ws['F1'] = 'Freq_tx_FR1'
            ws['G1'] = 'Diff'

            wb.save(filename)
            wb.close()

        logger.info('----------file exist----------')
        wb = openpyxl.load_workbook(filename)

        for d in data:
            sheetname = f'Raw_Data_ENDC_FR1_TxMax' if d[2] > 0 else f'Raw_Data_ENDC_FR1_-10dBm'
            ws = wb[sheetname]
            max_row = ws.max_row
            max_col = ws.max_column
            row = max_row + 1
            for col in range(max_col):
                ws.cell(row, col + 1).value = d[col]

        wb.save(filename)
        wb.close()

    def rx_power_relative_test_export_excel(self, data, band, bw, tx_freq_level,
                                            mode=0):  # mode general: 0,  mode LMH: 1
        """
        data is dict like:
        tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET]
        """
        logger.info('----------save to excel----------')
        filename = None
        if self.script == 'GENERAL':
            if mode == 1:
                if tx_freq_level >= 100:
                    filename = f'_{bw}MHZ_{self.tech}_LMH.xlsx'
                elif tx_freq_level <= 100:
                    filename = f'Sensitivty_{bw}MHZ_{self.tech}_LMH.xlsx'

                if pathlib.Path(filename).exists() is False:
                    logger.info('----------file does not exist----------')
                    wb = openpyxl.Workbook()
                    wb.remove(wb['Sheet'])
                    # to create sheet
                    if self.tech == 'LTE':
                        wb.create_sheet(f'Raw_Data_{self.mcs_lte}_TxMax')
                        wb.create_sheet(f'Raw_Data_{self.mcs_lte}_-10dBm')
                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'RX_Path'
                            ws['C1'] = 'Chan'
                            ws['D1'] = 'Tx_Freq'
                            ws['E1'] = 'Tx_level'
                            ws['F1'] = 'Power'
                            ws['G1'] = 'Rx_Level'
                            ws['H1'] = 'RSRP_RX0'
                            ws['I1'] = 'RSRP_RX1'
                            ws['J1'] = 'RSRP_RX2'
                            ws['K1'] = 'RSRP_RX3'
                            ws['L1'] = 'CINR_RX0'
                            ws['M1'] = 'CINR_RX1'
                            ws['N1'] = 'CINR_RX2'
                            ws['O1'] = 'CINR_RX3'
                            ws['P1'] = 'AGC_RX0'
                            ws['Q1'] = 'AGC_RX1'
                            ws['R1'] = 'AGC_RX2'
                            ws['S1'] = 'AGC_RX3'
                            ws['T1'] = 'TX_Path'

                        wb.create_sheet(f'Desens_{self.mcs_lte}')
                        ws = wb[f'Desens_{self.mcs_lte}']
                        ws['A1'] = 'Band'
                        ws['B1'] = 'Rx_Path'
                        ws['C1'] = 'Chan'
                        ws['D1'] = 'Diff'

                    elif self.tech == 'FR1':
                        wb.create_sheet(f'Raw_Data_{self.mcs_fr1}_TxMax')
                        wb.create_sheet(f'Raw_Data_{self.mcs_fr1}_-10dBm')
                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'RX_Path'
                            ws['C1'] = 'Chan'
                            ws['D1'] = 'Tx_Freq'
                            ws['E1'] = 'Tx_level'
                            ws['F1'] = 'Power'
                            ws['G1'] = 'Rx_Level'
                            ws['H1'] = 'RSRP_RX0'
                            ws['I1'] = 'RSRP_RX1'
                            ws['J1'] = 'RSRP_RX2'
                            ws['K1'] = 'RSRP_RX3'
                            ws['L1'] = 'CINR_RX0'
                            ws['M1'] = 'CINR_RX1'
                            ws['N1'] = 'CINR_RX2'
                            ws['O1'] = 'CINR_RX3'
                            ws['P1'] = 'AGC_RX0'
                            ws['Q1'] = 'AGC_RX1'
                            ws['R1'] = 'AGC_RX2'
                            ws['S1'] = 'AGC_RX3'
                            ws['T1'] = 'TX_Path'

                        wb.create_sheet(f'Desens_{self.mcs_fr1}')
                        ws = wb[f'Desens_{self.mcs_fr1}']
                        ws['A1'] = 'Band'
                        ws['B1'] = 'Rx_Path'
                        ws['C1'] = 'Chan'
                        ws['D1'] = 'Diff'
                        ws['E1'] = 'TX_Path'

                    elif self.tech == 'WCDMA':
                        wb.create_sheet(f'Raw_Data')
                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'RX_Path'
                            ws['C1'] = 'Chan'
                            ws['D1'] = 'Channel'
                            ws['E1'] = 'Tx_Freq'
                            ws['F1'] = 'Rx_Level'

                    elif self.tech == 'GSM':
                        wb.create_sheet(f'Raw_Data')
                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'RX_Path'
                            ws['C1'] = 'Chan'
                            ws['D1'] = 'Channel'
                            ws['E1'] = 'Tx_Freq'
                            ws['F1'] = 'Rx_Level'
                            ws['G1'] = 'RSSI'

                    wb.save(filename)
                    wb.close()

                logger.info('----------file exist----------')
                wb = openpyxl.load_workbook(filename)
                ws = None

                if self.tech == 'LTE':
                    sheetname = f'Raw_Data_{self.mcs_lte}_TxMax' if self.tx_level > 0 else f'Raw_Data_{self.mcs_lte}_-10dBm'
                    ws = wb[sheetname]
                elif self.tech == 'FR1':
                    sheetname = f'Raw_Data_{self.mcs_fr1}_TxMax' if self.tx_level > 0 else f'Raw_Data_{self.mcs_fr1}_-10dBm'
                    ws = wb[sheetname]
                elif self.tech == 'WCDMA':
                    sheetname = 'Raw_Data'
                    ws = wb[sheetname]
                elif self.tech == 'GSM':
                    sheetname = 'Raw_Data'
                    ws = wb[sheetname]

                if self.tech == 'LTE':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_lte(band, bw, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_lte
                            ws.cell(row, 16).value = self.rb_start_lte
                            ws.cell(row, 17).value = self.mcs_lte
                            ws.cell(row, 18).value = self.tx_path

                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            chan = self.chan_judge_lte(band, bw, tx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = self.rx_path_lte_dict[self.rx_path_lte]
                            ws.cell(row, 3).value = chan  # LMH
                            ws.cell(row, 4).value = tx_freq
                            ws.cell(row, 5).value = tx_freq_level  # this tx level
                            ws.cell(row, 6).value = measured_data[0]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[2][0]
                            ws.cell(row, 9).value = measured_data[2][1]
                            ws.cell(row, 10).value = measured_data[2][2]
                            ws.cell(row, 11).value = measured_data[2][3]
                            ws.cell(row, 12).value = measured_data[3][0]
                            ws.cell(row, 13).value = measured_data[3][1]
                            ws.cell(row, 14).value = measured_data[3][2]
                            ws.cell(row, 15).value = measured_data[3][3]
                            ws.cell(row, 16).value = measured_data[4][0]
                            ws.cell(row, 17).value = measured_data[4][1]
                            ws.cell(row, 18).value = measured_data[4][2]
                            ws.cell(row, 19).value = measured_data[4][3]
                            ws.cell(row, 20).value = self.tx_path
                            row += 1
                elif self.tech == 'FR1':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_fr1(band, bw, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_fr1
                            ws.cell(row, 16).value = self.rb_start_fr1
                            ws.cell(row, 17).value = self.type_fr1
                            ws.cell(row, 18).value = self.tx_path
                            ws.cell(row, 19).value = self.scs
                            ws.cell(row, 20).value = self.rb_state
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            chan = self.chan_judge_fr1(band, bw, tx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = self.rx_path_fr1_dict[self.rx_path_fr1]
                            ws.cell(row, 3).value = chan  # LMH
                            ws.cell(row, 4).value = tx_freq
                            ws.cell(row, 5).value = tx_freq_level  # this tx level
                            ws.cell(row, 6).value = measured_data[0]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[2][0]
                            ws.cell(row, 9).value = measured_data[2][1]
                            ws.cell(row, 10).value = measured_data[2][2]
                            ws.cell(row, 11).value = measured_data[2][3]
                            ws.cell(row, 12).value = measured_data[3][0]
                            ws.cell(row, 13).value = measured_data[3][1]
                            ws.cell(row, 14).value = measured_data[3][2]
                            ws.cell(row, 15).value = measured_data[3][3]
                            ws.cell(row, 16).value = measured_data[4][0]
                            ws.cell(row, 17).value = measured_data[4][1]
                            ws.cell(row, 18).value = measured_data[4][2]
                            ws.cell(row, 19).value = measured_data[4][3]
                            ws.cell(row, 20).value = self.tx_path
                            row += 1
                elif self.tech == 'WCDMA':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:  # this is not used for wcdma
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_wddma(band, bw, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_chan, rx_level in data.items():
                            chan = self.chan_judge_wcdma(band,
                                                         cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, tx_chan))
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = self.rx_path_wcdma_dict[self.rx_path_wcdma]
                            ws.cell(row, 3).value = chan  # LMH
                            ws.cell(row, 4).value = tx_chan  # channel
                            ws.cell(row, 5).value = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, tx_chan, 'tx')
                            ws.cell(row, 6).value = rx_level
                            row += 1
                elif self.tech == 'GSM':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:  # this is not used for gsm
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_gsm(band, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            row += 1

                    elif tx_freq_level <= 100:
                        for rx_chan, rx_level_rssi in data.items():
                            chan = self.chan_judge_gsm(band,
                                                       cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, rx_chan))
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = self.rx_path_gsm_dict[self.rx_path_gsm]
                            ws.cell(row, 3).value = chan  # LMH
                            ws.cell(row, 4).value = rx_chan  # channel
                            ws.cell(row, 5).value = cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, rx_chan, 'rx')
                            ws.cell(row, 6).value = rx_level_rssi[0]
                            ws.cell(row, 7).value = rx_level_rssi[1]
                            row += 1

                wb.save(filename)
                wb.close()

                return filename
            else:  # mode = 0
                if tx_freq_level >= 100:
                    filename = f'Sensitivity_Freq_sweep_{bw}MHZ_{self.tech}.xlsx'
                elif tx_freq_level <= 100:
                    filename = f'_{bw}MHZ_{self.tech}.xlsx'

                if pathlib.Path(filename).exists() is False:
                    logger.info('----------file does not exist----------')
                    wb = openpyxl.Workbook()
                    wb.remove(wb['Sheet'])
                    # to create sheet
                    if self.tech == 'LTE':
                        wb.create_sheet(f'Raw_Data_{self.mcs_lte}_TxMax')

                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Tx_Freq'
                            ws['C1'] = 'Tx_level'
                            ws['D1'] = 'Power'
                            ws['E1'] = 'Rx_Level'
                            ws['F1'] = 'RSRP_RX0'
                            ws['G1'] = 'RSRP_RX1'
                            ws['H1'] = 'RSRP_RX2'
                            ws['I1'] = 'RSRP_RX3'
                            ws['J1'] = 'CINR_RX0'
                            ws['K1'] = 'CINR_RX1'
                            ws['L1'] = 'CINR_RX2'
                            ws['M1'] = 'CINR_RX3'
                            ws['N1'] = 'AGC_RX0'
                            ws['O1'] = 'AGC_RX1'
                            ws['P1'] = 'AGC_RX2'
                            ws['Q1'] = 'AGC_RX3'
                            ws['R1'] = ''

                    elif self.tech == 'FR1':
                        wb.create_sheet(f'Raw_Data_{self.mcs_lte}_TxMax')

                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Tx_Freq'
                            ws['C1'] = 'Tx_level'
                            ws['D1'] = 'Power'
                            ws['E1'] = 'Rx_Level'
                            ws['F1'] = 'RSRP_RX0'
                            ws['G1'] = 'RSRP_RX1'
                            ws['H1'] = 'RSRP_RX2'
                            ws['I1'] = 'RSRP_RX3'
                            ws['J1'] = 'CINR_RX0'
                            ws['K1'] = 'CINR_RX1'
                            ws['L1'] = 'CINR_RX2'
                            ws['M1'] = 'CINR_RX3'
                            ws['N1'] = 'AGC_RX0'
                            ws['O1'] = 'AGC_RX1'
                            ws['P1'] = 'AGC_RX2'
                            ws['Q1'] = 'AGC_RX3'
                            ws['R1'] = ''

                    wb.save(filename)
                    wb.close()

                logger.info('----------file exist----------')
                wb = openpyxl.load_workbook(filename)
                ws = None
                if self.tech == 'LTE':
                    wb.create_sheet(f'Raw_Data_{self.mcs_lte}_TxMax')
                elif self.tech == 'FR1':
                    wb.create_sheet(f'Raw_Data_{self.mcs_fr1}_TxMax')

                if self.tech == 'LTE':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = self.rx_path_lte_dict[self.rx_path_lte]
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[0]  # power
                            ws.cell(row, 6).value = measured_data[1]  # rx_level
                            ws.cell(row, 7).value = measured_data[2][0]  # RSRP
                            ws.cell(row, 8).value = measured_data[2][1]
                            ws.cell(row, 9).value = measured_data[2][2]
                            ws.cell(row, 10).value = measured_data[2][3]
                            ws.cell(row, 11).value = measured_data[3][0]  # CINR
                            ws.cell(row, 12).value = measured_data[3][1]
                            ws.cell(row, 13).value = measured_data[3][2]
                            ws.cell(row, 14).value = measured_data[3][3]
                            ws.cell(row, 15).value = measured_data[4][0]  # AGC
                            ws.cell(row, 16).value = measured_data[4][1]
                            ws.cell(row, 17).value = measured_data[4][2]
                            ws.cell(row, 18).value = measured_data[4][3]
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq
                            ws.cell(row, 3).value = tx_freq_level  # this tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_lte
                            ws.cell(row, 15).value = self.rb_start_lte
                            ws.cell(row, 16).value = self.mcs_lte
                            ws.cell(row, 17).value = self.tx_path

                            row += 1
                elif self.tech == 'FR1':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = self.rx_path_lte_dict[self.rx_path_fr1]
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[0]  # power
                            ws.cell(row, 6).value = measured_data[1]  # rx_level
                            ws.cell(row, 7).value = measured_data[2][0]  # RSRP
                            ws.cell(row, 8).value = measured_data[2][1]
                            ws.cell(row, 9).value = measured_data[2][2]
                            ws.cell(row, 10).value = measured_data[2][3]
                            ws.cell(row, 11).value = measured_data[3][0]  # CINR
                            ws.cell(row, 12).value = measured_data[3][1]
                            ws.cell(row, 13).value = measured_data[3][2]
                            ws.cell(row, 14).value = measured_data[3][3]
                            ws.cell(row, 15).value = measured_data[4][0]  # AGC
                            ws.cell(row, 16).value = measured_data[4][1]
                            ws.cell(row, 17).value = measured_data[4][2]
                            ws.cell(row, 18).value = measured_data[4][3]
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq
                            ws.cell(row, 3).value = tx_freq_level  # this tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_fr1
                            ws.cell(row, 15).value = self.rb_start_fr1
                            ws.cell(row, 16).value = self.type_fr1
                            ws.cell(row, 17).value = self.tx_path
                            ws.cell(row, 18).value = self.scs
                            ws.cell(row, 19).value = self.rb_state

                            row += 1

                wb.save(filename)
                wb.close()

                return filename

    def rx_desense_process(self):
        wb = openpyxl.load_workbook(self.filename)
        self.mcs = self.mcs_lte if self.mcs_lte is not None else self.mcs_fr1

        ws_txmax = wb[f'Raw_Data_{self.mcs}_TxMax']
        ws_txmin = wb[f'Raw_Data_{self.mcs}_-10dBm']
        ws_desens = wb[f'Desens_{self.mcs}']
        for row in range(2, ws_txmax.max_row + 1):
            ws_desens.cell(row, 1).value = ws_txmax.cell(row, 1).value
            ws_desens.cell(row, 2).value = ws_txmax.cell(row, 2).value
            ws_desens.cell(row, 3).value = ws_txmax.cell(row, 3).value
            ws_desens.cell(row, 4).value = ws_txmax.cell(row, 7).value - ws_txmin.cell(row, 7).value
            ws_desens.cell(row, 5).value = ws_txmax.cell(row, 20).value

        wb.save(self.filename)
        wb.close()

    @staticmethod
    def rx_desense_endc_progress():
        wb = openpyxl.load_workbook('Sensitivty_ENDC.xlsx')
        ws_txmax = wb[f'Raw_Data_ENDC_FR1_TxMax']
        ws_txmin = wb[f'Raw_Data_ENDC_FR1_-10dBm']
        ws_desens = wb[f'Desens_ENDC']
        for row in range(2, ws_txmax.max_row + 1):
            ws_desens.cell(row, 1).value = ws_txmax.cell(row, 1).value
            ws_desens.cell(row, 2).value = ws_txmax.cell(row, 2).value
            ws_desens.cell(row, 3).value = ws_txmax.cell(row, 6).value
            ws_desens.cell(row, 4).value = ws_txmax.cell(row, 7).value
            ws_desens.cell(row, 5).value = ws_txmax.cell(row, 8).value
            ws_desens.cell(row, 6).value = ws_txmax.cell(row, 9).value
            ws_desens.cell(row, 7).value = ws_txmax.cell(row, 5).value - ws_txmin.cell(row, 5).value

        wb.save('Sensitivty_ENDC.xlsx')
        wb.close()

    def tx_power_fcc_ce_export_excel(self, data):
        logger.info('----------save to excel----------')
        filename = f'Power_{self.script}_{self.tech}.xlsx'
        if pathlib.Path(filename).exists() is False:
            logger.info('----------file does not exist----------')
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            # to create sheet
            sheetname = self.script
            ws = wb.create_sheet(sheetname)
            ws['A1'] = 'Band'
            ws['B1'] = 'BW'
            ws['C1'] = 'MCS'
            ws['D1'] = 'RB_size'
            ws['E1'] = 'RB_start'
            ws['F1'] = 'Tx_path'
            ws['G1'] = 'Chan'
            ws['H1'] = 'Tx_Freq'
            ws['I1'] = 'Tx_level'
            ws['J1'] = 'Power_measured'

            wb.save(filename)
            wb.close()

        logger.info('----------file exist----------')
        wb = openpyxl.load_workbook(filename)
        ws = wb[self.script]
        band = None
        bw = None
        mcs = None
        rb_size = None
        rb_start = None

        if self.tech == 'LTE':
            band = self.band_lte
            bw = self.bw_lte
            mcs = self.mcs_lte
            rb_size = self.rb_size_lte
            rb_start = self.rb_start_lte

        elif self.tech == 'FR1':
            band = self.band_fr1
            bw = self.bw_fr1
            mcs = self.mcs_fr1
            rb_size = self.rb_size_fr1
            rb_start = self.rb_start_fr1

        for tx_freq, measured_data in data.items():
            max_row = ws.max_row
            row = max_row + 1
            ws.cell(row, 1).value = band
            ws.cell(row, 2).value = bw
            ws.cell(row, 3).value = mcs
            ws.cell(row, 4).value = rb_size
            ws.cell(row, 5).value = rb_start
            ws.cell(row, 6).value = self.tx_path
            ws.cell(row, 7).value = measured_data[0]
            ws.cell(row, 8).value = tx_freq
            ws.cell(row, 9).value = self.tx_level
            ws.cell(row, 10).value = measured_data[1]

        wb.save(filename)
        wb.close()
        return filename

    def tx_power_relative_test_export_excel(self, data, band, bw, tx_freq_level,
                                            mode=0):  # mode general: 0,  mode LMH: 1
        """
        data is dict like:
        tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET]
        """
        logger.info('----------save to excel----------')
        filename = None
        if self.script == 'GENERAL':
            if mode == 1:
                if tx_freq_level >= 100:
                    filename = f'relative power_{bw}MHZ_{self.tech}_LMH.xlsx'
                elif tx_freq_level <= 100:
                    filename = f'TxP_ACLR_EVM_{bw}MHZ_{self.tech}_LMH.xlsx'

                if pathlib.Path(filename).exists() is False:
                    logger.info('----------file does not exist----------')
                    wb = openpyxl.Workbook()
                    wb.remove(wb['Sheet'])
                    # to create sheet
                    if self.tech == 'LTE':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}_PRB')
                            wb.create_sheet(f'Raw_Data_{mcs}_FRB')

                            for sheetname in wb.sheetnames:
                                ws = wb[sheetname]
                                ws['A1'] = 'Band'
                                ws['B1'] = 'Chan'
                                ws['C1'] = 'Tx_Freq'
                                ws['D1'] = 'Tx_level'
                                ws['E1'] = 'Measured_Power'
                                ws['F1'] = 'E_-1'
                                ws['G1'] = 'E_+1'
                                ws['H1'] = 'U_-1'
                                ws['I1'] = 'U_+1'
                                ws['J1'] = 'U_-2'
                                ws['K1'] = 'U_+2'
                                ws['L1'] = 'EVM'
                                ws['M1'] = 'Freq_Err'
                                ws['N1'] = 'IQ_OFFSET'
                                ws['O1'] = 'RB_num'
                                ws['P1'] = 'RB_start'
                                ws['Q1'] = 'MCS'
                                ws['R1'] = 'Tx_Path'
                                ws['S1'] = 'Sync_Path'
                                ws['T1'] = 'AS_Path'
                                ws['U1'] = 'Current(mA)'
                                ws['V1'] = 'Condition'
                                ws['W1'] = 'Temp0'
                                ws['X1'] = 'Temp1'

                    elif self.tech == 'FR1':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}')

                            for sheetname in wb.sheetnames:
                                ws = wb[sheetname]
                                ws['A1'] = 'Band'
                                ws['B1'] = 'Chan'
                                ws['C1'] = 'Tx_Freq'
                                ws['D1'] = 'Tx_level'
                                ws['E1'] = 'Measured_Power'
                                ws['F1'] = 'E_-1'
                                ws['G1'] = 'E_+1'
                                ws['H1'] = 'U_-1'
                                ws['I1'] = 'U_+1'
                                ws['J1'] = 'U_-2'
                                ws['K1'] = 'U_+2'
                                ws['L1'] = 'EVM'
                                ws['M1'] = 'Freq_Err'
                                ws['N1'] = 'IQ_OFFSET'
                                ws['O1'] = 'RB_num'
                                ws['P1'] = 'RB_start'
                                ws['Q1'] = 'Type'
                                ws['R1'] = 'Tx_Path'
                                ws['S1'] = 'SCS(KHz)'
                                ws['T1'] = 'RB_STATE'
                                ws['U1'] = 'Sync_Path'
                                ws['V1'] = 'SRS_Path'
                                ws['W1'] = 'Current(mA)'
                                ws['X1'] = 'Condition'
                                ws['Y1'] = 'Temp0'
                                ws['Z1'] = 'Temp1'

                    elif self.tech == 'WCDMA':

                        wb.create_sheet(f'Raw_Data')

                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Chan'
                            ws['C1'] = 'Channel'
                            ws['D1'] = 'Tx_Freq'
                            ws['E1'] = 'Tx_level'
                            ws['F1'] = 'Measured_Power'
                            ws['G1'] = 'U_-1'
                            ws['H1'] = 'U_+1'
                            ws['I1'] = 'U_-2'
                            ws['J1'] = 'U_+2'
                            ws['K1'] = 'OBW'
                            ws['L1'] = 'EVM'
                            ws['M1'] = 'Freq_Err'
                            ws['N1'] = 'IQ_OFFSET'
                            ws['O1'] = 'Tx_Path'
                            ws['P1'] = 'AS_Path'
                            ws['Q1'] = 'Current(mA)'
                            ws['R1'] = 'Condition'
                            ws['S1'] = 'Temp0'
                            ws['T1'] = 'Temp1'

                    elif self.tech == 'GSM':

                        wb.create_sheet(f'Raw_Data_GMSK')
                        wb.create_sheet(f'Raw_Data_EPSK')

                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Chan'
                            ws['C1'] = 'Channel'
                            ws['D1'] = 'Rx_Freq'
                            ws['E1'] = 'Tx_PCL'
                            ws['F1'] = 'Measured_Power'
                            ws['G1'] = 'Phase_rms'
                            ws['H1'] = 'EVM_rms'
                            ws['I1'] = 'Ferr'
                            ws['J1'] = 'ORFS_MOD_-200'
                            ws['K1'] = 'ORFS_MOD_+200'
                            ws['L1'] = 'ORFS_MOD_-400'
                            ws['M1'] = 'ORFS_MOD_+400'
                            ws['N1'] = 'ORFS_MOD_-600'
                            ws['O1'] = 'ORFS_MOD_+600'
                            ws['P1'] = 'ORFS_SW_-400'
                            ws['Q1'] = 'ORFS_SW_+400'
                            ws['R1'] = 'ORFS_SW_-600'
                            ws['S1'] = 'ORFS_SW_+600'
                            ws['T1'] = 'ORFS_SW_-1200'
                            ws['U1'] = 'ORFS_SW_+1200'
                            ws['V1'] = 'AS_Path'
                            ws['W1'] = 'Current(mA)'
                            ws['X1'] = 'Condition'
                            ws['Y1'] = 'Temp0'
                            ws['Z1'] = 'Temp1'

                    wb.save(filename)
                    wb.close()

                logger.info('----------file exist----------')
                wb = openpyxl.load_workbook(filename)
                ws = None
                if self.tech == 'LTE':
                    ws = wb[f'Raw_Data_{self.mcs_lte}_{self.rb_state}']
                elif self.tech == 'FR1':
                    ws = wb[f'Raw_Data_{self.mcs_fr1}']
                elif self.tech == 'WCDMA':
                    ws = wb[f'Raw_Data']
                elif self.tech == 'GSM':
                    ws = wb[f'Raw_Data_{self.mod_gsm}']

                if self.tech == 'LTE':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_lte(band, bw, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_lte
                            ws.cell(row, 16).value = self.rb_start_lte
                            ws.cell(row, 17).value = self.mcs_lte
                            ws.cell(row, 18).value = self.tx_path
                            ws.cell(row, 19).value = self.sync_path
                            ws.cell(row, 20).value = self.asw_path
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            chan = self.chan_judge_lte(band, bw, tx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # # LMH
                            ws.cell(row, 3).value = tx_freq
                            ws.cell(row, 4).value = tx_freq_level  # this tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_lte
                            ws.cell(row, 16).value = self.rb_start_lte
                            ws.cell(row, 17).value = self.mcs_lte
                            ws.cell(row, 18).value = self.tx_path
                            ws.cell(row, 19).value = self.sync_path
                            ws.cell(row, 20).value = self.asw_path
                            ws.cell(row, 21).value = measured_data[10]
                            ws.cell(row, 22).value = wt.condition
                            ws.cell(row, 23).value = measured_data[11]
                            ws.cell(row, 24).value = measured_data[12]

                            row += 1

                elif self.tech == 'FR1':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_fr1(band, bw, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_lte
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_fr1
                            ws.cell(row, 16).value = self.rb_start_fr1
                            ws.cell(row, 17).value = self.type_fr1
                            ws.cell(row, 18).value = self.tx_path
                            ws.cell(row, 19).value = self.scs
                            ws.cell(row, 20).value = self.rb_state
                            ws.cell(row, 21).value = self.sync_path
                            ws.cell(row, 22).value = self.srs_path

                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            chan = self.chan_judge_fr1(band, bw, tx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = tx_freq
                            ws.cell(row, 4).value = tx_freq_level  # this tx_level
                            ws.cell(row, 5).value = measured_data[3]
                            ws.cell(row, 6).value = measured_data[2]
                            ws.cell(row, 7).value = measured_data[4]
                            ws.cell(row, 8).value = measured_data[1]
                            ws.cell(row, 9).value = measured_data[5]
                            ws.cell(row, 10).value = measured_data[0]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = self.rb_size_fr1
                            ws.cell(row, 16).value = self.rb_start_fr1
                            ws.cell(row, 17).value = self.type_fr1
                            ws.cell(row, 18).value = self.tx_path
                            ws.cell(row, 19).value = self.scs
                            ws.cell(row, 20).value = self.rb_state
                            ws.cell(row, 21).value = self.sync_path
                            ws.cell(row, 22).value = self.srs_path
                            ws.cell(row, 23).value = measured_data[10]
                            ws.cell(row, 24).value = wt.condition
                            ws.cell(row, 25).value = measured_data[11]
                            ws.cell(row, 26).value = measured_data[12]

                            row += 1

                elif self.tech == 'WCDMA':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            chan = self.chan_judge_wcdma(band, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = cm_pmt_ftm.trandfer_freq2chan_wcdma(self.band_wcdma, tx_freq_level,
                                                                                        'tx')  # this channel
                            ws.cell(row, 4).value = tx_freq_level  # this tx_freq_wcdma
                            ws.cell(row, 5).value = tx_level
                            ws.cell(row, 6).value = measured_data[0]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[2]
                            ws.cell(row, 9).value = measured_data[3]
                            ws.cell(row, 10).value = measured_data[4]
                            ws.cell(row, 11).value = measured_data[5]
                            ws.cell(row, 12).value = measured_data[6]
                            ws.cell(row, 13).value = measured_data[7]
                            ws.cell(row, 14).value = measured_data[8]
                            ws.cell(row, 15).value = self.tx_path
                            ws.cell(row, 16).value = self.asw_path
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            chan = self.chan_judge_wcdma(band, tx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # # LMH
                            ws.cell(row, 3).value = cm_pmt_ftm.trandfer_freq2chan_wcdma(self.band_wcdma, tx_freq,
                                                                                        'tx')  # this channel
                            ws.cell(row, 4).value = tx_freq
                            ws.cell(row, 5).value = tx_freq_level  # this tx_level
                            ws.cell(row, 6).value = measured_data[0]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[2]
                            ws.cell(row, 9).value = measured_data[3]
                            ws.cell(row, 10).value = measured_data[4]
                            ws.cell(row, 11).value = measured_data[5]
                            ws.cell(row, 12).value = measured_data[6]
                            ws.cell(row, 13).value = measured_data[7]
                            ws.cell(row, 14).value = measured_data[8]
                            ws.cell(row, 15).value = self.tx_path
                            ws.cell(row, 16).value = self.asw_path
                            ws.cell(row, 17).value = measured_data[9]
                            ws.cell(row, 18).value = wt.condition
                            ws.cell(row, 19).value = measured_data[10]
                            ws.cell(row, 20).value = measured_data[11]
                            row += 1

                elif self.tech == 'GSM':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_pcl, measured_data in data.items():
                            chan = self.chan_judge_gsm(band, tx_freq_level)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = cm_pmt_ftm.transfer_freq2chan_gsm(self.band_gsm, tx_freq_level)
                            ws.cell(row, 4).value = tx_freq_level  # this rx_freq_gsm
                            ws.cell(row, 5).value = tx_pcl
                            ws.cell(row, 6).value = measured_data[0]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[2]
                            ws.cell(row, 9).value = measured_data[3]
                            ws.cell(row, 10).value = measured_data[4]
                            ws.cell(row, 11).value = measured_data[5]
                            ws.cell(row, 12).value = measured_data[6]
                            ws.cell(row, 13).value = measured_data[7]
                            ws.cell(row, 14).value = measured_data[8]
                            ws.cell(row, 15).value = measured_data[9]
                            ws.cell(row, 16).value = measured_data[10]
                            ws.cell(row, 17).value = measured_data[11]
                            ws.cell(row, 18).value = measured_data[12]
                            ws.cell(row, 19).value = measured_data[13]
                            ws.cell(row, 20).value = measured_data[14]
                            ws.cell(row, 21).value = measured_data[15]
                            ws.cell(row, 22).value = self.asw_path

                            row += 1

                    elif tx_freq_level <= 100:
                        for rx_freq, measured_data in data.items():
                            chan = self.chan_judge_gsm(band, rx_freq)
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = chan  # LMH
                            ws.cell(row, 3).value = cm_pmt_ftm.transfer_freq2chan_gsm(self.band_gsm, rx_freq)
                            ws.cell(row, 4).value = rx_freq  # this rx_freq_gsm
                            ws.cell(row, 5).value = tx_freq_level  # this pcl
                            ws.cell(row, 6).value = measured_data[0]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[2]
                            ws.cell(row, 9).value = measured_data[3]
                            ws.cell(row, 10).value = measured_data[4]
                            ws.cell(row, 11).value = measured_data[5]
                            ws.cell(row, 12).value = measured_data[6]
                            ws.cell(row, 13).value = measured_data[7]
                            ws.cell(row, 14).value = measured_data[8]
                            ws.cell(row, 15).value = measured_data[9]
                            ws.cell(row, 16).value = measured_data[10]
                            ws.cell(row, 17).value = measured_data[11]
                            ws.cell(row, 18).value = measured_data[12]
                            ws.cell(row, 19).value = measured_data[13]
                            ws.cell(row, 20).value = measured_data[14]
                            ws.cell(row, 21).value = measured_data[15]
                            ws.cell(row, 22).value = self.asw_path
                            ws.cell(row, 23).value = measured_data[16]
                            ws.cell(row, 24).value = wt.condition
                            ws.cell(row, 25).value = measured_data[17]
                            ws.cell(row, 26).value = measured_data[18]

                            row += 1

                wb.save(filename)
                wb.close()

                return filename
            else:  # mode = 0
                if tx_freq_level >= 100:
                    filename = f'Tx_level_sweep_{bw}MHZ_{self.tech}.xlsx'
                elif tx_freq_level <= 100:
                    if self.tx_1rb_filename_judge:
                        filename = f'Tx_1RB_sweep_{bw}MHZ_{self.tech}.xlsx'
                    else:
                        filename = f'Freq_sweep_{bw}MHZ_{self.tech}.xlsx'

                if pathlib.Path(filename).exists() is False:
                    logger.info('----------file does not exist----------')
                    wb = openpyxl.Workbook()
                    wb.remove(wb['Sheet'])
                    # to create sheet
                    if self.tech == 'LTE':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}_PRB')
                            wb.create_sheet(f'Raw_Data_{mcs}_FRB')

                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Tx_Freq'
                            ws['C1'] = 'Tx_level'
                            ws['D1'] = 'Measured_Power'
                            ws['E1'] = 'E_-1'
                            ws['F1'] = 'E_+1'
                            ws['G1'] = 'U_-1'
                            ws['H1'] = 'U_+1'
                            ws['I1'] = 'U_-2'
                            ws['J1'] = 'U_+2'
                            ws['K1'] = 'EVM'
                            ws['L1'] = 'Freq_Err'
                            ws['M1'] = 'IQ_OFFSET'
                            ws['N1'] = 'RB_num'
                            ws['O1'] = 'RB_start'
                            ws['P1'] = 'MCS'
                            ws['Q1'] = 'Tx_Path'
                            ws['R1'] = 'Sync_Path'
                            ws['S1'] = 'AS_Path'
                            ws['T1'] = 'Current(mA)'
                            ws['U1'] = 'Condition'

                    elif self.tech == 'FR1':
                        for mcs in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw10 might not have licesnse of Q256
                            wb.create_sheet(f'Raw_Data_{mcs}')

                            for sheetname in wb.sheetnames:
                                ws = wb[sheetname]
                                ws['A1'] = 'Band'
                                ws['B1'] = 'Tx_Freq'
                                ws['C1'] = 'Tx_level'
                                ws['D1'] = 'Measured_Power'
                                ws['E1'] = 'E_-1'
                                ws['F1'] = 'E_+1'
                                ws['G1'] = 'U_-1'
                                ws['H1'] = 'U_+1'
                                ws['I1'] = 'U_-2'
                                ws['J1'] = 'U_+2'
                                ws['K1'] = 'EVM'
                                ws['L1'] = 'Freq_Err'
                                ws['M1'] = 'IQ_OFFSET'
                                ws['N1'] = 'RB_num'
                                ws['O1'] = 'RB_start'
                                ws['P1'] = 'Type'
                                ws['Q1'] = 'Tx_Path'
                                ws['R1'] = 'SCS(KHz)'
                                ws['S1'] = 'RB_STATE'
                                ws['T1'] = 'Sync_Path'
                                ws['U1'] = 'SRS_Path'
                                ws['V1'] = 'Current(mA)'
                                ws['W1'] = 'Condition'

                    elif self.tech == 'WCDMA':
                        wb.create_sheet(f'Raw_Data')
                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Channel'
                            ws['C1'] = 'Tx_Freq'
                            ws['D1'] = 'Tx_level'
                            ws['E1'] = 'Measured_Power'
                            ws['F1'] = 'U_-1'
                            ws['G1'] = 'U_+1'
                            ws['H1'] = 'U_-2'
                            ws['I1'] = 'U_+2'
                            ws['J1'] = 'OBW'
                            ws['K1'] = 'EVM'
                            ws['L1'] = 'Freq_Err'
                            ws['M1'] = 'IQ_OFFSET'
                            ws['N1'] = 'Tx_Path'
                            ws['O1'] = 'AS_Path'
                            ws['P1'] = 'Current(mA)'
                            ws['Q1'] = 'Condition'

                    elif self.tech == 'GSM':
                        wb.create_sheet(f'Raw_Data_GMSK')
                        wb.create_sheet(f'Raw_Data_EPSK')

                        for sheetname in wb.sheetnames:
                            ws = wb[sheetname]
                            ws['A1'] = 'Band'
                            ws['B1'] = 'Channel'
                            ws['C1'] = 'Rx_Freq'
                            ws['D1'] = 'Tx_PCL'
                            ws['E1'] = 'Measured_Power'
                            ws['F1'] = 'Phase_rms'
                            ws['G1'] = 'EVM_rms'
                            ws['H1'] = 'Ferr'
                            ws['I1'] = 'ORFS_MOD_-200'
                            ws['J1'] = 'ORFS_MOD_+200'
                            ws['K1'] = 'ORFS_MOD_-400'
                            ws['L1'] = 'ORFS_MOD_+400'
                            ws['M1'] = 'ORFS_MOD_-600'
                            ws['N1'] = 'ORFS_MOD_+600'
                            ws['O1'] = 'ORFS_SW_-400'
                            ws['P1'] = 'ORFS_SW_+400'
                            ws['Q1'] = 'ORFS_SW_-600'
                            ws['R1'] = 'ORFS_SW_+600'
                            ws['S1'] = 'ORFS_SW_-1200'
                            ws['T1'] = 'ORFS_SW_+1200'
                            ws['U1'] = 'AS_Path'
                            ws['V1'] = 'Current(mA)'
                            ws['W1'] = 'Condition'

                    wb.save(filename)
                    wb.close()

                logger.info('----------file exist----------')
                wb = openpyxl.load_workbook(filename)
                ws = None
                if self.tech == 'LTE':
                    ws = wb[f'Raw_Data_{self.mcs_lte}_{self.rb_state}']
                elif self.tech == 'FR1':
                    ws = wb[f'Raw_Data_{self.mcs_fr1}']
                elif self.tech == 'WCDMA':
                    ws = wb[f'Raw_Data']
                elif self.tech == 'GSM':
                    ws = wb[f'Raw_Data_{self.mod_gsm}']

                if self.tech == 'LTE':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq_level  # this tx_freq
                            ws.cell(row, 3).value = tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_lte
                            ws.cell(row, 15).value = self.rb_start_lte
                            ws.cell(row, 16).value = self.mcs_lte
                            ws.cell(row, 17).value = self.tx_path
                            ws.cell(row, 18).value = self.sync_path
                            ws.cell(row, 19).value = self.asw_path
                            ws.cell(row, 20).value = measured_data[10]
                            ws.cell(row, 21).value = wt.condition
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq
                            ws.cell(row, 3).value = tx_freq_level  # this tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_lte
                            ws.cell(row, 15).value = self.rb_start_lte
                            ws.cell(row, 16).value = self.mcs_lte
                            ws.cell(row, 17).value = self.tx_path
                            ws.cell(row, 18).value = self.sync_path
                            ws.cell(row, 19).value = self.asw_path
                            row += 1

                elif self.tech == 'FR1':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq_level  # this tx_freq
                            ws.cell(row, 3).value = tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_fr1
                            ws.cell(row, 15).value = self.rb_start_fr1
                            ws.cell(row, 16).value = self.type_fr1
                            ws.cell(row, 17).value = self.tx_path
                            ws.cell(row, 18).value = self.scs
                            ws.cell(row, 19).value = self.rb_state
                            ws.cell(row, 20).value = self.sync_path
                            ws.cell(row, 21).value = self.srs_path
                            ws.cell(row, 22).value = measured_data[10]
                            ws.cell(row, 23).value = wt.condition
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq
                            ws.cell(row, 3).value = tx_freq_level  # this tx_level
                            ws.cell(row, 4).value = measured_data[3]
                            ws.cell(row, 5).value = measured_data[2]
                            ws.cell(row, 6).value = measured_data[4]
                            ws.cell(row, 7).value = measured_data[1]
                            ws.cell(row, 8).value = measured_data[5]
                            ws.cell(row, 9).value = measured_data[0]
                            ws.cell(row, 10).value = measured_data[6]
                            ws.cell(row, 11).value = measured_data[7]
                            ws.cell(row, 12).value = measured_data[8]
                            ws.cell(row, 13).value = measured_data[9]
                            ws.cell(row, 14).value = self.rb_size_fr1
                            ws.cell(row, 15).value = self.rb_start_fr1
                            ws.cell(row, 16).value = self.type_fr1
                            ws.cell(row, 17).value = self.tx_path
                            ws.cell(row, 18).value = self.scs
                            ws.cell(row, 19).value = self.rb_state
                            ws.cell(row, 20).value = self.sync_path
                            ws.cell(row, 21).value = self.srs_path
                            row += 1

                elif self.tech == 'WCDMA':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_level, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = cm_pmt_ftm.trandfer_freq2chan_wcdma(self.band_wcdma, tx_freq_level,
                                                                                        'tx')  # this channel
                            ws.cell(row, 3).value = tx_freq_level  # this tx_freq_wcdma
                            ws.cell(row, 4).value = tx_level
                            ws.cell(row, 5).value = measured_data[0]
                            ws.cell(row, 6).value = measured_data[1]
                            ws.cell(row, 7).value = measured_data[2]
                            ws.cell(row, 8).value = measured_data[3]
                            ws.cell(row, 9).value = measured_data[4]
                            ws.cell(row, 10).value = measured_data[5]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = self.tx_path
                            ws.cell(row, 15).value = self.asw_path
                            ws.cell(row, 16).value = measured_data[9]
                            ws.cell(row, 17).value = wt.condition
                            row += 1

                    elif tx_freq_level <= 100:
                        for tx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = tx_freq  # this channel
                            ws.cell(row, 3).value = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, tx_freq, 'tx')
                            ws.cell(row, 4).value = tx_freq_level  # this tx_level
                            ws.cell(row, 5).value = measured_data[0]
                            ws.cell(row, 6).value = measured_data[1]
                            ws.cell(row, 7).value = measured_data[2]
                            ws.cell(row, 8).value = measured_data[3]
                            ws.cell(row, 9).value = measured_data[4]
                            ws.cell(row, 10).value = measured_data[5]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = self.tx_path
                            ws.cell(row, 15).value = self.asw_path
                            row += 1

                elif self.tech == 'GSM':
                    max_row = ws.max_row
                    row = max_row + 1
                    if tx_freq_level >= 100:
                        for tx_pcl, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = cm_pmt_ftm.transfer_freq2chan_gsm(self.band_gsm, tx_freq_level)
                            ws.cell(row, 3).value = tx_freq_level  # this rx_freq_gsm
                            ws.cell(row, 4).value = tx_pcl
                            ws.cell(row, 5).value = measured_data[0]
                            ws.cell(row, 6).value = measured_data[1]
                            ws.cell(row, 7).value = measured_data[2]
                            ws.cell(row, 8).value = measured_data[3]
                            ws.cell(row, 9).value = measured_data[4]
                            ws.cell(row, 10).value = measured_data[5]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = measured_data[10]
                            ws.cell(row, 16).value = measured_data[11]
                            ws.cell(row, 17).value = measured_data[12]
                            ws.cell(row, 18).value = measured_data[13]
                            ws.cell(row, 19).value = measured_data[14]
                            ws.cell(row, 20).value = measured_data[15]
                            ws.cell(row, 21).value = self.asw_path
                            ws.cell(row, 22).value = measured_data[16]
                            ws.cell(row, 23).value = wt.condition

                            row += 1

                    elif tx_freq_level <= 100:
                        for rx_freq, measured_data in data.items():
                            ws.cell(row, 1).value = band
                            ws.cell(row, 2).value = cm_pmt_ftm.transfer_freq2chan_gsm(self.band_gsm, rx_freq)
                            ws.cell(row, 3).value = rx_freq  # this rx_freq_gsm
                            ws.cell(row, 4).value = tx_freq_level  # this pcl
                            ws.cell(row, 5).value = measured_data[0]
                            ws.cell(row, 6).value = measured_data[1]
                            ws.cell(row, 7).value = measured_data[2]
                            ws.cell(row, 8).value = measured_data[3]
                            ws.cell(row, 9).value = measured_data[4]
                            ws.cell(row, 10).value = measured_data[5]
                            ws.cell(row, 11).value = measured_data[6]
                            ws.cell(row, 12).value = measured_data[7]
                            ws.cell(row, 13).value = measured_data[8]
                            ws.cell(row, 14).value = measured_data[9]
                            ws.cell(row, 15).value = measured_data[10]
                            ws.cell(row, 16).value = measured_data[11]
                            ws.cell(row, 17).value = measured_data[12]
                            ws.cell(row, 18).value = measured_data[13]
                            ws.cell(row, 19).value = measured_data[14]
                            ws.cell(row, 20).value = measured_data[15]
                            ws.cell(row, 21).value = self.asw_path

                            row += 1

                wb.save(filename)
                wb.close()

                return filename

    def tx_power_relative_test_initial_gsm(self):
        logger.info('----------Relatvie test initial----------')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:PVT 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:SMOD 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SCO:SSW 5')
        self.command_cmw100_write(f'CONF:GSM:MEAS:SCEN:ACT STAN')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'ROUT:WCDMA:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:UMAR 10.00')
        self.command_cmw100_write(f"TRIG:GSM:MEAS:MEV:SOUR 'Power'")
        self.command_cmw100_write(f'TRIG:WCDM:MEAS:MEV:THR -30')
        self.command_cmw100_write(f'TRIG:GSM:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON, OFF, ON')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(
            f'CONF:GSM:MEAS:MEV:SMOD:OFR 100KHZ,200KHZ,250KHZ,400KHZ,600KHZ,800KHZ,1600KHZ,1800KHZ,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF')
        self.command_cmw100_write(f'CONF:GSM:MEAS:MEV:SMOD:EAR ON,6,45,ON,90,129')
        self.command_cmw100_write(
            f'CONF:GSM:MEAS:MEV:SSW:OFR 400KHZ,600KHZ,1200KHZ,1800KHZ,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF,OFF')
        self.command_cmw100_query(f'SYST:ERR:ALL?')

    def tx_power_relative_test_initial_wcdma(self):
        logger.info('----------Relatvie test initial----------')
        self.command_cmw100_write(f'*CLS')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:SCO:SPEC 5')
        self.command_cmw100_write(f'ROUT:WCDMA:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:UMAR 10.00')
        self.command_cmw100_write(f"TRIG:WCDM:MEAS:MEV:SOUR 'Free Run (Fast sync)'")
        self.command_cmw100_write(f'TRIG:WCDM:MEAS:MEV:THR -30')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:MEV:RES:ALL ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON,ON')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_query(f'SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:BAND OB{self.band_wcdma}')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:RFS:FREQ {self.tx_chan_wcdma} CH')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:DPDC ON')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:SFOR 0')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:SCOD 13496235')
        self.command_cmw100_write(f'CONF:WCDM:MEAS:UES:ULC WCDM')
        self.command_cmw100_query(f'SYST:ERR:ALL?')

    def tx_power_relative_test_initial_lte(self):
        logger.info('----------Relatvie test initial----------')
        mode = 'TDD' if self.band_lte in [38, 39, 40, 41, 42, 48] else 'FDD'
        self.command_cmw100_write(f'CONF:LTE:MEAS:DMODe {mode}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:BAND OB{self.band_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:FREQ {self.tx_freq_lte}KHz')
        self.command_cmw100_query(f'*OPC?')
        rb = f'0{self.bw_lte * 10}' if self.bw_lte < 10 else f'{self.bw_lte * 10}'
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CBAN B{rb}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOD:MSCH {self.mcs_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:NRB {self.rb_size_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:ORB {self.rb_start_lte}')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:CPR NORM')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:DSSP 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO OFF')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MOEX ON')
        lim1 = -10 if self.bw_lte == 1.4 else -13 if self.bw_lte == 3 else -15 if self.bw_lte == 5 else -18 if self.bw_lte == 10 else -20 if self.bw_lte == 15 else -21
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM1:CBAN{self.bw_lte * 10} ON,0MHz,1MHz,{lim1},K030')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM2:CBAN{self.bw_lte * 10} ON,1MHz,2.5MHz,-10,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN{self.bw_lte * 10} ON,2.5MHz,2.8MHz,-25,M1') if self.bw_lte < 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM3:CBAN{self.bw_lte * 10} ON,2.5MHz,2.8MHz,-10,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN{self.bw_lte * 10} ON,2.8MHz,5MHz,-10,M1') if self.bw_lte >= 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM4:CBAN{self.bw_lte * 10} OFF,2.8MHz,5MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN{self.bw_lte * 10} ON,5MHz,6MHz,-13,M1') if self.bw_lte > 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN{self.bw_lte * 10} OFF,5MHz,6MHz,-25,M1') if self.bw_lte < 3 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM5:CBAN{self.bw_lte * 10} ON,5MHz,6MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN{self.bw_lte * 10} ON,6MHz,10MHz,-13,M1') if self.bw_lte > 5 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN{self.bw_lte * 10} OFF,6MHz,10MHz,-25,M1') if self.bw_lte < 5 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM6:CBAN{self.bw_lte * 10} ON,6MHz,10MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN{self.bw_lte * 10} ON,10MHz,15MHz,-13,M1') if self.bw_lte > 10 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN{self.bw_lte * 10} OFF,10MHz,15MHz,-25,M1') if self.bw_lte < 10 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM7:CBAN{self.bw_lte * 10} ON,10MHz,15MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN{self.bw_lte * 10} ON,15MHz,20MHz,-13,M1') if self.bw_lte > 15 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN{self.bw_lte * 10} OFF,15MHz,20MHz,-25,M1') if self.bw_lte < 15 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM8:CBAN{self.bw_lte * 10} ON,15MHz,20MHz,-25,M1')
        self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN{self.bw_lte * 10} ON,20MHz,25MHz,-25,M1') if self.bw_lte == 20 else self.command_cmw100_write(
            f'CONF:LTE:MEAS:MEV:LIM:SEM:LIM9:CBAN{self.bw_lte * 10} OFF,20MHz,25MHz,-25,M1')

        self.command_cmw100_query(f'SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONFigure:LTE:MEAS:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:ENP {self.tx_level + 5}')
        self.command_cmw100_write(f'ROUT:LTE:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RBAL:AUTO ON')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f"TRIG:LTE:MEAS:MEV:SOUR 'GPRF Gen1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:LTE:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:MEV:MSUB 2, 10, 0')
        self.command_cmw100_write(f'CONF:LTE:MEAS:SCEN:ACT SAL')
        self.command_cmw100_query(f'SYST:ERR:ALL?')
        self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'ROUT:GPRF:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:SCO 2')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:REP SING')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:LIST OFF')
        self.command_cmw100_write(f"TRIGger:GPRF:MEAS:POWer:SOURce 'Free Run'")
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:TRIG:SLOP REDG')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:SLEN 5.0e-3')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:POW:MLEN 8.0e-4')
        self.command_cmw100_write(f'TRIGger:GPRF:MEAS:POWer:OFFSet 2.1E-3')
        self.command_cmw100_write(f'TRIG:GPRF:MEAS:POW:MODE ONCE')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:GPRF:MEAS:RFS:ENP {self.tx_level + 5}.00')

    def tx_power_relative_test_initial_fr1(self):
        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78,
                                     79] else 0  # for now FDD is forced to 15KHz and TDD is to be 30KHz
        scs = 15 * (2 ** scs)  # for now TDD only use 30KHz, FDD only use 15KHz
        logger.info('----------Relatvie test initial----------')
        mode = 'TDD' if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 75, 76, 77, 78, 79] else 'FDD'
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:DMODe {mode}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:BAND OB{self.band_fr1}')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:FREQ {self.tx_freq_fr1}KHz')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:PLC 0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:MOEX ON')
        bw = f'00{self.bw_fr1}' if self.bw_fr1 < 10 else f'0{self.bw_fr1}' if 10 <= self.bw_fr1 < 100 else self.bw_fr1
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:BWC S{scs}K, B{bw}')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA1:CBAN{self.bw_fr1}   ON, 0.015MHz, 0.0985MHz, {round(-13.5 - 10 * math.log10(self.bw_fr1 / 5), 1)},K030')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA2:CBAN{self.bw_fr1}   ON,   1.5MHz,    4.5MHz,  -8.5,  M1')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA3:CBAN{self.bw_fr1}   ON,   5.5MHz,   {round(-0.5 + self.bw_fr1, 1)}MHz, -11.5,  M1')
        self.command_cmw100_write(
            f'CONF:NRS:MEAS:MEV:LIM:SEM:AREA4:CBAN{self.bw_fr1}   ON, 0 {round(0.5 + self.bw_fr1, 1)}MHz,  {round(4.5 + self.bw_fr1, 1)}MHz, -23.5,  M1')
        _256Q_flag = 2 if self.mcs_fr1 == 'Q256' else 0
        self.command_cmw100_write(
            f'CONFigure:NRSub:MEASurement:MEValuation:PUSChconfig {self.mcs_fr1},A,OFF,{self.rb_size_fr1},{self.rb_start_fr1},14,0,T1,SING,{_256Q_flag},2')
        type_ = 'ON' if self.type_fr1 == 'DFTS' else 'OFF'  # DFTS: ON, CP: OFF
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:DFTPrecoding {type_}')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PCOMp OFF, 6000E+6')
        self.command_cmw100_query(f'*OPC?')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:REPetition SING')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:PLCid 0')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:CTYPe PUSC')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PER MS25')
        self.command_cmw100_write(f'CONF:NRS:MEAS:ULDL:PATT S{scs}K, 3,0,1,14 ')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'ROUT:NRS:MEAS:SCEN:SAL R1{self.port_tx}, RX1')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:UMAR 10.000000')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:MOD 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:ACLR 5')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:SCO:SPEC:SEM 5')
        self.command_cmw100_write(f"TRIG:NRS:MEAS:MEV:SOUR 'GPRF GEN1: Restart Marker'")
        self.command_cmw100_write(f'TRIG:NRS:MEAS:MEV:THR -20.0')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:REP SING')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:RES:ALL ON, ON, ON, ON, ON, ON, ON, ON, ON, ON')
        self.command_cmw100_write(f'CONF:NRS:MEAS:MEV:NSUB 10')
        self.command_cmw100_write(f'CONFigure:NRSub:MEASurement:MEValuation:MSLot ALL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:SCEN:ACT SAL')
        self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:EATT {self.loss_tx}')
        self.command_cmw100_query(f'*OPC?')

    def tx_power_aclr_evm_lmh_pipeline_gsm(self):
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.mod_gsm = wt.mod_gsm
        self.tsc = 0 if self.mod_gsm == 'GMSK' else 5
        for tech in wt.tech:
            if tech == 'GSM' and wt.gsm_bands != []:
                self.tech = 'GSM'
                for band in wt.gsm_bands:
                    self.pcl = wt.tx_pcl_lb if band in [850, 900] else wt.tx_pcl_mb
                    self.band_gsm = band
                    self.tx_power_aclr_evm_lmh_gsm(plot=False)
                self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode

    def tx_power_aclr_evm_lmh_pipeline_wcdma(self):
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        for tech in wt.tech:
            if tech == 'WCDMA' and wt.wcdma_bands != []:
                self.tech = 'WCDMA'
                for band in wt.wcdma_bands:
                    self.band_wcdma = band
                    self.tx_power_aclr_evm_lmh_wcdma(plot=False)
                self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode

    def tx_power_aclr_evm_lmh_pipeline_lte(self):
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        items = [
            (tech, tx_path, bw, band)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.lte_bandwidths
            for band in wt.lte_bands
        ]
        for item in items:
            if item[0] == 'LTE' and wt.lte_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_lte = item[2]
                self.band_lte = item[3]
                if self.bw_lte in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                    self.tx_power_aclr_evm_lmh_lte(plot=False)
                else:
                    logger.info(f'B{self.band_lte} does not have BW {self.bw_lte}MHZ')
        for bw in wt.lte_bandwidths:
            try:
                self.filename = f'TxP_ACLR_EVM_{bw}MHZ_{self.tech}_LMH.xlsx'
                self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
            except TypeError:
                logger.info(f'there is no data to plot because the band does not have this BW ')
            except FileNotFoundError:
                logger.info(f'there is not file to plot BW{bw} ')

    def tx_power_aclr_evm_lmh_pipeline_fr1(self):
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.sa_nsa_mode = wt.sa_nsa
        items = [
            (tech, tx_path, bw, band, type_)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.fr1_bandwidths
            for band in wt.fr1_bands
            for type_ in wt.type_fr1
        ]

        for item in items:
            if item[0] == 'FR1' and wt.fr1_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_fr1 = item[2]
                self.band_fr1 = item[3]
                self.type_fr1 = item[4]
                if self.bw_fr1 in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                    self.tx_power_aclr_evm_lmh_fr1(plot=False)
                else:
                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')
        for bw in wt.fr1_bandwidths:
            try:
                self.filename = f'TxP_ACLR_EVM_{bw}MHZ_{self.tech}_LMH.xlsx'
                self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
            except TypeError:
                logger.info(f'there is no data to plot because the band does not have this BW ')
            except FileNotFoundError:
                logger.info(f'there is not file to plot BW{bw} ')

    def tx_power_pipline_fcc_fr1(self):  # band > bw > mcs > rb
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.sa_nsa_mode = wt.sa_nsa
        items = [
            (tech, tx_path, bw, band, type_)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.fr1_bandwidths
            for band in wt.fr1_bands
            for type_ in wt.type_fr1
        ]

        for item in items:
            if item[0] == 'FR1' and wt.fr1_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_fr1 = item[2]
                self.band_fr1 = item[3]
                self.type_fr1 = item[4]
                if self.bw_fr1 in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                    self.tx_power_fcc_fr1()
                else:
                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')

    def tx_power_pipline_ce_fr1(self):  # band > bw > mcs > rb
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.sa_nsa_mode = wt.sa_nsa
        items = [
            (tech, tx_path, bw, band, type_)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.fr1_bandwidths
            for band in wt.fr1_bands
            for type_ in wt.type_fr1
        ]

        for item in items:
            if item[0] == 'FR1' and wt.fr1_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_fr1 = item[2]
                self.band_fr1 = item[3]
                self.type_fr1 = item[4]
                if self.bw_fr1 in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                    self.tx_power_ce_fr1()
                else:
                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')

    def tx_freq_sweep_pipline_gsm(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        # self.chan = wt.channel
        self.mod_gsm = wt.mod_gsm
        self.tsc = 0 if self.mod_gsm == 'GMSK' else 5
        for tech in wt.tech:
            if tech == 'GSM' and wt.gsm_bands != []:
                self.tech = tech
                for band in wt.gsm_bands:
                    self.pcl = wt.tx_pcl_lb if band in [850, 900] else wt.tx_pcl_mb
                    self.band_gsm = band
                    self.tx_freq_sweep_progress_gsm(plot=False)
                self.txp_aclr_evm_plot(self.filename, mode=0)

    def tx_freq_sweep_pipline_wcdma(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        # self.chan = wt.channel
        for tech in wt.tech:
            if tech == 'WCDMA' and wt.wcdma_bands != []:
                self.tech = tech
                for band in wt.wcdma_bands:
                    self.band_wcdma = band
                    self.tx_freq_sweep_progress_wcdma(plot=False)
                self.txp_aclr_evm_plot(self.filename, mode=0)

    def tx_freq_sweep_pipline_lte(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        # self.chan = wt.channel
        items = [
            (tech, tx_path, bw, band)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.lte_bandwidths
            for band in wt.lte_bands
        ]
        for item in items:
            if item[0] == 'LTE' and wt.lte_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_lte = item[2]
                self.band_lte = item[3]
                if self.bw_lte in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                    self.tx_freq_sweep_progress_lte(plot=False)
                else:
                    logger.info(f'B{self.band_lte} does not have BW {self.bw_lte}MHZ')

        for bw in wt.lte_bandwidths:
            try:
                self.filename = f'Freq_sweep_{bw}MHZ_{self.tech}.xlsx'
                self.txp_aclr_evm_plot(self.filename, mode=0)
            except TypeError:
                logger.info(f'there is no data to plot because the band does not have this BW ')
            except FileNotFoundError:
                logger.info(f'there is not file to plot BW{bw} ')

    def tx_freq_sweep_pipline_fr1(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        # self.chan = wt.channel
        self.sa_nsa_mode = wt.sa_nsa
        items = [
            (tech, tx_path, bw, band, type_)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.fr1_bandwidths
            for band in wt.fr1_bands
            for type_ in wt.type_fr1
        ]

        for item in items:
            if item[0] == 'FR1' and wt.fr1_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_fr1 = item[2]
                self.band_fr1 = item[3]
                self.type_fr1 = item[4]
                if self.bw_fr1 in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                    self.tx_freq_sweep_progress_fr1(plot=False)
                else:
                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')
        for bw in wt.fr1_bandwidths:
            try:
                self.filename = f'Freq_sweep_{bw}MHZ_{self.tech}.xlsx'
                self.txp_aclr_evm_plot(self.filename, mode=0)
            except TypeError:
                logger.info(f'there is no data to plot because the band does not have this BW ')
            except FileNotFoundError:
                logger.info(f'there is not file to plot BW{bw} ')

    def tx_level_sweep_pipeline_gsm(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.mod_gsm = wt.mod_gsm
        self.tsc = 0 if self.mod_gsm == 'GMSK' else 5
        for tech in wt.tech:
            if tech == 'GSM' and wt.gsm_bands != []:
                self.tech = tech
                for band in wt.gsm_bands:
                    self.pcl = wt.tx_pcl_lb if band in [850, 900] else wt.tx_pcl_mb
                    self.band_gsm = band
                    self.tx_level_sweep_progress_gsm(plot=False)
                self.txp_aclr_evm_plot(self.filename, mode=0)

    def tx_level_sweep_pipeline_wcdma(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        for tech in wt.tech:
            if tech == 'WCDMA' and wt.wcdma_bands != []:
                self.tech = tech
                for tx_path in wt.tx_paths:
                    self.tx_path = tx_path
                    for band in wt.wcdma_bands:
                        self.band_wcdma = band
                        self.tx_level_sweep_progress_wcdma(plot=False)
                    self.txp_aclr_evm_plot(self.filename, mode=0)

    def tx_level_sweep_pipeline_lte(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel

        items = [
            (tech, tx_path, bw, band)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.lte_bandwidths
            for band in wt.lte_bands
        ]
        for item in items:
            if item[0] == 'LTE' and wt.lte_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_lte = item[2]
                self.band_lte = item[3]
                if self.bw_lte in cm_pmt_ftm.bandwidths_selected_lte(self.band_lte):
                    self.tx_level_sweep_progress_lte(plot=False)
                else:
                    logger.info(f'B{self.band_lte} does not have BW {self.bw_lte}MHZ')
        for bw in wt.lte_bandwidths:
            try:
                self.filename = f'Tx_level_sweep_{bw}MHZ_{self.tech}.xlsx'
                self.txp_aclr_evm_plot(self.filename, mode=0)
            except TypeError:
                logger.info(f'there is no data to plot because the band does not have this BW ')
            except FileNotFoundError:
                logger.info(f'there is not file to plot BW{bw} ')

    def tx_level_sweep_pipeline_fr1(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.sa_nsa_mode = wt.sa_nsa
        items = [
            (tech, tx_path, bw, band, type_)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.fr1_bandwidths
            for band in wt.fr1_bands
            for type_ in wt.type_fr1
        ]

        for item in items:
            if item[0] == 'FR1' and wt.fr1_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_fr1 = item[2]
                self.band_fr1 = item[3]
                self.type_fr1 = item[4]
                if self.bw_fr1 in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                    self.tx_level_sweep_progress_fr1(plot=False)
                else:
                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')
        for bw in wt.fr1_bandwidths:
            try:
                self.filename = f'Tx_level_sweep_{bw}MHZ_{self.tech}.xlsx'
                self.txp_aclr_evm_plot(self.filename, mode=0)
            except TypeError:
                logger.info(f'there is no data to plot because the band does not have this BW ')
            except FileNotFoundError:
                logger.info(f'there is not file to plot BW{bw} ')

    def tx_1rb_sweep_pipeline_fr1(self):
        self.rx_level = wt.init_rx_sync_level
        self.tx_level = wt.tx_level
        self.port_tx = wt.port_tx
        self.chan = wt.channel
        self.sa_nsa_mode = wt.sa_nsa
        self.tx_1rb_filename_judge = True
        items = [
            (tech, tx_path, bw, band, type_)
            for tech in wt.tech
            for tx_path in wt.tx_paths
            for bw in wt.fr1_bandwidths
            for band in wt.fr1_bands
            for type_ in wt.type_fr1
        ]

        for item in items:
            if item[0] == 'FR1' and wt.fr1_bands != []:
                self.tech = item[0]
                self.tx_path = item[1]
                self.bw_fr1 = item[2]
                self.band_fr1 = item[3]
                self.type_fr1 = item[4]
                if self.bw_fr1 in cm_pmt_ftm.bandwidths_selected_fr1(self.band_fr1):
                    self.tx_1rb_sweep_progress_fr1(plot=False)
                else:
                    logger.info(f'B{self.band_fr1} does not have BW {self.bw_fr1}MHZ')
        for bw in wt.fr1_bandwidths:
            try:
                self.filename = f'Tx_1RB_sweep_{bw}MHZ_{self.tech}.xlsx'
                self.txp_aclr_evm_plot(self.filename, mode=0)
            except TypeError:
                logger.info(f'there is no data to plot because the band does not have this BW ')
            except FileNotFoundError:
                logger.info(f'there is not file to plot BW{bw} ')
        self.tx_1rb_filename_judge = False

    def tx_measure_single(self):  # this is incompleted
        if 'LTE' in wt.tech:
            self.port_tx = wt.port_tx
            self.band_lte = wt.band_lte
            self.bw_lte = wt.bw_lte
            self.chan = wt.channel
            rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte, self.bw_lte)
            self.chan_single_lte = None
            if self.chan_singl_lte == 'L':
                self.tx_freq_lte = cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[0])
            elif self.chan_single == 'M':
                self.tx_freq_lte_lte = cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[1])
            elif self.chan_single == 'H':
                self.tx_freq_lte_lte = cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[2])
            self.rx_freq_lte = cm_pmt_ftm.transfer_freq_tx2rx_lte(self.band_lte, self.bw_lte)
            self.chan = wt.channel
            self.mcs_lte = wt.mcs_lte  # PRB, FRB
            self.rb_size_lte, self.rb_start_lte = scripts.GENERAL_LTE[self.bw_lte][
                self.rb_select_lte_dict[self.rb_state]]
            self.tx_level = wt.tx_level
            self.loss_tx = get_loss(self.tx_freq_lte)
            self.loss_rx = get_loss(self.rx_freq_lte)
            self.preset_instrument()
            self.set_test_end_lte()
            self.antenna_switch_v2()
            self.set_test_mode_lte()
            self.command_cmw100_query('*OPC?')
            self.sig_gen_lte()
            self.sync_lte()
            self.tx_set_lte()
            self.tx_measure_lte()

    def power_init_gsm(self):
        if self.band_gsm in [850, 900]:
            self.pwr_init_gsm = 33 - 2 * (self.pcl - 5)
        elif self.band_gsm in [1800, 1900]:
            self.pwr_init_gsm = 30 - 2 * (self.pcl - 0)

    def tx_power_aclr_evm_lmh_gsm(self, plot=True):
        """
                order: tx_path > band > chan
                band_gsm:
                tx_pcl:
                rf_port:
                freq_select: 'LMH'
                tx_path:
                data: {rx_freq: [power, phase_err_rms, phase_peak, ferr,orfs_mod_-200,orfs_mod_200,...orfs_sw-400,orfs_sw400,...], ...}
        """
        rx_chan_list = cm_pmt_ftm.dl_chan_select_gsm(self.band_gsm)

        rx_chan_select_list = []
        for chan in self.chan:
            if chan == 'L':
                rx_chan_select_list.append(rx_chan_list[0])
            elif chan == 'M':
                rx_chan_select_list.append(rx_chan_list[1])
            elif chan == 'H':
                rx_chan_select_list.append(rx_chan_list[2])

        self.preset_instrument()
        self.set_test_mode_gsm()
        self.set_test_end_gsm()

        for script in wt.scripts:
            if script == 'GENERAL':
                self.script = script
                data_chan = {}
                for rx_chan_gsm in rx_chan_select_list:
                    logger.info('----------Test LMH progress---------')
                    self.rx_chan_gsm = rx_chan_gsm
                    self.rx_freq_gsm = cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, self.rx_chan_gsm, 'rx')
                    self.tx_freq_gsm = cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, self.rx_chan_gsm, 'tx')
                    self.loss_rx = get_loss(self.rx_freq_gsm)
                    self.loss_tx = get_loss(self.tx_freq_gsm)
                    self.set_test_mode_gsm()
                    self.antenna_switch_v2()
                    self.sig_gen_gsm()
                    self.sync_gsm()
                    self.tx_set_gsm()
                    aclr_mod_current_results = aclr_mod_results = self.tx_measure_gsm()
                    logger.debug(aclr_mod_results)
                    aclr_mod_current_results.append(self.measure_current())
                    data_chan[self.rx_freq_gsm] = aclr_mod_current_results + self.get_temperature()
                logger.debug(data_chan)
                # ready to export to excel
                self.filename = self.tx_power_relative_test_export_excel(data_chan, self.band_gsm, 0,
                                                                         self.pcl,
                                                                         mode=1)  # mode=1: LMH mode
        self.set_test_end_gsm()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        else:
            pass

    def tx_power_aclr_evm_lmh_wcdma(self, plot=True):
        """
                order: tx_path > bw > band > mcs > rb > chan
                band_wcdma:
                tx_level:
                rf_port:
                freq_select: 'LMH'
                tx_path:
                data: {tx_level: [Pwr, U_-1, U_+1, U_-2, U_+2, OBW, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_chan_list = cm_pmt_ftm.dl_chan_select_wcdma(self.band_wcdma)
        tx_chan_list = [cm_pmt_ftm.transfer_chan_rx2tx_wcdma(self.band_wcdma, rx_chan) for rx_chan in rx_chan_list]
        tx_rx_chan_list = list(zip(tx_chan_list, rx_chan_list))  # [(tx_chan, rx_chan),...]

        tx_rx_chan_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_rx_chan_select_list.append(tx_rx_chan_list[0])
            elif chan == 'M':
                tx_rx_chan_select_list.append(tx_rx_chan_list[1])
            elif chan == 'H':
                tx_rx_chan_select_list.append(tx_rx_chan_list[2])

        self.preset_instrument()

        for script in wt.scripts:
            if script == 'GENERAL':
                self.script = script
                data_chan = {}
                for tx_rx_chan_wcdma in tx_rx_chan_select_list:
                    logger.info('----------Test LMH progress---------')
                    self.rx_chan_wcdma = tx_rx_chan_wcdma[1]
                    self.tx_chan_wcdma = tx_rx_chan_wcdma[0]
                    self.rx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.rx_chan_wcdma, 'rx')
                    self.tx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.tx_chan_wcdma, 'tx')
                    self.loss_rx = get_loss(self.rx_freq_wcdma)
                    self.loss_tx = get_loss(self.tx_freq_wcdma)
                    self.set_test_end_wcdma()
                    self.set_test_mode_wcdma()
                    self.command_cmw100_query('*OPC?')
                    self.sig_gen_wcdma()
                    self.sync_wcdma()
                    self.tx_chan_wcdma = tx_rx_chan_wcdma[0]
                    self.tx_set_wcdma()
                    self.antenna_switch_v2()
                    aclr_mod_current_results = aclr_mod_results = self.tx_measure_wcdma()
                    logger.debug(aclr_mod_results)
                    aclr_mod_current_results.append(self.measure_current())
                    data_chan[
                        cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma,
                                                            self.tx_chan_wcdma)] = aclr_mod_current_results + self.get_temperature()
                logger.debug(data_chan)
                # ready to export to excel
                self.filename = self.tx_power_relative_test_export_excel(data_chan, self.band_wcdma, 5,
                                                                         self.tx_level,
                                                                         mode=1)  # mode=1: LMH mode
        self.set_test_end_wcdma()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        else:
            pass

    def tx_power_aclr_evm_lmh_lte(self, plot=True):
        """
        order: tx_path > bw > band > mcs > rb > chan
        band_lte:
        bw_lte:
        tx_level:
        rf_port:
        freq_select: 'LMH'
        tx_path:
        data: {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte,
                                                   self.bw_lte)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        self.rx_freq_lte = rx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        logger.info('----------Test LMH progress---------')
        self.preset_instrument()
        self.set_test_end_lte()
        self.set_test_mode_lte()
        self.antenna_switch_v2()
        self.sig_gen_lte()
        self.sync_lte()

        tx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[0]))
            elif chan == 'M':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[1]))
            elif chan == 'H':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[2]))

        for mcs in wt.mcs_lte:
            self.mcs_lte = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_lte:  # PRB, FRB
                        self.rb_size_lte, self.rb_start_lte = scripts.GENERAL_LTE[self.bw_lte][
                            self.rb_select_lte_dict[rb_ftm]]  # PRB: 0, # FRB: 1
                        self.rb_state = rb_ftm  # PRB, FRB
                        data_freq = {}
                        for tx_freq_lte in tx_freq_select_list:
                            self.tx_freq_lte = tx_freq_lte
                            self.loss_tx = get_loss(self.tx_freq_lte)
                            self.tx_set_lte()
                            aclr_mod_current_results = aclr_mod_results = self.tx_measure_lte()
                            logger.debug(aclr_mod_results)
                            aclr_mod_current_results.append(self.measure_current())
                            data_freq[self.tx_freq_lte] = aclr_mod_current_results + self.get_temperature()
                        logger.debug(data_freq)
                        # ready to export to excel
                        self.filename = self.tx_power_relative_test_export_excel(data_freq, self.band_lte, self.bw_lte,
                                                                                 self.tx_level,
                                                                                 mode=1)  # mode=1: LMH mode
        self.set_test_end_lte()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        else:
            pass

    def tx_power_aclr_evm_lmh_fr1(self, plot=True):
        """
        order: tx_path > bw > band > mcs > rb > chan
        band_fr1:
        bw_fr1:
        tx_level:
        rf_port:
        freq_select: 'LMH'
        tx_path:
        data: {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1,
                                                   self.bw_fr1)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        # self.rx_freq_fr1 = rx_freq_list[1]
        # self.loss_rx = get_loss(rx_freq_list[1])
        logger.info('----------Test LMH progress---------')
        self.preset_instrument()
        # self.set_test_end_fr1()
        # self.set_test_mode_fr1()
        # if self.srs_path_enable:
        #     self.srs_switch()
        # else:
        #     self.antenna_switch_v2()
        # self.sig_gen_fr1()
        # self.sync_fr1()

        scs = 1 if self.band_fr1 in [34, 38, 39, 40, 41, 42, 48, 77, 78,  # temp
                                     79] else 0  # for now FDD is forced to 15KHz and TDD is to be 30KHz  # temp
        scs = 15 * (2 ** scs)  # temp
        self.scs = scs  # temp

        tx_freq_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[0]))
            elif chan == 'M':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[1]))
            elif chan == 'H':
                tx_freq_select_list.append(cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[2]))

        for mcs in wt.mcs_fr1:
            self.mcs_fr1 = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_fr1:  # INNER_FULL, OUTER_FULL
                        self.rb_size_fr1, self.rb_start_fr1 = scripts.GENERAL_FR1[self.bw_fr1][self.scs][self.type_fr1][
                            self.rb_alloc_fr1_dict[rb_ftm]]
                        self.rb_state = rb_ftm  # INNER_FULL, OUTER_FULL
                        data_freq = {}
                        for tx_freq_fr1 in tx_freq_select_list:
                            self.tx_freq_fr1 = tx_freq_fr1
                            self.rx_freq_fr1 = cm_pmt_ftm.transfer_freq_tx2rx_fr1(self.band_fr1, tx_freq_fr1)  # temp
                            self.loss_tx = get_loss(self.tx_freq_fr1)
                            self.loss_rx = get_loss(rx_freq_list[1])  # temp
                            self.set_test_end_fr1()  # temp
                            self.set_test_mode_fr1()  # temp
                            if self.srs_path_enable:  # temp
                                self.srs_switch()  # temp
                            else:  # temp
                                self.antenna_switch_v2()  # temp
                            self.sig_gen_fr1()  # temp
                            self.sync_fr1()  # temp
                            self.tx_set_fr1()
                            aclr_mod_current_results = aclr_mod_results = self.tx_measure_fr1()
                            logger.debug(aclr_mod_results)
                            aclr_mod_current_results.append(self.measure_current())
                            data_freq[self.tx_freq_fr1] = aclr_mod_current_results + self.get_temperature()
                        logger.debug(data_freq)
                        # ready to export to excel
                        self.filename = self.tx_power_relative_test_export_excel(data_freq, self.band_fr1, self.bw_fr1,
                                                                                 self.tx_level,
                                                                                 mode=1)  # mode=1: LMH mode
        self.set_test_end_fr1()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        else:
            pass

    def tx_power_fcc_fr1(self):
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1,
                                                   self.bw_fr1)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        self.rx_freq_fr1 = rx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        logger.info('----------Test FCC LMH progress---------')
        self.preset_instrument()
        self.set_gprf_measurement()
        self.set_test_end_fr1()
        self.set_test_mode_fr1()
        if self.srs_path_enable:
            self.srs_switch()
        else:
            self.antenna_switch_v2()
        self.command_cmw100_query('*OPC?')

        tx_freq_select_list = []
        try:
            tx_freq_select_list = [int(freq * 1000) for freq in
                                   fcc.tx_freq_list_fr1[self.band_fr1][self.bw_fr1]]  # band > bw > tx_fre1_list
        except KeyError as err:
            logger.info(f"this Band: {err} don't have to  test this BW: {self.bw_fr1} for FCC")

        for mcs in wt.mcs_fr1:
            self.mcs_fr1 = mcs
            for script in wt.scripts:
                if script == 'FCC':
                    self.script = script
                    self.rb_state = 'FCC'
                    try:
                        for self.rb_size_fr1, self.rb_start_fr1 in scripts.FCC_FR1[self.band_fr1][self.bw_fr1][self.mcs_fr1]:
                            data = {}
                            for num, tx_freq_fr1 in enumerate(tx_freq_select_list):
                                chan_mark = f'chan{num}'
                                self.tx_freq_fr1 = tx_freq_fr1
                                self.loss_tx = get_loss(self.tx_freq_fr1)
                                self.set_gprf_tx_freq()
                                self.set_duty_cycle()
                                self.tx_set_no_sync_fr1()
                                power_results = self.get_gprf_power()
                                data[self.tx_freq_fr1] = (
                                    chan_mark, power_results)  # data = {tx_freq:(chan_mark, power)}
                            logger.debug(data)
                            # ready to export to excel
                            self.filename = self.tx_power_fcc_ce_export_excel(data)
                    except KeyError as err:
                        logger.debug(f'show error: {err}')
                        logger.info(
                            f"Band {self.band_fr1}, BW: {self.bw_fr1} don't need to test this MCS: {self.mcs_fr1} for FCC")

        self.set_test_end_fr1()
        # if plot == True:
        #     self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        # else:
        #     pass

    def tx_power_ce_fr1(self):
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1,
                                                   self.bw_fr1)  # [L_rx_freq, M_rx_ferq, H_rx_freq]
        self.rx_freq_fr1 = rx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        logger.info('----------Test CE LMH progress---------')
        self.preset_instrument()
        self.set_gprf_measurement()
        self.set_test_end_fr1()
        self.set_test_mode_fr1()
        if self.srs_path_enable:
            self.srs_switch()
        else:
            self.antenna_switch_v2()
        self.command_cmw100_query('*OPC?')

        tx_freq_select_list = []
        try:
            tx_freq_select_list = [int(freq * 1000) for freq in
                                   ce.tx_freq_list_fr1[self.band_fr1][self.bw_fr1]]  # band > bw > tx_fre1_list
        except KeyError as err:
            logger.info(f"this Band: {err} don't have to  test this BW: {self.bw_fr1} for CE")

        for mcs in wt.mcs_fr1:
            self.mcs_fr1 = mcs
            for script in wt.scripts:
                if script == 'CE':
                    self.script = script
                    self.rb_state = 'CE'
                    try:
                        for self.rb_size_fr1, self.rb_start_fr1 in scripts.CE_FR1[self.band_fr1][self.bw_fr1][self.mcs_fr1]:
                            data = {}
                            for num, tx_freq_fr1 in enumerate(tx_freq_select_list):
                                chan_mark = f'chan{num}'
                                self.tx_freq_fr1 = tx_freq_fr1
                                self.loss_tx = get_loss(self.tx_freq_fr1)
                                self.set_gprf_tx_freq()
                                self.set_duty_cycle()
                                self.tx_set_no_sync_fr1()
                                power_results = self.get_gprf_power()
                                data[self.tx_freq_fr1] = (
                                    chan_mark, power_results)  # data = {tx_freq:(chan_mark, power)}
                            logger.debug(data)
                            # ready to export to excel
                            self.filename = self.tx_power_fcc_ce_export_excel(data)
                    except KeyError as err:
                        logger.debug(f'show error: {err}')
                        logger.info(
                            f"Band {self.band_fr1}, BW: {self.bw_fr1} don't need to test this MCS: {self.mcs_fr1} for CE")

        self.set_test_end_fr1()
        # if plot == True:
        #     self.txp_aclr_evm_plot(self.filename, mode=1)  # mode=1: LMH mode
        # else:
        #     pass

    def tx_freq_sweep_progress_gsm(self, plot=True):
        """
        band_gsm:
        tx_freq_gsm:
        tx_pcl:
        rf_port:
        freq_range_list: [freq_level_1, freq_level_2, freq_step]
        tx_path:
        data: {tx_freq: [power, phase_err_rms, phase_peak, ferr,orfs_mod_-200,orfs_mod_200,...orfs_sw-400,orfs_sw400,...], ...}
        """
        logger.info('----------Freq Sweep progress ---------')
        rx_chan_list = cm_pmt_ftm.dl_chan_select_gsm(self.band_gsm)

        self.preset_instrument()
        self.set_test_mode_gsm()
        self.set_test_end_gsm()

        rx_chan_range_list = [rx_chan_list[0], rx_chan_list[2], 0.2]
        step = int(rx_chan_range_list[2] * 1000 * 5)
        rx_freq_range_list = [cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, rx_chan) for rx_chan in
                              rx_chan_range_list[:2]]
        rx_freq_range_list.append(step)

        for script in wt.scripts:
            if script == 'GENERAL':
                self.script = script
                data = {}
                for rx_freq_gsm in range(rx_freq_range_list[0], rx_freq_range_list[1] + 1, step):
                    self.rx_freq_gsm = rx_freq_gsm
                    self.tx_freq_gsm = cm_pmt_ftm.transfer_freq_rx2tx_gsm(self.band_gsm, self.rx_freq_gsm)
                    self.rx_chan_gsm = cm_pmt_ftm.transfer_freq2chan_gsm(self.band_gsm, self.rx_freq_gsm)
                    self.loss_rx = get_loss(self.rx_freq_gsm)
                    self.loss_tx = get_loss(self.tx_freq_gsm)
                    self.set_test_mode_gsm()
                    self.antenna_switch_v2()
                    self.sig_gen_gsm()
                    self.sync_gsm()
                    self.tx_set_gsm()
                    aclr_mod_results = self.tx_measure_gsm()
                    logger.debug(aclr_mod_results)
                    data[self.rx_freq_gsm] = aclr_mod_results
                logger.debug(data)
                self.filename = self.tx_power_relative_test_export_excel(data, self.band_gsm, 0,
                                                                         self.pcl, mode=0)
        self.set_test_end_gsm()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_freq_sweep_progress_wcdma(self, plot=True):
        """
        band_lte:
        bw_lte:
        tx_freq_lte:
        rb_num:
        rb_start:
        mcs:
        tx_level:
        rf_port:
        freq_range_list: [freq_level_1, freq_level_2, freq_step]
        tx_path:
        data: {tx_freq: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        logger.info('----------Freq Sweep progress ---------')
        rx_chan_list = cm_pmt_ftm.dl_chan_select_wcdma(self.band_wcdma)
        tx_chan_list = [cm_pmt_ftm.transfer_chan_rx2tx_wcdma(self.band_wcdma, rx_chan) for rx_chan in rx_chan_list]

        self.preset_instrument()

        tx_chan_range_list = [tx_chan_list[0], tx_chan_list[2], 1]
        step = tx_chan_range_list[2]

        for script in wt.scripts:
            if script == 'GENERAL':
                self.script = script
                data = {}
                for tx_chan_wcdma in range(tx_chan_range_list[0], tx_chan_range_list[1] + step, step):
                    self.tx_chan_wcdma = tx_chan_wcdma
                    self.rx_chan_wcdma = cm_pmt_ftm.transfer_chan_tx2rx_wcdma(self.band_wcdma, tx_chan_wcdma)
                    self.tx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.tx_chan_wcdma, 'tx')
                    self.rx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.rx_chan_wcdma, 'rx')
                    self.loss_rx = get_loss(self.rx_freq_wcdma)
                    self.loss_tx = get_loss(self.tx_freq_wcdma)
                    self.set_test_end_wcdma()
                    self.set_test_mode_wcdma()
                    self.command_cmw100_query('*OPC?')
                    self.sig_gen_wcdma()
                    self.sync_wcdma()
                    # self.antenna_switch_v2()
                    self.tx_set_wcdma()
                    self.antenna_switch_v2()
                    aclr_mod_results = self.tx_measure_wcdma()
                    logger.debug(aclr_mod_results)
                    data[self.tx_chan_wcdma] = aclr_mod_results
                logger.debug(data)
                self.filename = self.tx_power_relative_test_export_excel(data, self.band_wcdma, 5,
                                                                         self.tx_level, mode=0)
        self.set_test_end_wcdma()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_freq_sweep_progress_lte(self, plot=True):
        """
        band_lte:
        bw_lte:
        tx_freq_lte:
        rb_num:
        rb_start:
        mcs:
        tx_level:
        rf_port:
        freq_range_list: [freq_level_1, freq_level_2, freq_step]
        tx_path:
        data: {tx_freq: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        logger.info('----------Freq Sweep progress ---------')
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte, self.bw_lte)
        tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq) for rx_freq in rx_freq_list]
        self.rx_freq_lte = rx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        self.preset_instrument()
        self.set_test_end_lte()
        self.set_test_mode_lte()
        self.antenna_switch_v2()
        self.sig_gen_lte()
        self.sync_lte()

        freq_range_list = [tx_freq_list[0], tx_freq_list[2], 1000]
        step = freq_range_list[2]

        for mcs in wt.mcs_lte:
            self.mcs_lte = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_lte:  # PRB, FRB
                        self.rb_size_lte, self.rb_start_lte = scripts.GENERAL_LTE[self.bw_lte][
                            self.rb_select_lte_dict[rb_ftm]]  # PRB: 0, # FRB: 1
                        self.rb_state = rb_ftm  # PRB, FRB
                        data = {}
                        for tx_freq_lte in range(freq_range_list[0], freq_range_list[1] + step, step):
                            self.tx_freq_lte = tx_freq_lte
                            self.loss_tx = get_loss(self.tx_freq_lte)
                            self.tx_set_lte()
                            aclr_mod_results = self.tx_measure_lte()
                            logger.debug(aclr_mod_results)
                            data[self.tx_freq_lte] = aclr_mod_results
                        logger.debug(data)
                        self.filename = self.tx_power_relative_test_export_excel(data, self.band_lte, self.bw_lte,
                                                                                 self.tx_level, mode=0)
        self.set_test_end_lte()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_freq_sweep_progress_fr1(self, plot=True):
        """
        band_fr1:
        bw_fr1:
        tx_freq_fr1:
        rb_num:
        rb_start:
        mcs:
        tx_level:
        rf_port:
        freq_range_list: [freq_level_1, freq_level_2, freq_step]
        tx_path:
        data: {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        logger.info('----------Freq Sweep progress ---------')
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1, self.bw_fr1)
        tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq) for rx_freq in rx_freq_list]
        self.rx_freq_fr1 = rx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        self.preset_instrument()
        self.set_test_end_fr1()
        self.set_test_mode_fr1()
        if self.srs_path_enable:
            self.srs_switch()
        else:
            self.antenna_switch_v2()
        self.sig_gen_fr1()
        self.sync_fr1()

        freq_range_list = [tx_freq_list[0], tx_freq_list[2], 1000]
        step = freq_range_list[2]

        for mcs in wt.mcs_fr1:
            self.mcs_fr1 = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_fr1:  # PRB, FRB
                        self.rb_size_fr1, self.rb_start_fr1 = scripts.GENERAL_FR1[self.bw_fr1][self.scs][self.type_fr1][
                            self.rb_alloc_fr1_dict[rb_ftm]]  # PRB: 0, # FRB: 1
                        self.rb_state = rb_ftm  # PRB, FRB
                        data = {}
                        for tx_freq_fr1 in range(freq_range_list[0], freq_range_list[1] + step, step):
                            self.tx_freq_fr1 = tx_freq_fr1
                            self.loss_tx = get_loss(self.tx_freq_fr1)
                            self.tx_set_fr1()
                            aclr_mod_results = self.tx_measure_fr1()
                            logger.debug(aclr_mod_results)
                            data[self.tx_freq_fr1] = aclr_mod_results
                        logger.debug(data)
                        self.filename = self.tx_power_relative_test_export_excel(data, self.band_fr1, self.bw_fr1,
                                                                                 self.tx_level, mode=0)
        self.set_test_end_fr1()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_1rb_sweep_progress_fr1(self, plot=True):
        logger.info('----------1RB Sweep progress ---------')
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1, self.bw_fr1)
        # tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq) for rx_freq in rx_freq_list]
        self.rx_freq_fr1 = rx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        self.preset_instrument()
        self.set_test_end_fr1()
        self.set_test_mode_fr1()
        if self.srs_path_enable:
            self.srs_switch()
        else:
            self.antenna_switch_v2()
        self.sig_gen_fr1()
        self.sync_fr1()

        for mcs in wt.mcs_fr1:
            self.mcs_fr1 = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    tx_freq_select_list = []
                    for chan in self.chan:
                        if chan == 'L':
                            tx_freq_select_list.append(
                                cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[0]))
                        elif chan == 'M':
                            tx_freq_select_list.append(
                                cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[1]))
                        elif chan == 'H':
                            tx_freq_select_list.append(
                                cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[2]))

                    for tx_freq_select in tx_freq_select_list:
                        self.tx_freq_fr1 = tx_freq_select
                        self.rb_size_fr1, rb_sweep_fr1 = scripts.GENERAL_FR1[self.bw_fr1][self.scs][self.type_fr1][
                            self.rb_alloc_fr1_dict['EDGE_1RB_RIGHT']]  # capture EDGE_1RB_RIGHT
                        self.rb_state = '1rb_sweep'
                        data = {}
                        for rb_start in range(rb_sweep_fr1 + 1):
                            self.rb_start_fr1 = rb_start
                            self.loss_tx = get_loss(self.tx_freq_fr1)
                            self.tx_set_fr1()
                            aclr_mod_results = self.tx_measure_fr1()
                            logger.debug(aclr_mod_results)
                            data[self.tx_freq_fr1] = aclr_mod_results
                            logger.debug(data)
                            self.filename = self.tx_power_relative_test_export_excel(data, self.band_fr1, self.bw_fr1,
                                                                                     self.tx_level, mode=0)
        self.set_test_end_fr1()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_level_sweep_progress_gsm(self, plot=True):
        """
        band_gsm:
        tx_freq_gsm:
        pwr:
        rf_port:
        loss:
        tx_path:
        data {tx_pcl: [power, phase_err_rms, phase_peak, ferr,orfs_mod_-200,orfs_mod_200,...orfs_sw-400,orfs_sw400,...], ...}
        """
        rx_chan_list = cm_pmt_ftm.dl_chan_select_gsm(self.band_gsm)

        rx_chan_select_list = []
        for chan in self.chan:
            if chan == 'L':
                rx_chan_select_list.append(rx_chan_list[0])
            elif chan == 'M':
                rx_chan_select_list.append(rx_chan_list[1])
            elif chan == 'H':
                rx_chan_select_list.append(rx_chan_list[2])

        self.preset_instrument()
        self.set_test_mode_gsm()
        self.set_test_end_gsm()

        for script in wt.scripts:
            if script == 'GENERAL':
                self.script = script
                #  initial all before tx level prgress
                for rx_chan_gsm in rx_chan_select_list:
                    self.rx_chan_gsm = rx_chan_gsm
                    self.rx_freq_gsm = cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, self.rx_chan_gsm, 'rx')
                    self.tx_freq_gsm = cm_pmt_ftm.transfer_chan2freq_gsm(self.band_gsm, self.rx_chan_gsm, 'tx')
                    self.loss_rx = get_loss(self.rx_freq_gsm)
                    self.loss_tx = get_loss(self.tx_freq_gsm)
                    self.set_test_mode_gsm()
                    self.antenna_switch_v2()
                    self.sig_gen_gsm()
                    self.sync_gsm()

                    # self.tx_power_relative_test_initial_gsm()

                    tx_range_list = wt.tx_pcl_range_list_lb if self.band_gsm in [850,
                                                                                 900] else wt.tx_pcl_range_list_mb  # [tx_pcl_1, tx_pcl_2]

                    logger.info('----------TX Level Sweep progress---------')
                    logger.info(f'----------from PCL{tx_range_list[0]} to PCL{tx_range_list[1]}----------')

                    step = -1 if tx_range_list[0] > tx_range_list[1] else 1

                    #  following is real change tx pcl prgress

                    data = {}
                    for tx_pcl in range(tx_range_list[0], tx_range_list[1] + step, step):
                        self.pcl = tx_pcl
                        logger.info(f'========Now Tx PCL = PCL{self.pcl} ========')
                        self.tx_set_gsm()
                        mod_orfs_current_results = mod_orfs_results = self.tx_measure_gsm()
                        logger.debug(mod_orfs_results)
                        mod_orfs_current_results.append(self.measure_current())
                        data[tx_pcl] = mod_orfs_current_results
                    logger.debug(data)
                    self.filename = self.tx_power_relative_test_export_excel(data, self.band_gsm, 0,
                                                                             self.rx_freq_gsm)
        self.set_test_end_gsm()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_level_sweep_progress_wcdma(self, plot=True):
        """
        band_wcdma:
        bw_wcdma:
        tx_freq_wcdma:
        rb_num:
        rb_start:
        mcs:
        pwr:
        rf_port:
        loss:
        tx_path:
        data {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_chan_list = cm_pmt_ftm.dl_chan_select_wcdma(self.band_wcdma)
        tx_chan_list = [cm_pmt_ftm.transfer_chan_rx2tx_wcdma(self.band_wcdma, rx_chan) for rx_chan in rx_chan_list]
        tx_rx_chan_list = list(zip(tx_chan_list, rx_chan_list))  # [(tx_chan, rx_chan),...]

        tx_rx_chan_select_list = []
        for chan in self.chan:
            if chan == 'L':
                tx_rx_chan_select_list.append(tx_rx_chan_list[0])
            elif chan == 'M':
                tx_rx_chan_select_list.append(tx_rx_chan_list[1])
            elif chan == 'H':
                tx_rx_chan_select_list.append(tx_rx_chan_list[2])

        self.preset_instrument()

        for script in wt.scripts:
            if script == 'GENERAL':
                self.script = script
                #  initial all before tx level prgress
                for tx_rx_chan_wcdma in tx_rx_chan_select_list:
                    self.rx_chan_wcdma = tx_rx_chan_wcdma[1]
                    self.tx_chan_wcdma = tx_rx_chan_wcdma[0]
                    self.rx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.rx_chan_wcdma, 'rx')
                    self.tx_freq_wcdma = cm_pmt_ftm.transfer_chan2freq_wcdma(self.band_wcdma, self.tx_chan_wcdma, 'tx')
                    self.loss_rx = get_loss(self.rx_freq_wcdma)
                    self.loss_tx = get_loss(self.tx_freq_wcdma)
                    self.set_test_end_wcdma()
                    self.set_test_mode_wcdma()
                    self.command_cmw100_query('*OPC?')
                    self.sig_gen_wcdma()
                    self.sync_wcdma()

                    # self.tx_power_relative_test_initial_wcdma()

                    tx_range_list = wt.tx_level_range_list  # [tx_level_1, tx_level_2]

                    logger.info('----------TX Level Sweep progress---------')
                    logger.info(f'----------from {tx_range_list[0]} dBm to {tx_range_list[1]} dBm----------')

                    step = -1 if tx_range_list[0] > tx_range_list[1] else 1

                    #  following is real change tx level prgress

                    data = {}
                    for tx_level in range(tx_range_list[0], tx_range_list[1] + step, step):
                        self.tx_level = tx_level
                        logger.info(f'========Now Tx level = {self.tx_level} dBm========')
                        self.tx_set_wcdma()
                        # self.command(f'AT+HTXPERSTART={self.tx_chan_wcdma}')
                        # self.command(f'AT+HSETMAXPOWER={self.tx_level * 10}')
                        self.antenna_switch_v2()
                        spectrum_mod_current_results = spectrum_mod_results = self.tx_measure_wcdma()
                        # self.command_cmw100_write(f'CONF:WCDMA:MEAS:RFS:UMAR 10.00')
                        # self.command_cmw100_write(f'CONF:WCDM:MEAS:RFS:ENP {self.tx_level + 5}')
                        # mod_results = self.command_cmw100_query(
                        #     f'READ:WCDM:MEAS:MEV:MOD:AVER?')  # P1 is EVM, P4 is Ferr, P8 is IQ Offset
                        # mod_results = mod_results.split(',')
                        # mod_results = [mod_results[1], mod_results[4], mod_results[8]]
                        # mod_results = [eval(m) for m in mod_results]
                        # logger.info(
                        #     f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
                        # self.command_cmw100_write(f'INIT:WCDM:MEAS:MEV')
                        # self.command_cmw100_query(f'*OPC?')
                        # f_state = self.command_cmw100_query(f'FETC:WCDM:MEAS:MEV:STAT?')
                        # while f_state != 'RDY':
                        #     time.sleep(0.2)
                        #     f_state = self.command_cmw100_query('FETC:WCDM:MEAS:MEV:STAT?')
                        #     self.command_cmw100_query('*OPC?')
                        # spectrum_results = self.command_cmw100_query(
                        #     f'FETC:WCDM:MEAS:MEV:SPEC:AVER?')  # P1: Power, P2: ACLR_-2, P3: ACLR_-1, P4:ACLR_+1, P5:ACLR_+2, P6:OBW
                        # spectrum_results = spectrum_results.split(',')
                        # spectrum_results = [
                        #     round(eval(spectrum_results[1]), 2),
                        #     round(eval(spectrum_results[3]) - eval(spectrum_results[1]), 2),
                        #     round(eval(spectrum_results[4]) - eval(spectrum_results[1]), 2),
                        #     round(eval(spectrum_results[2]) - eval(spectrum_results[1]), 2),
                        #     round(eval(spectrum_results[5]) - eval(spectrum_results[1]), 2),
                        #     round(eval(spectrum_results[6]) / 1000000, 2)
                        # ]
                        # logger.info(
                        #     f'Power: {spectrum_results[0]:.2f}, ACLR_-1: {spectrum_results[2]:.2f}, ACLR_1: {spectrum_results[3]:.2f}, ACLR_-2: {spectrum_results[1]:.2f}, ACLR_+2: {spectrum_results[4]:.2f}, OBW: {spectrum_results[5]:.2f}MHz')
                        # self.command_cmw100_write(f'STOP:WCDM:MEAS:MEV')
                        # self.command_cmw100_query(f'*OPC?')

                        logger.debug(spectrum_mod_results)
                        spectrum_mod_current_results.append(self.measure_current())
                        data[tx_level] = spectrum_mod_current_results
                    logger.debug(data)
                    self.filename = self.tx_power_relative_test_export_excel(data, self.band_wcdma, 5,
                                                                             self.tx_freq_wcdma)
        self.set_test_end_wcdma()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_level_sweep_progress_lte(self, plot=True):
        """
        band_lte:
        bw_lte:
        tx_freq_lte:
        rb_num:
        rb_start:
        mcs:
        pwr:
        rf_port:
        loss:
        tx_path:
        data {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('LTE', self.band_lte, self.bw_lte)
        tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq) for rx_freq in rx_freq_list]
        self.rx_freq_lte = rx_freq_list[1]
        self.tx_freq_lte = tx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        self.preset_instrument()
        self.set_test_end_lte()
        self.set_test_mode_lte()
        self.antenna_switch_v2()
        self.command_cmw100_query('*OPC?')
        self.sig_gen_lte()
        self.sync_lte()

        for mcs in wt.mcs_lte:
            self.mcs_lte = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_lte:  # PRB, FRB
                        self.rb_size_lte, self.rb_start_lte = scripts.GENERAL_LTE[self.bw_lte][
                            self.rb_select_lte_dict[rb_ftm]]  # PRB: 0, # FRB: 1
                        self.rb_state = rb_ftm  # PRB, FRB

                        tx_freq_select_list = []
                        for chan in self.chan:
                            if chan == 'L':
                                tx_freq_select_list.append(
                                    cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[0]))
                            elif chan == 'M':
                                tx_freq_select_list.append(
                                    cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[1]))
                            elif chan == 'H':
                                tx_freq_select_list.append(
                                    cm_pmt_ftm.transfer_freq_rx2tx_lte(self.band_lte, rx_freq_list[2]))

                        #  initial all before tx level prgress
                        for tx_freq_select in tx_freq_select_list:
                            self.tx_freq_lte = tx_freq_select
                            self.loss_tx = get_loss(self.tx_freq_lte)
                            self.tx_set_lte()
                            self.tx_power_relative_test_initial_lte()

                            tx_range_list = wt.tx_level_range_list  # [tx_level_1, tx_level_2]

                            logger.info('----------TX Level Sweep progress---------')
                            logger.info(f'----------from {tx_range_list[0]} dBm to {tx_range_list[1]} dBm----------')

                            step = -1 if tx_range_list[0] > tx_range_list[1] else 1

                            #  following is real change tx level prgress

                            data = {}
                            for tx_level in range(tx_range_list[0], tx_range_list[1] + step, step):
                                self.tx_level = tx_level
                                logger.info(f'========Now Tx level = {self.tx_level} dBm========')
                                self.command(f'AT+LTXPWRLVLSET={self.tx_level}')
                                self.command(f'AT+LTXCHNSDREQ')
                                self.command_cmw100_write('CONF:LTE:MEAS:RFS:UMAR 10.000000')
                                self.command_cmw100_write(f'CONF:LTE:MEAS:RFS:ENP {self.tx_level + 5}.00')
                                mod_results = self.command_cmw100_query(f'READ:LTE:MEAS:MEV:MOD:AVER?')
                                mod_results = mod_results.split(',')
                                mod_results = [mod_results[3], mod_results[15], mod_results[14]]
                                mod_results = [eval(m) for m in mod_results]
                                # logger.info(f'mod_results = {mod_results}')
                                logger.info(
                                    f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
                                self.command_cmw100_write('INIT:LTE:MEAS:MEV')
                                self.command_cmw100_query('*OPC?')
                                f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
                                while f_state != 'RDY':
                                    f_state = self.command_cmw100_query('FETC:LTE:MEAS:MEV:STAT?')
                                    self.command_cmw100_query('*OPC?')
                                aclr_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:ACLR:AVER?')
                                aclr_results = aclr_results.split(',')[1:]
                                aclr_results = [eval(aclr) * -1 if eval(aclr) > 30 else eval(aclr) for aclr in
                                                aclr_results]  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2
                                logger.info(
                                    f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
                                iem_results = self.command_cmw100_query('FETC:LTE:MEAS:MEV:IEM:MARG?')
                                iem_results = iem_results.split(',')
                                logger.info(f'InBandEmissions Margin: {eval(iem_results[2]):.2f}dB')
                                # logger.info(f'IEM_MARG results: {iem_results}')
                                esfl_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:ESFL:EXTR?')
                                esfl_results = esfl_results.split(',')
                                ripple1 = f'{eval(esfl_results[2]):.2f}' if esfl_results[2] != 'NCAP' else 'NCAP'
                                ripple2 = f'{eval(esfl_results[3]):.2f}' if esfl_results[3] != 'NCAP' else 'NCAP'
                                logger.info(
                                    f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
                                time.sleep(0.2)
                                # logger.info(f'ESFL results: {esfl_results}')
                                sem_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:MARG?')
                                logger.info(f'SEM_MARG results: {sem_results}')
                                sem_avg_results = self.command_cmw100_query(f'FETC:LTE:MEAS:MEV:SEM:AVER?')
                                sem_avg_results = sem_avg_results.split(',')
                                logger.info(
                                    f'OBW: {eval(sem_avg_results[2]) / 1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
                                # logger.info(f'SEM_AVER results: {sem_avg_results}')
                                self.command_cmw100_write(f'STOP:LTE:MEAS:MEV')
                                self.command_cmw100_query('*OPC?')
                                aclr_mod_current_results = aclr_mod_results = aclr_results + mod_results
                                logger.debug(aclr_mod_results)
                                aclr_mod_current_results.append(self.measure_current())
                                data[tx_level] = aclr_mod_current_results
                            logger.debug(data)
                            self.filename = self.tx_power_relative_test_export_excel(data, self.band_lte, self.bw_lte,
                                                                                     self.tx_freq_lte)
        self.set_test_end_lte()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def tx_level_sweep_progress_fr1(self, plot=True):
        """
        band_fr1:
        bw_fr1:
        tx_freq_fr1:
        rb_num:
        rb_start:
        mcs:
        pwr:
        rf_port:
        loss:
        tx_path:
        data {tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET], ...}
        """
        rx_freq_list = cm_pmt_ftm.dl_freq_selected('FR1', self.band_fr1, self.bw_fr1)
        tx_freq_list = [cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq) for rx_freq in rx_freq_list]
        self.rx_freq_fr1 = rx_freq_list[1]
        self.tx_freq_fr1 = tx_freq_list[1]
        self.loss_rx = get_loss(rx_freq_list[1])
        self.preset_instrument()
        self.set_test_end_fr1()
        self.set_test_mode_fr1()
        if self.srs_path_enable:
            self.srs_switch()
        else:
            self.antenna_switch_v2()
        self.sig_gen_fr1()
        self.sync_fr1()

        for mcs in wt.mcs_fr1:
            self.mcs_fr1 = mcs
            for script in wt.scripts:
                if script == 'GENERAL':
                    self.script = script
                    for rb_ftm in wt.rb_ftm_fr1:  # INNER, OUTER
                        self.rb_size_fr1, self.rb_start_fr1 = scripts.GENERAL_FR1[self.bw_fr1][self.scs][self.type_fr1][
                            self.rb_alloc_fr1_dict[rb_ftm]]  # INNER: 0, # OUTER: 1
                        self.rb_state = rb_ftm  # INNER, OUTER

                        tx_freq_select_list = []
                        for chan in self.chan:
                            if chan == 'L':
                                tx_freq_select_list.append(
                                    cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[0]))
                            elif chan == 'M':
                                tx_freq_select_list.append(
                                    cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[1]))
                            elif chan == 'H':
                                tx_freq_select_list.append(
                                    cm_pmt_ftm.transfer_freq_rx2tx_fr1(self.band_fr1, rx_freq_list[2]))

                        #  initial all before tx level prgress
                        for tx_freq_select in tx_freq_select_list:
                            self.tx_freq_fr1 = tx_freq_select
                            self.loss_tx = get_loss(self.tx_freq_fr1)
                            self.tx_set_fr1()
                            self.tx_power_relative_test_initial_fr1()

                            tx_range_list = wt.tx_level_range_list  # [tx_level_1, tx_level_2]

                            logger.info('----------TX Level Sweep progress---------')
                            logger.info(f'----------from {tx_range_list[0]} dBm to {tx_range_list[1]} dBm----------')

                            step = -1 if tx_range_list[0] > tx_range_list[1] else 1

                            #  following is real change tx level prgress

                            data = {}
                            for tx_level in range(tx_range_list[0], tx_range_list[1] + step, step):
                                self.tx_level = tx_level
                                logger.info(f'========Now Tx level = {self.tx_level} dBm========')
                                self.command(f'AT+NTXPWRLVLSET={self.tx_level}')
                                self.command_cmw100_write('CONF:NRS:MEAS:RFS:UMAR 10.000000')
                                self.command_cmw100_write(f'CONF:NRS:MEAS:RFS:ENP {self.tx_level + 5}.00')
                                self.command_cmw100_write(f'INIT:NRS:MEAS:MEV')
                                self.command_cmw100_query('*OPC?')
                                f_state = self.command_cmw100_query('FETC:NRS:MEAS:MEV:STAT?')
                                while f_state != 'RDY':
                                    f_state = self.command_cmw100_query('FETC:NRS:MEAS:MEV:STAT?')
                                    self.command_cmw100_query('*OPC?')
                                mod_results = self.command_cmw100_query(
                                    f'FETC:NRS:MEAS:MEV:MOD:AVER?')  # P3 is EVM, P15 is Ferr, P14 is IQ Offset
                                mod_results = mod_results.split(',')
                                mod_results = [mod_results[3], mod_results[15], mod_results[14]]
                                mod_results = [eval(m) for m in mod_results]
                                # logger.info(f'mod_results = {mod_results}')
                                logger.info(
                                    f'EVM: {mod_results[0]:.2f}, FREQ_ERR: {mod_results[1]:.2f}, IQ_OFFSET: {mod_results[2]:.2f}')
                                aclr_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:ACLR:AVER?')
                                aclr_results = aclr_results.split(',')[1:]
                                aclr_results = [eval(aclr) * -1 if eval(aclr) > 30 else eval(aclr) for aclr in
                                                aclr_results]  # U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2
                                logger.info(
                                    f'Power: {aclr_results[3]:.2f}, E-UTRA: [{aclr_results[2]:.2f}, {aclr_results[4]:.2f}], UTRA_1: [{aclr_results[1]:.2f}, {aclr_results[5]:.2f}], UTRA_2: [{aclr_results[0]:.2f}, {aclr_results[6]:.2f}]')
                                iem_results = self.command_cmw100_query('FETC:NRS:MEAS:MEV:IEM:MARG:AVER?')
                                iem_results = iem_results.split(',')
                                iem = f'{eval(iem_results[2]):.2f}' if iem_results[2] != 'INV' else 'INV'
                                logger.info(f'InBandEmissions Margin: {iem}dB')
                                # logger.info(f'IEM_MARG results: {iem_results}')
                                esfl_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:ESFL:EXTR?')
                                esfl_results = esfl_results.split(',')
                                ripple1 = f'{eval(esfl_results[2]):.2f}' if esfl_results[2] != 'NCAP' else 'NCAP'
                                ripple2 = f'{eval(esfl_results[3]):.2f}' if esfl_results[3] != 'NCAP' else 'NCAP'
                                logger.info(
                                    f'Equalize Spectrum Flatness: Ripple1:{ripple1} dBpp, Ripple2:{ripple2} dBpp')
                                time.sleep(0.2)
                                # logger.info(f'ESFL results: {esfl_results}')
                                sem_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:MARG:ALL?')
                                logger.info(f'SEM_MARG results: {sem_results}')
                                sem_avg_results = self.command_cmw100_query(f'FETC:NRS:MEAS:MEV:SEM:AVERage?')
                                sem_avg_results = sem_avg_results.split(',')
                                logger.info(
                                    f'OBW: {eval(sem_avg_results[2]) / 1000000:.3f} MHz, Total TX Power: {eval(sem_avg_results[3]):.2f} dBm')
                                # logger.info(f'SEM_AVER results: {sem_avg_results}')
                                self.command_cmw100_write(f'STOP:NRS:MEAS:MEV')
                                self.command_cmw100_query('*OPC?')
                                aclr_mod_current_results = aclr_mod_results = aclr_results + mod_results
                                logger.debug(aclr_mod_results)
                                aclr_mod_current_results.append(self.measure_current())
                                data[tx_level] = aclr_mod_current_results
                            logger.debug(data)
                            self.filename = self.tx_power_relative_test_export_excel(data, self.band_fr1, self.bw_fr1,
                                                                                     self.tx_freq_fr1)
        self.set_test_end_fr1()
        if plot:
            self.txp_aclr_evm_plot(self.filename, mode=0)
        else:
            pass

    def txp_aclr_evm_plot(self, filename, mode=0):  # mode 0: general, mode 1: LMH
        logger.info('----------Plot Chart---------')
        wb = openpyxl.load_workbook(filename)
        if self.script == 'GENERAL':
            if self.tech == 'LTE':
                if mode == 1:
                    for ws_name in wb.sheetnames:
                        logger.info(f'========={ws_name}==========')
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "Y1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=6, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "Y39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "Y77")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=21, min_row=1, max_col=21, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "Y115")

                    wb.save(filename)
                    wb.close()
                else:  # mode 0
                    for ws_name in wb.sheetnames:
                        logger.info(f'========={ws_name}==========')
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "V1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=10, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "V39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "V77")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=20, min_row=1, max_col=20, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "V115")

                    wb.save(filename)
                    wb.close()
            elif self.tech == 'FR1':
                if mode == 1:
                    for ws_name in wb.sheetnames:
                        logger.info(f'========={ws_name}==========')
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "AA1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=6, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "AA39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "AA77")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=23, min_row=1, max_col=23, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "AA115")

                    wb.save(filename)
                    wb.close()
                else:
                    for ws_name in wb.sheetnames:
                        logger.info(f'========={ws_name}==========')
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "X1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=10, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "X39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "X77")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=22, min_row=1, max_col=22, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "X115")
                    wb.save(filename)
                    wb.close()
            elif self.tech == 'WCDMA':
                if mode == 1:  # for LMH use
                    for ws_name in wb.sheetnames:
                        logger.info(f'========={ws_name}==========')
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "U1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=7, min_row=1, max_col=10, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for UTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for UTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-2
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+2

                        ws.add_chart(chart, "U39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "U77")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=17, min_row=1, max_col=17, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "U115")

                    wb.save(filename)
                    wb.close()
                else:  # for non LMH such as sweep
                    for ws_name in wb.sheetnames:
                        logger.info(f'========={ws_name}==========')
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "S1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=6, min_row=1, max_col=9, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for UTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for UTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-2
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+2

                        ws.add_chart(chart, "S39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "S77")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=16, min_row=1, max_col=16, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "S115")

                    wb.save(filename)
                    wb.close()
            elif self.tech == 'GSM':
                if mode == 1:
                    for ws_name in wb.sheetnames:
                        logger.info(f'========={ws_name}==========')
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "AA1")

                        logger.info('----------ORFS_MOD---------')
                        chart = LineChart()
                        chart.title = 'ORFS_MOD'
                        chart.y_axis.title = 'ORFS_MOD(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -80
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=10, min_row=1, max_col=15, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                        chart.series[2].marker.size = 10
                        chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                        chart.series[3].marker.size = 10
                        chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-200
                        chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+200
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-600
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+600

                        ws.add_chart(chart, "AA39")

                        logger.info('----------ORFS_SW---------')
                        chart = LineChart()
                        chart.title = 'ORFS_SW'
                        chart.y_axis.title = 'ORFS_SW(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -80
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=16, min_row=1, max_col=21, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                        chart.series[2].marker.size = 10
                        chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                        chart.series[3].marker.size = 10
                        chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-600
                        chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+600
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-1200
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+1200

                        ws.add_chart(chart, "AA77")

                        if 'GMSK' in ws_name:
                            logger.info('----------PHASE_RMS---------')
                            chart = LineChart()
                            chart.title = 'PHASE'
                            chart.y_axis.title = 'PHASE(degree)'
                            chart.x_axis.title = 'Band'
                            chart.x_axis.tickLblPos = 'low'

                            chart.height = 20
                            chart.width = 32

                            y_data = Reference(ws, min_col=7, min_row=1, max_col=7, max_row=ws.max_row)
                            x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                            chart.add_data(y_data, titles_from_data=True)
                            chart.set_categories(x_data)

                            chart.series[0].marker.symbol = 'circle'
                            chart.series[0].marker.size = 10

                            ws.add_chart(chart, "AA115")

                        elif 'EPSK' in ws_name:
                            logger.info('----------EVM_RMS---------')
                            chart = LineChart()
                            chart.title = 'EVM'
                            chart.y_axis.title = 'EVM(%)'
                            chart.x_axis.title = 'Band'
                            chart.x_axis.tickLblPos = 'low'

                            chart.height = 20
                            chart.width = 32

                            y_data = Reference(ws, min_col=8, min_row=1, max_col=8, max_row=ws.max_row)
                            x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                            chart.add_data(y_data, titles_from_data=True)
                            chart.set_categories(x_data)

                            chart.series[0].marker.symbol = 'circle'
                            chart.series[0].marker.size = 10

                            ws.add_chart(chart, "AA115")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=23, min_row=1, max_col=23, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "AA153")

                    wb.save(filename)
                    wb.close()
                else:  # mode = 0
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "X1")

                        logger.info('----------ORFS_MOD---------')
                        chart = LineChart()
                        chart.title = 'ORFS_MOD'
                        chart.y_axis.title = 'ORFS_MOD(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -80
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=9, min_row=1, max_col=14, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                        chart.series[2].marker.size = 10
                        chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                        chart.series[3].marker.size = 10
                        chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-200
                        chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+200
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-600
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+600

                        ws.add_chart(chart, "X39")

                        logger.info('----------ORFS_SW---------')
                        chart = LineChart()
                        chart.title = 'ORFS_SW'
                        chart.y_axis.title = 'ORFS_SW(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -80
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=15, min_row=1, max_col=20, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                        chart.series[2].marker.size = 10
                        chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                        chart.series[3].marker.size = 10
                        chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-600
                        chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+600
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-1200
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+1200

                        ws.add_chart(chart, "X77")

                        if 'GMSK' in ws_name:
                            logger.info('----------PHASE_RMS---------')
                            chart = LineChart()
                            chart.title = 'PHASE'
                            chart.y_axis.title = 'PHASE(degree)'
                            chart.x_axis.title = 'Band'
                            chart.x_axis.tickLblPos = 'low'

                            chart.height = 20
                            chart.width = 32

                            y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                            x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                            chart.add_data(y_data, titles_from_data=True)
                            chart.set_categories(x_data)

                            chart.series[0].marker.symbol = 'circle'
                            chart.series[0].marker.size = 10

                            ws.add_chart(chart, "X115")

                        elif 'EPSK' in ws_name:
                            logger.info('----------EVM_RMS---------')
                            chart = LineChart()
                            chart.title = 'EVM'
                            chart.y_axis.title = 'EVM(%)'
                            chart.x_axis.title = 'Band'
                            chart.x_axis.tickLblPos = 'low'

                            chart.height = 20
                            chart.width = 32

                            y_data = Reference(ws, min_col=7, min_row=1, max_col=7, max_row=ws.max_row)
                            x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                            chart.add_data(y_data, titles_from_data=True)
                            chart.set_categories(x_data)

                            chart.series[0].marker.symbol = 'circle'
                            chart.series[0].marker.size = 10

                            ws.add_chart(chart, "X115")

                        logger.info('----------Current---------')
                        chart = LineChart()
                        chart.title = 'Current'
                        chart.y_axis.title = 'Current(mA)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=22, min_row=1, max_col=22, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "X153")

                    wb.save(filename)
                    wb.close()

    def rxs_relative_plot(self, filename, mode=0):
        logger.info('----------Plot Chart---------')
        # self.tech ='LTE'
        # self.mcs_lte = 'QPSK'
        wb = openpyxl.load_workbook(filename)
        if self.script == 'GENERAL':
            if self.tech == 'LTE':
                if mode == 1:
                    ws_desens = wb[f'Desens_{self.mcs_lte}']
                    ws_txmax = wb[f'Raw_Data_{self.mcs_lte}_TxMax']
                    ws_txmin = wb[f'Raw_Data_{self.mcs_lte}_-10dBm']

                    if ws_desens._charts != []:  # if there is charts, delete it
                        ws_desens._charts.clear()

                    chart1 = LineChart()
                    chart1.title = 'Sensitivity'
                    chart1.y_axis.title = 'Rx_Level(dBm)'
                    chart1.x_axis.title = 'Band'
                    chart1.x_axis.tickLblPos = 'low'
                    chart1.height = 20
                    chart1.width = 32
                    y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
                    y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
                    y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
                    x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

                    series_txmax = Series(y_data_txmax, title="Tx_Max")
                    series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

                    chart1.append(series_txmax)
                    chart1.append(series_txmin)
                    chart1.set_categories(x_data)
                    chart1.y_axis.majorGridlines = None

                    chart2 = BarChart()
                    chart2.add_data(y_data_desens, titles_from_data=True)
                    chart2.y_axis.axId = 200
                    chart2.y_axis.title = 'Diff(dB)'

                    chart1.y_axis.crosses = "max"
                    chart1 += chart2

                    ws_desens.add_chart(chart1, "F1")

                    # logger.info('----------Power---------')
                    # chart = LineChart()
                    # chart.title = 'Power'
                    # chart.y_axis.title = 'Power(dBm)'
                    # chart.x_axis.title = 'Band'
                    # chart.x_axis.tickLblPos = 'low'
                    #
                    # chart.height = 20
                    # chart.width = 32
                    #
                    # y_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                    # x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    # chart.add_data(y_data, titles_from_data=True)
                    # chart.set_categories(x_data)
                    #
                    # chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    # chart.series[0].marker.size = 10
                    #
                    # ws.add_chart(chart, "U1")
                    #
                    # logger.info('----------ACLR---------')
                    # chart = LineChart()
                    # chart.title = 'ACLR'
                    # chart.y_axis.title = 'ACLR(dB)'
                    # chart.x_axis.title = 'Band'
                    # chart.x_axis.tickLblPos = 'low'
                    # chart.y_axis.scaling.min = -60
                    # chart.y_axis.scaling.max = -20
                    #
                    # chart.height = 20
                    # chart.width = 32
                    #
                    # y_data = Reference(ws, min_col=6, min_row=1, max_col=11, max_row=ws.max_row)
                    # x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    # chart.add_data(y_data, titles_from_data=True)
                    # chart.set_categories(x_data)
                    #
                    # chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                    # chart.series[0].marker.size = 10
                    # chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                    # chart.series[1].marker.size = 10
                    # chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                    # chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                    # chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                    # chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2
                    #
                    # ws.add_chart(chart, "U39")
                    #
                    # logger.info('----------EVM---------')
                    # chart = LineChart()
                    # chart.title = 'EVM'
                    # chart.y_axis.title = 'EVM(%)'
                    # chart.x_axis.title = 'Band'
                    # chart.x_axis.tickLblPos = 'low'
                    #
                    # chart.height = 20
                    # chart.width = 32
                    #
                    # y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                    # x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    # chart.add_data(y_data, titles_from_data=True)
                    # chart.set_categories(x_data)
                    #
                    # chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    # chart.series[0].marker.size = 10
                    #
                    # ws.add_chart(chart, "U77")

                    wb.save(filename)
                    wb.close()
                else:
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "R1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=10, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "R39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "R77")

                    wb.save(filename)
                    wb.close()
            elif self.tech == 'FR1':
                if mode == 1:
                    ws_desens = wb[f'Desens_{self.mcs_fr1}']
                    ws_txmax = wb[f'Raw_Data_{self.mcs_fr1}_TxMax']
                    ws_txmin = wb[f'Raw_Data_{self.mcs_fr1}_-10dBm']

                    if ws_desens._charts != []:  # if there is charts, delete it
                        ws_desens._charts.clear()

                    chart1 = LineChart()
                    chart1.title = 'Sensitivity'
                    chart1.y_axis.title = 'Rx_Level(dBm)'
                    chart1.x_axis.title = 'Band'
                    chart1.x_axis.tickLblPos = 'low'
                    chart1.height = 20
                    chart1.width = 32
                    y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
                    y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
                    y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
                    x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

                    series_txmax = Series(y_data_txmax, title="Tx_Max")
                    series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

                    chart1.append(series_txmax)
                    chart1.append(series_txmin)
                    chart1.set_categories(x_data)
                    chart1.y_axis.majorGridlines = None

                    chart2 = BarChart()
                    chart2.add_data(y_data_desens, titles_from_data=True)
                    chart2.y_axis.axId = 200
                    chart2.y_axis.title = 'Diff(dB)'

                    chart1.y_axis.crosses = "max"
                    chart1 += chart2

                    ws_desens.add_chart(chart1, "F1")

                    wb.save(filename)
                    wb.close()
                else:
                    for ws_name in wb.sheetnames:
                        ws = wb[ws_name]

                        if ws._charts != []:  # if there is charts, delete it
                            ws._charts.clear()

                        logger.info('----------Power---------')
                        chart = LineChart()
                        chart.title = 'Power'
                        chart.y_axis.title = 'Power(dBm)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "R1")

                        logger.info('----------ACLR---------')
                        chart = LineChart()
                        chart.title = 'ACLR'
                        chart.y_axis.title = 'ACLR(dB)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'
                        chart.y_axis.scaling.min = -60
                        chart.y_axis.scaling.max = -20

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=5, min_row=1, max_col=10, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                        chart.series[0].marker.size = 10
                        chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[1].marker.size = 10
                        chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                        chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                        chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                        chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                        ws.add_chart(chart, "R39")

                        logger.info('----------EVM---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                        chart.series[0].marker.size = 10

                        ws.add_chart(chart, "R77")

                    wb.save(filename)
                    wb.close()
            elif self.tech == 'WCDMA':
                if mode == 1:
                    ws = wb[f'Raw_Data']

                    if ws._charts != []:  # if there is charts, delete it
                        ws._charts.clear()

                    chart1 = LineChart()
                    chart1.title = 'Sensitivity'
                    chart1.y_axis.title = 'Rx_Level(dBm)'
                    chart1.x_axis.title = 'Band'
                    chart1.x_axis.tickLblPos = 'low'
                    chart1.height = 20
                    chart1.width = 32
                    y_data = Reference(ws, min_col=6, min_row=2, max_col=6, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)

                    series_pure_sens = Series(y_data, title="Pure_Sensititvity_FTM")

                    chart1.append(series_pure_sens)
                    chart1.set_categories(x_data)

                    ws.add_chart(chart1, "H1")

                    wb.save(filename)
                    wb.close()
                else:
                    pass
                    # for ws_name in wb.sheetnames:
                    #     ws = wb[ws_name]
                    #
                    #     if ws._charts != []:  # if there is charts, delete it
                    #         ws._charts.clear()
                    #
                    #     logger.info('----------Power---------')
                    #     chart = LineChart()
                    #     chart.title = 'Power'
                    #     chart.y_axis.title = 'Power(dBm)'
                    #     chart.x_axis.title = 'Band'
                    #     chart.x_axis.tickLblPos = 'low'
                    #
                    #     chart.height = 20
                    #     chart.width = 32
                    #
                    #     y_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                    #     x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    #     chart.add_data(y_data, titles_from_data=True)
                    #     chart.set_categories(x_data)
                    #
                    #     chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    #     chart.series[0].marker.size = 10
                    #
                    #     ws.add_chart(chart, "R1")
                    #
                    #     logger.info('----------ACLR---------')
                    #     chart = LineChart()
                    #     chart.title = 'ACLR'
                    #     chart.y_axis.title = 'ACLR(dB)'
                    #     chart.x_axis.title = 'Band'
                    #     chart.x_axis.tickLblPos = 'low'
                    #     chart.y_axis.scaling.min = -60
                    #     chart.y_axis.scaling.max = -20
                    #
                    #     chart.height = 20
                    #     chart.width = 32
                    #
                    #     y_data = Reference(ws, min_col=5, min_row=1, max_col=10, max_row=ws.max_row)
                    #     x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    #     chart.add_data(y_data, titles_from_data=True)
                    #     chart.set_categories(x_data)
                    #
                    #     chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                    #     chart.series[0].marker.size = 10
                    #     chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                    #     chart.series[1].marker.size = 10
                    #     chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                    #     chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                    #     chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                    #     chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2
                    #
                    #     ws.add_chart(chart, "R39")
                    #
                    #     logger.info('----------EVM---------')
                    #     chart = LineChart()
                    #     chart.title = 'EVM'
                    #     chart.y_axis.title = 'EVM(%)'
                    #     chart.x_axis.title = 'Band'
                    #     chart.x_axis.tickLblPos = 'low'
                    #
                    #     chart.height = 20
                    #     chart.width = 32
                    #
                    #     y_data = Reference(ws, min_col=11, min_row=1, max_col=11, max_row=ws.max_row)
                    #     x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    #     chart.add_data(y_data, titles_from_data=True)
                    #     chart.set_categories(x_data)
                    #
                    #     chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    #     chart.series[0].marker.size = 10
                    #
                    #     ws.add_chart(chart, "R77")
                    #
                    # wb.save(filename)
                    # wb.close()
            elif self.tech == 'GSM':
                if mode == 1:
                    ws = wb[f'Raw_Data']

                    if ws._charts != []:  # if there is charts, delete it
                        ws._charts.clear()

                    chart1 = LineChart()
                    chart1.title = 'Sensitivity'
                    chart1.y_axis.title = 'Rx_Level(dBm)'
                    chart1.x_axis.title = 'Band'
                    chart1.x_axis.tickLblPos = 'low'
                    chart1.height = 20
                    chart1.width = 32
                    y_data = Reference(ws, min_col=6, min_row=1, max_col=7, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)

                    chart1.add_data(y_data, titles_from_data=True)

                    chart1.set_categories(x_data)

                    ws.add_chart(chart1, "H1")

                    wb.save(filename)
                    wb.close()
                else:
                    pass

    @staticmethod
    def rxs_endc_plot(filename):
        logger.info('----------Plot Chart---------')
        wb = openpyxl.load_workbook(filename)
        ws_desens = wb[f'Desens_ENDC']
        ws_txmax = wb[f'Raw_Data_ENDC_FR1_TxMax']
        ws_txmin = wb[f'Raw_Data_ENDC_FR1_-10dBm']

        if ws_desens._charts != []:  # if there is charts, delete it
            ws_desens._charts.clear()

        chart1 = LineChart()
        chart1.title = 'Sensitivity'
        chart1.y_axis.title = 'Rx_Level(dBm)'
        chart1.x_axis.title = 'Band'
        chart1.x_axis.tickLblPos = 'low'
        chart1.height = 20
        chart1.width = 32
        y_data_txmax = Reference(ws_txmax, min_col=5, min_row=2, max_col=5, max_row=ws_txmax.max_row)
        y_data_txmin = Reference(ws_txmin, min_col=5, min_row=2, max_col=5, max_row=ws_txmin.max_row)
        y_data_desens = Reference(ws_desens, min_col=7, min_row=1, max_col=7, max_row=ws_desens.max_row)
        x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=6, max_row=ws_desens.max_row)

        series_txmax = Series(y_data_txmax, title="Tx_Max")
        series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

        chart1.append(series_txmax)
        chart1.append(series_txmin)
        chart1.set_categories(x_data)
        chart1.y_axis.majorGridlines = None

        chart2 = BarChart()
        chart2.add_data(y_data_desens, titles_from_data=True)
        chart2.y_axis.axId = 200
        chart2.y_axis.title = 'Diff(dB)'

        chart1.y_axis.crosses = "max"
        chart1 += chart2

        ws_desens.add_chart(chart1, "I1")

        wb.save(filename)
        wb.close()

    def run(self):
        for tech in wt.tech:
            if tech == 'LTE':
                self.tx_power_aclr_evm_lmh_pipeline_lte()
            elif tech == 'FR1':
                for script in wt.scripts:
                    if script == 'GENERAL':
                        self.tx_power_aclr_evm_lmh_pipeline_fr1()
                    elif script == 'FCC':
                        self.tx_power_pipline_fcc_fr1()
                    elif script == 'CE':
                        self.tx_power_pipline_ce_fr1()
                    elif tech == 'WCDMA':
                        self.tx_power_aclr_evm_lmh_pipeline_wcdma()
            elif tech == 'WCDMA':
                self.tx_power_aclr_evm_lmh_pipeline_wcdma()
            elif tech == 'GSM':
                self.tx_power_aclr_evm_lmh_pipeline_gsm()

    def run(self):
        for tech in wt.tech:
            if tech == 'LTE':
                self.search_sensitivity_pipline_lte()
            elif tech == 'FR1':
                for script in wt.scripts:
                    if script == 'GENERAL':
                        self.search_sensitivity_pipline_fr1()
                    elif script == 'ENDC' and wt.sa_nsa == 1:
                        self.sensitivity_pipline_endc()
            elif tech == 'WCDMA':
                self.search_sensitivity_pipline_wcdma()
            elif tech == 'GSM':
                self.search_sensitivity_pipline_gsm()

    def run_tx_level_sweep(self):
        for tech in wt.tech:
            if tech == 'LTE':
                self.tx_level_sweep_pipeline_lte()
            elif tech == 'FR1':
                self.tx_level_sweep_pipeline_fr1()
            elif tech == 'WCDMA':
                self.tx_level_sweep_pipeline_wcdma()
            elif tech == 'GSM':
                self.tx_level_sweep_pipeline_gsm()

    def run_tx_freq_sweep(self):
        for tech in wt.tech:
            if tech == 'LTE':
                self.tx_freq_sweep_pipline_lte()
            elif tech == 'FR1':
                self.tx_freq_sweep_pipline_fr1()
            elif tech == 'WCDMA':
                self.tx_freq_sweep_pipline_wcdma()
            elif tech == 'GSM':
                self.tx_freq_sweep_pipline_gsm()

    def run_tx_1rb_sweep(self):
        for tech in wt.tech:
            if tech == 'FR1':
                self.tx_1rb_sweep_pipeline_fr1()
            elif tech == 'LTE':
                pass
            else:
                pass

    def command_cmw100_query(self, tcpip_command):
        tcpip_response = self.cmw100.query(tcpip_command).strip()
        logger.info(f'TCPIP::<<{tcpip_command}')
        logger.info(f'TCPIP::>>{tcpip_response}')
        return tcpip_response

    def command_cmw100_write(self, tcpip_command):
        self.cmw100.write(tcpip_command)
        logger.info(f'TCPIP::<<{tcpip_command}')

    def command(self, command='at', delay=0.2):
        logger.info(f'MTM: <<{command}')
        command = command + '\r'
        self.ser.write(command.encode())
        time.sleep(delay)
        response = self.ser.readlines()
        for res in response:
            r = res.decode().split()
            if len(r) > 1:  # with more than one response
                for rr in r:
                    logger.info(f'MTM: >>{rr}')
            else:
                if r == []:  # sometimes there is not \r\n in the middle response
                    continue
                else:  # only one response
                    logger.info(f'MTM: >>{r[0]}')

        while b'OK\r\n' not in response:
            logger.info('OK is not at the end, so repeat again')
            logger.info(f'==========repeat to response again=========')
            response = self.ser.readlines()
            time.sleep(1)
            for res in response:
                r = res.decode().split()
                if len(r) > 1:  # with more than one response
                    for rr in r:
                        logger.info(f'MTM: >>{rr}')
                else:
                    if r == []:  # sometimes there is not \r\n in the middle response
                        continue
                    else:  # only one response
                        logger.info(f'MTM: >>{r[0]}')

        return response


def main():
    start = datetime.datetime.now()
    cmw100 = Cmw100()
    # cmw100.tx_power_aclr_evm_lmh_pipeline_lte()
    # cmw100.tx_freq_sweep_pipline_lte()
    # cmw100.tx_level_sweep_pipeline_lte()

    # cmw100.tx_power_aclr_evm_lmh_pipeline_fr1()
    # cmw100.tx_freq_sweep_pipline_fr1()
    # cmw100.tx_level_sweep_pipeline_fr1()

    # cmw100.search_sensitivity_pipline_lte()
    # cmw100.rx_desense_process()
    # cmw100.rxs_relative_plot('Sensitivty_10MHZ_LTE_LMH.xlsx', mode=1)

    # cmw100.search_sensitivity_pipline_lte()
    # cmw100.search_sensitivity_pipline_fr1()
    # cmw100.search_sensitivity_pipline_v2_fr1()
    # cmw100.tx_freq_fr1 = 3750000
    # cmw100.port_tx = 1
    # cmw100.bw_fr1 = 100
    # cmw100.tx_level = 26
    # cmw100.loss_tx = 1
    # cmw100.set_measurement_gprf()
    # cmw100.set_tx_freq_gprf()
    # cmw100.set_measure_start_on_gprf()
    # cmw100.search_sensitivity_pipline_endc()
    # cmw100.preset_instrument()
    # cmw100.tx_monitor_lte()
    # cmw100.set_test_mode_wcdma()
    # cmw100.tx_power_aclr_evm_lmh_pipeline_wcdma()
    # cmw100.tx_level_sweep_pipeline_wcdma()
    # cmw100.tx_freq_sweep_pipline_wcdma()
    # cmw100.search_sensitivity_pipline_wcdma()
    cmw100.tx_power_aclr_evm_lmh_pipeline_gsm()
    cmw100.tx_level_sweep_pipeline_gsm()
    cmw100.tx_freq_sweep_pipline_gsm()
    # cmw100.search_sensitivity_pipline_gsm()
    stop = datetime.datetime.now()

    logger.info(f'Timer: {stop - start}')


if __name__ == '__main__':
    main()
