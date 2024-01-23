import logging
import openpyxl
from openpyxl.chart import LineChart, Reference, BarChart, Series
from pathlib import Path

from utils.log_init import log_set
import utils.parameters.common_parameters_ftm as cm_pmt_ftm
import utils.parameters.common_parameters_anritsu as cm_pmt_anritsu
import utils.parameters.external_paramters as ext_pmt
from utils.channel_handler import chan_judge_fr1, chan_judge_lte, chan_judge_wcdma, chan_judge_gsm
from exception.custom_exception import FileNotFoundException
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from spec_limits.spec_limit_handler import import_aclr_limits, import_evm_limits

logger = log_set('excel_hdl')

rx_path_gsm_dict = {
    2: 'RX0',
    1: 'RX1',
    4: 'RX2',
    8: 'RX3',
    3: 'RX0+RX1',
    12: 'RX2+RX3',
    15: 'ALL PATH',
}
rx_path_wcdma_dict = {
    2: 'RX0',
    1: 'RX1',
    4: 'RX2',
    8: 'RX3',
    3: 'RX0+RX1',
    12: 'RX2+RX3',
    15: 'ALL PATH',
}
rx_path_lte_dict = {
    2: 'RX0',
    1: 'RX1',
    4: 'RX2',
    8: 'RX3',
    3: 'RX0+RX1',
    12: 'RX2+RX3',
    15: 'ALL PATH',
}
rx_path_fr1_dict = {
    2: 'RX0',
    1: 'RX1',
    4: 'RX2',
    8: 'RX3',
    3: 'RX0+RX1',
    12: 'RX2+RX3',
    15: 'ALL PATH',
}


def select_file_name_endc_ftm():
    if ext_pmt.part_number == "":
        return f'Sensitivty_ENDC.xlsx'
    else:
        return f'Sensitivty_ENDC_{ext_pmt.part_number}.xlsx'


def select_file_name_fcc_ce_ftm(script, tech):
    if ext_pmt.part_number == "":
        return f'Power_{script}_{tech}.xlsx'
    else:
        return f'Power_{script}_{tech}_{ext_pmt.part_number}.xlsx'


def select_file_name_rx_ftm(bw, tech):
    if ext_pmt.part_number == "":
        return f'Sensitivty_{bw}MHZ_{tech}_LMH.xlsx'
    else:
        return f'Sensitivty_{bw}MHZ_{tech}_LMH_{ext_pmt.part_number}.xlsx'


def select_file_name_genre_tx_ftm(bw, tech, test_item='lmh'):
    """
    select:
        level_sweep
        lmh
        freq_sweep
        1rb_sweep
        harmonics
    """
    if ext_pmt.part_number == "":
        if test_item == 'level_sweep':
            return f'Tx_level_sweep_{bw}MHZ_{tech}.xlsx'
        elif test_item == 'lmh':
            return f'Tx_Pwr_ACLR_EVM_{bw}MHZ_{tech}_LMH.xlsx'
        elif test_item == 'freq_sweep':
            return f'Tx_freq_sweep_{bw}MHZ_{tech}.xlsx'
        elif test_item == '1rb_sweep':
            return f'Tx_1RB_sweep_{bw}MHZ_{tech}.xlsx'
        elif test_item == 'harmonics':
            return f'Tx_harmonics_{bw}MHZ_{tech}.xlsx'
        elif test_item == 'cbe':
            return f'Tx_cbe_{bw}MHZ_{tech}.xlsx'
        elif test_item == 'apt_sweep':
            return f'Tx_apt_sweep_candidate_{bw}MHZ_{tech}.xlsx'

    else:
        if test_item == 'level_sweep':
            return f'Tx_level_sweep_{bw}MHZ_{tech}_{ext_pmt.part_number}.xlsx'
        elif test_item == 'lmh':
            return f'Tx_Pwr_ACLR_EVM_{bw}MHZ_{tech}_LMH_{ext_pmt.part_number}.xlsx'
        elif test_item == 'freq_sweep':
            return f'Tx_freq_sweep_{bw}MHZ_{tech}_{ext_pmt.part_number}.xlsx'
        elif test_item == '1rb_sweep':
            return f'Tx_1RB_sweep_{bw}MHZ_{tech}_{ext_pmt.part_number}.xlsx'
        elif test_item == 'harmonics':
            return f'Tx_harmonics_{bw}MHZ_{tech}_{ext_pmt.part_number}.xlsx'
        elif test_item == 'cbe':
            return f'Tx_cbe_{bw}MHZ_{tech}_{ext_pmt.part_number}.xlsx'
        elif test_item == 'apt_sweep':
            return f'Tx_apt_sweep_candidate_{bw}MHZ_{tech}_{ext_pmt.part_number}.xlsx'


def select_file_name_rx_sig(bw, tech):
    if ext_pmt.part_number == "":
        return f'Sensitivty_sig_{bw}MHZ_{tech}_LMH.xlsx'
    else:
        return f'Sensitivty_sig_{bw}MHZ_{tech}_LMH_{ext_pmt.part_number}.xlsx'


def select_file_name_rx_freq_sweep_sig(bw, tech):
    if ext_pmt.part_number == "":
        return f'Sensitivty_Freq_Sweep_sig_{bw}MHZ_{tech}_LMH.xlsx'
    else:
        return f'Sensitivty_Freq_Sweep_sig_{bw}MHZ_{tech}_LMH_{ext_pmt.part_number}.xlsx'


def select_file_name_genre_tx_sig(bw, standard, chcoding=None):
    """
    this is for signaling on Tx
    """
    if ext_pmt.part_number == "":
        if standard == 'LTE':
            return f'Tx_sig_{bw}MHZ_{standard}.xlsx'
        elif standard == 'WCDMA' and chcoding == 'REFMEASCH':  # WCDMA
            return f'Tx_sig_WCDMA.xlsx'
        elif standard == 'WCDMA' and chcoding == 'EDCHTEST':  # HSUPA
            return f'Tx_sig_HSUPA.xlsx'
        elif standard == 'WCDMA' and chcoding == 'FIXREFCH':  # HSDPA
            return f'Tx_sig_HSDPA.xlsx'
    else:
        if standard == 'LTE':
            return f'Tx_sig_{bw}MHZ_{standard}_{ext_pmt.part_number}.xlsx'
        elif standard == 'WCDMA' and chcoding == 'REFMEASCH':  # WCDMA
            return f'Tx_sig_WCDMA_{ext_pmt.part_number}.xlsx'
        elif standard == 'WCDMA' and chcoding == 'EDCHTEST':  # HSUPA
            return f'Tx_sig_HSUPA_{ext_pmt.part_number}.xlsx'
        elif standard == 'WCDMA' and chcoding == 'FIXREFCH':  # HSDPA
            return f'Tx_sig_HSDPA_{ext_pmt.part_number}.xlsx'


def excel_folder_create():
    file_dir = excel_folder_path()
    file_dir.mkdir(parents=True, exist_ok=True)
    logger.info('----------folder create ----------')
    logger.info(file_dir)


def excel_folder_path():
    path = Path('output') / Path(ext_pmt.devices_serial)
    logger.info(f'========== folder path: {path}==========')
    return path


def tx_power_fcc_ce_export_excel_ftm(data, parameters_dict):
    script = parameters_dict['script']
    tech = parameters_dict['tech']
    band = parameters_dict['band']
    bw = parameters_dict['bw']
    rb_size = parameters_dict['rb_size']
    rb_start = parameters_dict['rb_start']
    tx_level = parameters_dict['tx_level']
    mcs = parameters_dict['mcs']
    tx_path = parameters_dict['tx_path']
    logger.info('----------save to excel----------')
    filename = select_file_name_fcc_ce_ftm(script, tech)
    file_path = Path(excel_folder_path()) / Path(filename)
    if Path(file_path).exists() is False:
        logger.info('----------file does not exist----------')
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        # to create sheet
        sheetname = script
        ws = wb.create_sheet(sheetname)
        ws['A1'] = 'Band'
        ws['B1'] = 'BW'
        ws['C1'] = 'MCS'
        ws['D1'] = 'RB_size'
        ws['E1'] = 'RB_start'
        ws['F1'] = 'Tx_path'
        ws['G1'] = 'Chan'
        ws['H1'] = 'Tx_Freq'
        ws['I1'] = 'Tx_level'
        ws['J1'] = 'Power_measured'

        wb.save(file_path)
        wb.close()

    logger.info('----------file exist----------')
    wb = openpyxl.load_workbook(file_path)
    ws = wb[script]

    for tx_freq, measured_data in data.items():
        max_row = ws.max_row
        row = max_row + 1
        ws.cell(row, 1).value = band
        ws.cell(row, 2).value = bw
        ws.cell(row, 3).value = mcs
        ws.cell(row, 4).value = rb_size
        ws.cell(row, 5).value = rb_start
        ws.cell(row, 6).value = tx_path
        ws.cell(row, 7).value = measured_data[0]
        ws.cell(row, 8).value = tx_freq
        ws.cell(row, 9).value = tx_level
        ws.cell(row, 10).value = measured_data[1]

    wb.save(file_path)
    wb.close()
    return file_path


def tx_ulca_power_relative_test_export_excel_ftm(tech, data, sub_info):
    """
    input data:
    [6 items] + [OBW, sem_pwr]+ [U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2] + [Power, EVM, Freq_Err, IQ]*2 +
    [4 items] + path_setting(4 items)
    total 29 information
    """
    logger.info('----------save to excel----------')

    # filename choosen
    filename = None
    if ext_pmt.part_number == "" and sub_info['test_item'] != 'cbe':
        filename = f'Tx_ulca_{tech}.xlsx'
    elif ext_pmt.part_number != "" and sub_info['test_item'] != 'cbe':
        filename = f'Tx_ulca_{tech}_{ext_pmt.part_number}.xlsx'
    elif ext_pmt.part_number == "" and sub_info['test_item'] == 'cbe':
        filename = f'Tx_ulca_cbe_{tech}_{ext_pmt.part_number}.xlsx'
    elif ext_pmt.part_number != "" and sub_info['test_item'] == 'cbe':
        filename = f'Tx_ulca_cbe_{tech}_{ext_pmt.part_number}.xlsx'

    # file path by combinating the path and filename
    file_path = Path(excel_folder_path()) / Path(filename)

    if Path(file_path).exists() is False:
        logger.info('----------file does not exist----------')
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        # to create sheet
        if tech == 'LTE':
            # create dashboard
            for _ in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw100 might not have licesnse of Q256
                wb.create_sheet(f'Dashboard_{_}')
                # wb.create_sheet(f'Dashboard_{_}_PRB')
                # wb.create_sheet(f'Dashboard_{_}_FRB')

            # create the Raw data sheets
            for _ in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw100 might not have licesnse of Q256
                wb.create_sheet(f'Raw_Data_{_}')
                # wb.create_sheet(f'Raw_Data_{_}_PRB')
                # wb.create_sheet(f'Raw_Data_{_}_FRB')

            # create the title for every sheets
            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'Chan_LMH'
                    ws['C1'] = 'Tx_level'
                    ws['D1'] = 'CC1_BW'
                    ws['E1'] = 'CC2_BW'
                    ws['F1'] = 'CC1_chan'
                    ws['G1'] = 'CC2_chan'
                    ws['H1'] = 'OBW(MHz)'
                    ws['I1'] = 'SEM Power'
                    ws['J1'] = 'Carrier Power'
                    ws['K1'] = 'E_-1'
                    ws['L1'] = 'E_+1'
                    ws['M1'] = 'U_-1'
                    ws['N1'] = 'U_+1'
                    ws['O1'] = 'U_-2'
                    ws['P1'] = 'U_+2'
                    ws['Q1'] = 'CC1_PWR'
                    ws['R1'] = 'CC1_EVM'
                    ws['S1'] = 'CC1_Freq_Err'
                    ws['T1'] = 'CC1_IQ_OFFSET'
                    ws['U1'] = 'CC2_PWR'
                    ws['V1'] = 'CC2_EVM'
                    ws['W1'] = 'CC2_Freq_Err'
                    ws['X1'] = 'CC2_IQ_OFFSET'
                    ws['Y1'] = 'CC1_RB_num'
                    ws['Z1'] = 'CC1_RB_start'
                    ws['AA1'] = 'CC2_RB_num'
                    ws['AB1'] = 'CC2_RB_start'
                    ws['AC1'] = 'MCS'
                    ws['AD1'] = 'Tx_Path'
                    ws['AE1'] = 'CC1_RB_STATE'
                    ws['AF1'] = 'CC2_RB_STATE'
                    ws['AG1'] = 'Sync_Path'
                    ws['AH1'] = 'AS_Path'
                    ws['AI1'] = 'Current(mA)'
                    ws['AJ1'] = 'Condition'
                    ws['AK1'] = 'Temp0'
                    ws['AL1'] = 'Temp1'

                else:  # to pass the dashboard
                    pass

        # elif tech == 'FR1':  # this is not used for ulca
        #     # create dashboard
        #     for _ in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw100 might not have licesnse of Q256
        #         wb.create_sheet(f'Dashboard_{_}')
        #
        #     # create the Raw data sheets
        #     for _ in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw100 might not have licesnse of Q256
        #         wb.create_sheet(f'Raw_Data_{_}')
        #
        #     # create the title for every sheets
        #     for sheetname in wb.sheetnames:
        #         if 'Raw_Data' in sheetname:
        #             ws = wb[sheetname]
        #             ws['A1'] = 'Band'
        #             ws['B1'] = 'BW'
        #             ws['C1'] = 'Tx_Freq'
        #             ws['D1'] = 'Chan'
        #             ws['E1'] = 'Tx_level'
        #             ws['F1'] = 'Measured_Power'
        #             ws['G1'] = 'E_-1'
        #             ws['H1'] = 'E_+1'
        #             ws['I1'] = 'U_-1'
        #             ws['J1'] = 'U_+1'
        #             ws['K1'] = 'U_-2'
        #             ws['L1'] = 'U_+2'
        #             ws['M1'] = 'EVM'
        #             ws['N1'] = 'Freq_Err'
        #             ws['O1'] = 'IQ_OFFSET'
        #             ws['P1'] = 'RB_num'
        #             ws['Q1'] = 'RB_start'
        #             ws['R1'] = 'MCS'
        #             ws['S1'] = 'Type'
        #             ws['T1'] = 'Tx_Path'
        #             ws['U1'] = 'SCS(KHz)'
        #             ws['V1'] = 'RB_STATE'
        #             ws['W1'] = 'Sync_Path'
        #             ws['X1'] = 'AS_SRS_Path'
        #             ws['Y1'] = 'Current(mA)'
        #             ws['Z1'] = 'Condition'
        #             ws['AA1'] = 'Temp0'
        #             ws['AB1'] = 'Temp1'
        #             ws['AC1'] = '2f0' if test_item == 'harmonics' else None
        #             ws['AD1'] = '3f0' if test_item == 'harmonics' else None
        #         else:  # to pass the dashboard
        #             pass

        # save and close file
        wb.save(file_path)
        wb.close()

    logger.info('----------file exist----------')
    wb = openpyxl.load_workbook(file_path)
    ws = None
    if tech == 'LTE':
        mcs = data[27]
        ws = wb[f'Raw_Data_{mcs}']
    # elif tech == 'FR1':  this is not use for ULCA
    #     ws = wb[f'Raw_Data_{mcs}']

    if tech == 'LTE':
        max_row = ws.max_row
        row = max_row + 1

        ws.cell(row, 1).value = data[0]  # band
        ws.cell(row, 2).value = data[1]  # chan_lmh
        ws.cell(row, 3).value = ext_pmt.tx_level  # Tx_level
        ws.cell(row, 4).value = data[2]  # cc1_bw
        ws.cell(row, 5).value = data[3]  # cc2_bw
        ws.cell(row, 6).value = data[4]  # cc1_channel
        ws.cell(row, 7).value = data[5]  # cc2_channel
        ws.cell(row, 8).value = data[6] / 1000000  # OBW
        ws.cell(row, 9).value = data[7]  # sem power
        ws.cell(row, 10).value = data[11]  # carrier power
        ws.cell(row, 11).value = data[10]  # E_-1
        ws.cell(row, 12).value = data[12]  # E_+1
        ws.cell(row, 13).value = data[9]  # U_-1
        ws.cell(row, 14).value = data[13]  # U_+1
        ws.cell(row, 15).value = data[8]  # U_-2
        ws.cell(row, 16).value = data[14]  # U_+2
        ws.cell(row, 17).value = data[18]  # cc1_pwr
        ws.cell(row, 18).value = data[15]  # cc1_evm
        ws.cell(row, 19).value = data[16]  # cc1_freq_err
        ws.cell(row, 20).value = data[17]  # cc1_iq
        ws.cell(row, 21).value = data[22]  # cc2_pwr
        ws.cell(row, 22).value = data[19]  # cc2_evm
        ws.cell(row, 23).value = data[20]  # cc2_freq_err
        ws.cell(row, 24).value = data[21]  # cc2_iq
        ws.cell(row, 25).value = data[23]  # cc1_rb_num
        ws.cell(row, 26).value = data[24]  # cc1_rb_start
        ws.cell(row, 27).value = data[25]  # cc2_rb_num
        ws.cell(row, 28).value = data[26]  # cc2_iq
        ws.cell(row, 29).value = data[27]  # mcs
        ws.cell(row, 30).value = data[28]  # tx_path
        ws.cell(row, 31).value = sub_info['cc1_alloc']  # cc1_rb_state
        ws.cell(row, 32).value = sub_info['cc2_alloc']  # cc2_rb_state
        ws.cell(row, 33).value = data[29]  # Sync_Path
        ws.cell(row, 34).value = data[30]  # AS_Path
        ws.cell(row, 35).value = None  # Current(mA)
        ws.cell(row, 36).value = None  # Condition
        ws.cell(row, 37).value = sub_info['temp0']  # Temp0
        ws.cell(row, 38).value = sub_info['temp1']  # Temp1

    # elif tech == 'FR1':  # this is not for FR1
    #     max_row = ws.max_row
    #     row = max_row + 1
    #
    #     for tx_freq, measured_data in data.items():
    #         chan = chan_judge_fr1(band, bw, tx_freq) if test_item != 'freq_sweep' else None
    #         ws.cell(row, 1).value = band
    #         ws.cell(row, 2).value = bw
    #         ws.cell(row, 3).value = tx_freq
    #         ws.cell(row, 4).value = chan  # LMH
    #         ws.cell(row, 5).value = tx_freq_level  # this tx_level
    #         ws.cell(row, 6).value = measured_data[3]
    #         ws.cell(row, 7).value = measured_data[2]
    #         ws.cell(row, 8).value = measured_data[4]
    #         ws.cell(row, 9).value = measured_data[1]
    #         ws.cell(row, 10).value = measured_data[5]
    #         ws.cell(row, 11).value = measured_data[0]
    #         ws.cell(row, 12).value = measured_data[6]
    #         ws.cell(row, 13).value = measured_data[7]
    #         ws.cell(row, 14).value = measured_data[8]
    #         ws.cell(row, 15).value = measured_data[9]
    #         ws.cell(row, 16).value = rb_size
    #         ws.cell(row, 17).value = rb_start
    #         ws.cell(row, 18).value = mcs
    #         ws.cell(row, 19).value = type_
    #         ws.cell(row, 20).value = tx_path
    #         ws.cell(row, 21).value = scs
    #         ws.cell(row, 22).value = rb_state
    #         ws.cell(row, 23).value = sync_path
    #         ws.cell(row, 24).value = asw_srs_path
    #         ws.cell(row, 25).value = measured_data[10] if test_item == 'lmh' else None
    #         ws.cell(row, 26).value = ext_pmt.condition if test_item == 'lmh' else None
    #         ws.cell(row, 27).value = measured_data[11] if test_item == 'lmh' else None
    #         ws.cell(row, 28).value = measured_data[12] if test_item == 'lmh' else None
    #         ws.cell(row, 29).value = measured_data[13][1] if test_item == 'harmonics' else None  # 2f0
    #         ws.cell(row, 30).value = measured_data[14][1] if test_item == 'harmonics' else None  # 3f0
    #         row += 1

    wb.save(file_path)
    wb.close()

    return file_path


def tx_power_relative_test_export_excel_ftm(data, parameters_dict):
    """
    data is dict like:
    tx_level: [ U_-2, U_-1, E_-1, Pwr, E_+1, U_+1, U_+2, EVM, Freq_Err, IQ_OFFSET]
    """
    script = parameters_dict['script']
    tech = parameters_dict['tech']
    band = parameters_dict['band']
    bw = parameters_dict['bw']
    tx_freq_level = parameters_dict['tx_freq_level']
    mcs = parameters_dict['mcs']
    tx_path = parameters_dict['tx_path']
    mod = parameters_dict['mod']
    rb_state = parameters_dict['rb_state']
    rb_size = parameters_dict['rb_size']
    rb_start = parameters_dict['rb_start']
    sync_path = parameters_dict['sync_path']
    asw_srs_path = parameters_dict['asw_srs_path']
    scs = parameters_dict['scs']
    type_ = parameters_dict['type']
    test_item = parameters_dict['test_item']
    logger.info('----------save to excel----------')
    filename = None
    if script in ['GENERAL', 'CSE', 'APT']:
        if tx_freq_level >= 100:
            filename = select_file_name_genre_tx_ftm(bw, tech, test_item)
        elif tx_freq_level <= 100:
            filename = select_file_name_genre_tx_ftm(bw, tech, test_item)

        file_path = Path(excel_folder_path()) / Path(filename)

        if Path(file_path).exists() is False:
            logger.info('----------file does not exist----------')
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            # to create sheet
            if tech == 'LTE':
                # create dashboard
                for _ in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw100 might not have licesnse of Q256
                    wb.create_sheet(f'Dashboard_{_}')
                    # wb.create_sheet(f'Dashboard_{_}_PRB')
                    # wb.create_sheet(f'Dashboard_{_}_FRB')

                # create the Raw data sheets
                for _ in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw100 might not have licesnse of Q256
                    wb.create_sheet(f'Raw_Data_{_}')
                    # wb.create_sheet(f'Raw_Data_{_}_PRB')
                    # wb.create_sheet(f'Raw_Data_{_}_FRB')

                # create the title for every sheets
                for sheetname in wb.sheetnames:
                    if 'Raw_Data' in sheetname:
                        ws = wb[sheetname]
                        ws['A1'] = 'Band'
                        ws['B1'] = 'BW'
                        ws['C1'] = 'Tx_Freq'
                        ws['D1'] = 'Chan'
                        ws['E1'] = 'Tx_level'
                        ws['F1'] = 'Measured_Power'
                        ws['G1'] = 'E_-1'
                        ws['H1'] = 'E_+1'
                        ws['I1'] = 'U_-1'
                        ws['J1'] = 'U_+1'
                        ws['K1'] = 'U_-2'
                        ws['L1'] = 'U_+2'
                        ws['M1'] = 'EVM'
                        ws['N1'] = 'Freq_Err'
                        ws['O1'] = 'IQ_OFFSET'
                        ws['P1'] = 'RB_num'
                        ws['Q1'] = 'RB_start'
                        ws['R1'] = 'MCS'
                        ws['S1'] = 'Tx_Path'
                        ws['T1'] = 'RB_STATE'
                        ws['U1'] = 'Sync_Path'
                        ws['V1'] = 'AS_Path'
                        ws['W1'] = 'Current(mA)'
                        ws['X1'] = 'Condition'
                        ws['Y1'] = 'Temp0'
                        ws['Z1'] = 'Temp1'
                        if test_item == 'harmonics':
                            ws['AA1'] = '2f0'
                            ws['AB1'] = '3f0'
                        elif test_item in ['lmh', 'level_sweep']:
                            ws['AA1'] = 'Voltage'
                            ws['AB1'] = 'FBRX power'
                            ws['AC1'] = 'MIPI read'

                        elif test_item in ['apt_sweep']:
                            ws['AA1'] = 'Vcc'
                            ws['AB1'] = 'Bias0'
                            ws['AC1'] = 'Bias1'
                    else:  # to pass the dashboard
                        pass

            elif tech == 'FR1':
                # create dashboard
                for _ in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw100 might not have licesnse of Q256
                    wb.create_sheet(f'Dashboard_{_}')

                # create the Raw data sheets
                for _ in ['QPSK', 'Q16', 'Q64', 'Q256', 'BPSK']:  # some cmw100 might not have licesnse of Q256
                    wb.create_sheet(f'Raw_Data_{_}')

                # create the title for every sheets
                for sheetname in wb.sheetnames:
                    if 'Raw_Data' in sheetname:
                        ws = wb[sheetname]
                        ws['A1'] = 'Band'
                        ws['B1'] = 'BW'
                        ws['C1'] = 'Tx_Freq'
                        ws['D1'] = 'Chan'
                        ws['E1'] = 'Tx_level'
                        ws['F1'] = 'Measured_Power'
                        ws['G1'] = 'E_-1'
                        ws['H1'] = 'E_+1'
                        ws['I1'] = 'U_-1'
                        ws['J1'] = 'U_+1'
                        ws['K1'] = 'U_-2'
                        ws['L1'] = 'U_+2'
                        ws['M1'] = 'EVM'
                        ws['N1'] = 'Freq_Err'
                        ws['O1'] = 'IQ_OFFSET'
                        ws['P1'] = 'RB_num'
                        ws['Q1'] = 'RB_start'
                        ws['R1'] = 'MCS'
                        ws['S1'] = 'Type'
                        ws['T1'] = 'Tx_Path'
                        ws['U1'] = 'SCS(KHz)'
                        ws['V1'] = 'RB_STATE'
                        ws['W1'] = 'Sync_Path'
                        ws['X1'] = 'AS_SRS_Path'
                        ws['Y1'] = 'Current(mA)'
                        ws['Z1'] = 'Condition'
                        ws['AA1'] = 'Temp0'
                        ws['AB1'] = 'Temp1'
                        if test_item == 'harmonics':
                            ws['AC1'] = '2f0'
                            ws['AD1'] = '3f0'
                        elif test_item in ['lmh', 'level_sweep']:
                            ws['AC1'] = 'Voltage_mipi'
                            ws['AD1'] = 'FBRX power'
                            ws['AE1'] = 'MIPI read'
                        elif test_item in ['apt_sweep']:
                            ws['AC1'] = 'Vcc'
                            ws['AD1'] = 'Bias0'
                            ws['AE1'] = 'Bias1'
                    else:  # to pass the dashboard
                        pass

            elif tech == 'WCDMA':
                # create dashboard
                wb.create_sheet(f'Dashboard')

                # create the Raw data sheets
                wb.create_sheet(f'Raw_Data')

                # create the title for every sheets
                for sheetname in wb.sheetnames:
                    if 'Raw_Data' in sheetname:
                        ws = wb[sheetname]
                        ws['A1'] = 'Band'
                        ws['B1'] = 'Channel'
                        ws['C1'] = 'Chan'
                        ws['D1'] = 'Tx_Freq'
                        ws['E1'] = 'Tx_level'
                        ws['F1'] = 'Measured_Power'
                        ws['G1'] = 'U_-1'
                        ws['H1'] = 'U_+1'
                        ws['I1'] = 'U_-2'
                        ws['J1'] = 'U_+2'
                        ws['K1'] = 'OBW'
                        ws['L1'] = 'EVM'
                        ws['M1'] = 'Freq_Err'
                        ws['N1'] = 'IQ_OFFSET'
                        ws['O1'] = 'Tx_Path'
                        ws['P1'] = 'AS_Path'
                        ws['Q1'] = 'Current(mA)'
                        ws['R1'] = 'Condition'
                        ws['S1'] = 'Temp0'
                        ws['T1'] = 'Temp1'
                        if test_item == 'harmonics':
                            ws['U1'] = '2f0'
                            ws['V1'] = '3f0'
                        elif test_item in ['lmh', 'level_sweep']:
                            ws['U1'] = 'Voltage_mipi'
                            ws['V1'] = 'MIPI read'
                    else:  # to pass the dashboard
                        pass

            elif tech == 'GSM':
                # create dashboard
                wb.create_sheet(f'Dashboard_GMSK')
                wb.create_sheet(f'Dashboard_EPSK')

                # create the Raw data sheets
                wb.create_sheet(f'Raw_Data_GMSK')
                wb.create_sheet(f'Raw_Data_EPSK')

                # creat the title for every sheets
                for sheetname in wb.sheetnames:
                    if 'Raw_Data' in sheetname:
                        ws = wb[sheetname]
                        ws['A1'] = 'Band'
                        ws['B1'] = 'Channel'
                        ws['C1'] = 'Chan'
                        ws['D1'] = 'Rx_Freq'
                        ws['E1'] = 'Tx_PCL'
                        ws['F1'] = 'Measured_Power'
                        ws['G1'] = 'Phase_rms'
                        ws['H1'] = 'EVM_rms'
                        ws['I1'] = 'Ferr'
                        ws['J1'] = 'ORFS_MOD_-200'
                        ws['K1'] = 'ORFS_MOD_+200'
                        ws['L1'] = 'ORFS_MOD_-400'
                        ws['M1'] = 'ORFS_MOD_+400'
                        ws['N1'] = 'ORFS_MOD_-600'
                        ws['O1'] = 'ORFS_MOD_+600'
                        ws['P1'] = 'ORFS_SW_-400'
                        ws['Q1'] = 'ORFS_SW_+400'
                        ws['R1'] = 'ORFS_SW_-600'
                        ws['S1'] = 'ORFS_SW_+600'
                        ws['T1'] = 'ORFS_SW_-1200'
                        ws['U1'] = 'ORFS_SW_+1200'
                        ws['V1'] = 'AS_Path'
                        ws['W1'] = 'Current(mA)'
                        ws['X1'] = 'Condition'
                        ws['Y1'] = 'Temp0'
                        ws['Z1'] = 'Temp1'
                        ws['AA1'] = '2f0' if test_item == 'harmonics' else None
                        ws['AB1'] = '3f0' if test_item == 'harmonics' else None
                    else:  # to pass the dashboard
                        pass

            # save and close file
            wb.save(file_path)
            wb.close()

        logger.info('----------file exist----------')
        wb = openpyxl.load_workbook(file_path)
        ws = None
        if tech == 'LTE':
            ws = wb[f'Raw_Data_{mcs}']
        elif tech == 'FR1':
            ws = wb[f'Raw_Data_{mcs}']
        elif tech == 'WCDMA':
            ws = wb[f'Raw_Data']
        elif tech == 'GSM':
            ws = wb[f'Raw_Data_{mod}']

        if tech == 'LTE':
            max_row = ws.max_row
            row = max_row + 1
            if tx_freq_level >= 100:  # level_sweep
                for tx_level, measured_data in data.items():
                    chan = chan_judge_lte(band, bw, tx_freq_level)
                    ws.cell(row, 1).value = band
                    ws.cell(row, 2).value = bw
                    ws.cell(row, 3).value = tx_freq_level  # this freq_lte
                    ws.cell(row, 4).value = chan  # LMH
                    ws.cell(row, 5).value = tx_level
                    ws.cell(row, 6).value = measured_data[3]
                    ws.cell(row, 7).value = measured_data[2]
                    ws.cell(row, 8).value = measured_data[4]
                    ws.cell(row, 9).value = measured_data[1]
                    ws.cell(row, 10).value = measured_data[5]
                    ws.cell(row, 11).value = measured_data[0]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = measured_data[9]
                    ws.cell(row, 16).value = rb_size
                    ws.cell(row, 17).value = rb_start
                    ws.cell(row, 18).value = mcs
                    ws.cell(row, 19).value = tx_path
                    ws.cell(row, 20).value = rb_state
                    ws.cell(row, 21).value = sync_path
                    ws.cell(row, 22).value = asw_srs_path
                    ws.cell(row, 23).value = measured_data[10]
                    ws.cell(row, 24).value = ext_pmt.condition
                    ws.cell(row, 25).value = measured_data[11]
                    ws.cell(row, 26).value = measured_data[12]
                    ws.cell(row, 27).value = measured_data[13] if ext_pmt.volt_mipi_en else None  # volt_mipi
                    ws.cell(row, 28).value = measured_data[14] if ext_pmt.fbrx_en else None  # fbrx_power
                    ws.cell(row, 29).value = measured_data[15] if ext_pmt.mipi_read_en else None  # mipi_read
                    row += 1

            elif tx_freq_level <= 100:  # 1rb_sweep, lmh, freq_sweep, cbe
                for tx_freq, measured_data in data.items():
                    chan = chan_judge_lte(band, bw, tx_freq) if test_item != 'freq_sweep' else None
                    ws.cell(row, 1).value = band
                    ws.cell(row, 2).value = bw
                    ws.cell(row, 3).value = tx_freq
                    ws.cell(row, 4).value = chan  # LMH
                    ws.cell(row, 5).value = tx_freq_level  # this tx_level
                    ws.cell(row, 6).value = measured_data[3]
                    ws.cell(row, 7).value = measured_data[2]
                    ws.cell(row, 8).value = measured_data[4]
                    ws.cell(row, 9).value = measured_data[1]
                    ws.cell(row, 10).value = measured_data[5]
                    ws.cell(row, 11).value = measured_data[0]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = measured_data[9]
                    ws.cell(row, 16).value = rb_size
                    ws.cell(row, 17).value = rb_start
                    ws.cell(row, 18).value = mcs
                    ws.cell(row, 19).value = tx_path
                    ws.cell(row, 20).value = rb_state
                    ws.cell(row, 21).value = sync_path
                    ws.cell(row, 22).value = asw_srs_path
                    ws.cell(row, 23).value = measured_data[10] if test_item == 'lmh' else None
                    ws.cell(row, 24).value = ext_pmt.condition if test_item == 'lmh' else None
                    ws.cell(row, 25).value = measured_data[11] if test_item in ['lmh', 'cbe'] else None
                    ws.cell(row, 26).value = measured_data[12] if test_item in ['lmh', 'cbe'] else None
                    if test_item != 'harmonics':
                        ws.cell(row, 27).value = measured_data[13] if ext_pmt.volt_mipi_en else None  # volt_mipi
                        ws.cell(row, 28).value = measured_data[14] if ext_pmt.fbrx_en else None  # fbrx_power
                        ws.cell(row, 29).value = measured_data[15] if ext_pmt.mipi_read_en else None  # mipi_read

                    elif test_item == 'harmonics':
                        ws.cell(row, 27).value = measured_data[13][1]  # 2f0
                        ws.cell(row, 28).value = measured_data[14][1]  # 3f0

                    row += 1

        elif tech == 'FR1':
            max_row = ws.max_row
            row = max_row + 1
            if tx_freq_level >= 100:  # level_sweep
                for tx_level, measured_data in data.items():
                    chan = chan_judge_fr1(band, bw, tx_freq_level)
                    ws.cell(row, 1).value = band
                    ws.cell(row, 2).value = bw
                    ws.cell(row, 3).value = tx_freq_level  # this freq_fr1
                    ws.cell(row, 4).value = chan  # LMH
                    ws.cell(row, 5).value = tx_level
                    ws.cell(row, 6).value = measured_data[3]
                    ws.cell(row, 7).value = measured_data[2]
                    ws.cell(row, 8).value = measured_data[4]
                    ws.cell(row, 9).value = measured_data[1]
                    ws.cell(row, 10).value = measured_data[5]
                    ws.cell(row, 11).value = measured_data[0]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = measured_data[9]
                    ws.cell(row, 16).value = rb_size
                    ws.cell(row, 17).value = rb_start
                    ws.cell(row, 18).value = mcs
                    ws.cell(row, 19).value = type_
                    ws.cell(row, 20).value = tx_path
                    ws.cell(row, 21).value = scs
                    ws.cell(row, 22).value = rb_state
                    ws.cell(row, 23).value = sync_path
                    ws.cell(row, 24).value = asw_srs_path
                    ws.cell(row, 25).value = measured_data[10]
                    ws.cell(row, 26).value = ext_pmt.condition
                    ws.cell(row, 27).value = measured_data[11]
                    ws.cell(row, 28).value = measured_data[12]
                    if parameters_dict['test_item'] != 'apt_sweep':
                        ws.cell(row, 29).value = measured_data[13] if ext_pmt.volt_mipi_en else None  # volt_mipi
                        ws.cell(row, 30).value = measured_data[14] if ext_pmt.fbrx_en else None  # fbrx_power
                        ws.cell(row, 31).value = measured_data[15] if ext_pmt.mipi_read_en else None  # mipi_read
                    elif parameters_dict['test_item'] == 'apt_sweep':
                        ws.cell(row, 29).value = measured_data[11]  # vcc
                        ws.cell(row, 30).value = measured_data[12]  # bias0
                        ws.cell(row, 31).value = measured_data[13]  # bias1

                    row += 1

            elif tx_freq_level <= 100:  # 1rb_sweep, lmh, freq_sweep, cbe
                for tx_freq, measured_data in data.items():
                    chan = chan_judge_fr1(band, bw, tx_freq) if test_item != 'freq_sweep' else None
                    ws.cell(row, 1).value = band
                    ws.cell(row, 2).value = bw
                    ws.cell(row, 3).value = tx_freq
                    ws.cell(row, 4).value = chan  # LMH
                    ws.cell(row, 5).value = tx_freq_level  # this tx_level
                    ws.cell(row, 6).value = measured_data[3]
                    ws.cell(row, 7).value = measured_data[2]
                    ws.cell(row, 8).value = measured_data[4]
                    ws.cell(row, 9).value = measured_data[1]
                    ws.cell(row, 10).value = measured_data[5]
                    ws.cell(row, 11).value = measured_data[0]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = measured_data[9]
                    ws.cell(row, 16).value = rb_size
                    ws.cell(row, 17).value = rb_start
                    ws.cell(row, 18).value = mcs
                    ws.cell(row, 19).value = type_
                    ws.cell(row, 20).value = tx_path
                    ws.cell(row, 21).value = scs
                    ws.cell(row, 22).value = rb_state
                    ws.cell(row, 23).value = sync_path
                    ws.cell(row, 24).value = asw_srs_path
                    ws.cell(row, 25).value = measured_data[10] if test_item == 'lmh' else None
                    ws.cell(row, 26).value = ext_pmt.condition if test_item == 'lmh' else None
                    ws.cell(row, 27).value = measured_data[11] if test_item in ['lmh', 'cbe'] else None
                    ws.cell(row, 28).value = measured_data[12] if test_item in ['lmh', 'cbe'] else None
                    if test_item != 'harmonics':
                        ws.cell(row, 29).value = measured_data[13] if ext_pmt.volt_mipi_en else None  # volt_mipi
                        ws.cell(row, 30).value = measured_data[14] if ext_pmt.fbrx_en else None  # fbrx_power
                        ws.cell(row, 31).value = measured_data[15] if ext_pmt.mipi_read_en else None  # mipi_read
                    elif test_item == 'harmonics':
                        ws.cell(row, 29).value = measured_data[13][1]  # 2f0
                        ws.cell(row, 30).value = measured_data[14][1]  # 3f0

                    row += 1

        elif tech == 'WCDMA':
            max_row = ws.max_row
            row = max_row + 1
            if tx_freq_level >= 100:  # level_sweep
                for tx_level, measured_data in data.items():
                    chan = chan_judge_wcdma(band, tx_freq_level)
                    ws.cell(row, 1).value = band
                    # this channel
                    ws.cell(row, 2).value = cm_pmt_ftm.transfer_freq2chan_wcdma(band, tx_freq_level, 'tx')
                    ws.cell(row, 3).value = chan  # LMH
                    ws.cell(row, 4).value = tx_freq_level  # this tx_freq_wcdma
                    ws.cell(row, 5).value = tx_level
                    ws.cell(row, 6).value = measured_data[5]
                    ws.cell(row, 7).value = measured_data[0]
                    ws.cell(row, 8).value = measured_data[1]
                    ws.cell(row, 9).value = measured_data[2]
                    ws.cell(row, 10).value = measured_data[3]
                    ws.cell(row, 11).value = measured_data[4]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = tx_path
                    ws.cell(row, 16).value = asw_srs_path
                    ws.cell(row, 17).value = measured_data[9]
                    ws.cell(row, 18).value = ext_pmt.condition
                    ws.cell(row, 19).value = measured_data[10]
                    ws.cell(row, 20).value = measured_data[11]
                    ws.cell(row, 21).value = measured_data[12] if ext_pmt.volt_mipi_en else None  # volt_mipi
                    ws.cell(row, 22).value = measured_data[13] if ext_pmt.mipi_read_en else None  # mipi_read
                    row += 1

            elif tx_freq_level <= 100:  # 1rb_sweep, lmh, freq_sweep
                for tx_freq, measured_data in data.items():
                    chan = chan_judge_wcdma(band, tx_freq) if test_item != 'freq_sweep' else None
                    ws.cell(row, 1).value = band
                    ws.cell(row, 2).value = cm_pmt_ftm.transfer_freq2chan_wcdma(band, tx_freq, 'tx')  # this channel
                    ws.cell(row, 3).value = chan  # LMH
                    ws.cell(row, 4).value = tx_freq
                    ws.cell(row, 5).value = tx_freq_level  # this tx_level
                    ws.cell(row, 6).value = measured_data[5]
                    ws.cell(row, 7).value = measured_data[0]
                    ws.cell(row, 8).value = measured_data[1]
                    ws.cell(row, 9).value = measured_data[2]
                    ws.cell(row, 10).value = measured_data[3]
                    ws.cell(row, 11).value = measured_data[4]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = tx_path
                    ws.cell(row, 16).value = asw_srs_path
                    ws.cell(row, 17).value = measured_data[9] if test_item == 'lmh' else None
                    ws.cell(row, 18).value = ext_pmt.condition if test_item == 'lmh' else None
                    ws.cell(row, 19).value = measured_data[10] if test_item == 'lmh' else None
                    ws.cell(row, 20).value = measured_data[11] if test_item == 'lmh' else None
                    if test_item != 'harmonics':
                        ws.cell(row, 21).value = measured_data[12] if ext_pmt.volt_mipi_en else None  # volt_mipi
                        ws.cell(row, 22).value = measured_data[13] if ext_pmt.mipi_read_en else None  # mipi_read
                    elif test_item == 'harmonics':
                        ws.cell(row, 21).value = measured_data[12][1]  # 2f0
                        ws.cell(row, 22).value = measured_data[13][1]  # 3f0

                    row += 1

        elif tech == 'GSM':
            max_row = ws.max_row
            row = max_row + 1
            if tx_freq_level >= 100:  # level_sweep
                for tx_pcl, measured_data in data.items():
                    chan = chan_judge_gsm(band, tx_freq_level)
                    ws.cell(row, 1).value = band
                    ws.cell(row, 2).value = cm_pmt_ftm.transfer_freq2chan_gsm(band, tx_freq_level)  # channel
                    ws.cell(row, 3).value = chan  # LMH
                    ws.cell(row, 4).value = tx_freq_level  # this rx_freq_gsm
                    ws.cell(row, 5).value = tx_pcl
                    ws.cell(row, 6).value = measured_data[0]
                    ws.cell(row, 7).value = measured_data[1]
                    ws.cell(row, 8).value = measured_data[2]
                    ws.cell(row, 9).value = measured_data[3]
                    ws.cell(row, 10).value = measured_data[4]
                    ws.cell(row, 11).value = measured_data[5]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = measured_data[9]
                    ws.cell(row, 16).value = measured_data[10]
                    ws.cell(row, 17).value = measured_data[11]
                    ws.cell(row, 18).value = measured_data[12]
                    ws.cell(row, 19).value = measured_data[13]
                    ws.cell(row, 20).value = measured_data[14]
                    ws.cell(row, 21).value = measured_data[15]
                    ws.cell(row, 22).value = asw_srs_path
                    ws.cell(row, 23).value = measured_data[16]
                    ws.cell(row, 24).value = ext_pmt.condition
                    ws.cell(row, 25).value = measured_data[17]
                    ws.cell(row, 26).value = measured_data[18]

                    row += 1

            elif tx_freq_level <= 100:  # 1rb_sweep, lmh, freq_sweep
                for rx_freq, measured_data in data.items():
                    chan = chan_judge_gsm(band, rx_freq) if test_item != 'freq_sweep' else None
                    ws.cell(row, 1).value = band
                    ws.cell(row, 2).value = cm_pmt_ftm.transfer_freq2chan_gsm(band, rx_freq)
                    ws.cell(row, 3).value = chan  # LMH
                    ws.cell(row, 4).value = rx_freq  # this rx_freq_gsm
                    ws.cell(row, 5).value = tx_freq_level  # this pcl
                    ws.cell(row, 6).value = measured_data[0]
                    ws.cell(row, 7).value = measured_data[1]
                    ws.cell(row, 8).value = measured_data[2]
                    ws.cell(row, 9).value = measured_data[3]
                    ws.cell(row, 10).value = measured_data[4]
                    ws.cell(row, 11).value = measured_data[5]
                    ws.cell(row, 12).value = measured_data[6]
                    ws.cell(row, 13).value = measured_data[7]
                    ws.cell(row, 14).value = measured_data[8]
                    ws.cell(row, 15).value = measured_data[9]
                    ws.cell(row, 16).value = measured_data[10]
                    ws.cell(row, 17).value = measured_data[11]
                    ws.cell(row, 18).value = measured_data[12]
                    ws.cell(row, 19).value = measured_data[13]
                    ws.cell(row, 20).value = measured_data[14]
                    ws.cell(row, 21).value = measured_data[15]
                    ws.cell(row, 22).value = asw_srs_path
                    ws.cell(row, 23).value = measured_data[16] if test_item == 'lmh' else None
                    ws.cell(row, 24).value = ext_pmt.condition if test_item == 'lmh' else None
                    ws.cell(row, 25).value = measured_data[17] if test_item == 'lmh' else None
                    ws.cell(row, 26).value = measured_data[18] if test_item == 'lmh' else None
                    ws.cell(row, 27).value = measured_data[19][1] if test_item == 'harmonics' else None  # 2f0
                    ws.cell(row, 28).value = measured_data[20][1] if test_item == 'harmonics' else None  # 3f0
                    row += 1

        wb.save(file_path)
        wb.close()

        return file_path


def txp_aclr_evm_current_plot_ftm(file_path, parameters_dict):
    script = parameters_dict['script']
    tech = parameters_dict['tech']
    # band = parameters_dict['band']
    # bw = parameters_dict['bw']
    # tx_freq_level = parameters_dict['tx_freq_level']
    # mcs = parameters_dict['mcs']
    # tx_path = parameters_dict['tx_path']
    # mod = parameters_dict['mod']
    # rb_state = parameters_dict['rb_state']
    # rb_size = parameters_dict['rb_size']
    # rb_start = parameters_dict['rb_start']
    # sync_path = parameters_dict['sync_path']
    # asw_srs_path = parameters_dict['asw_srs_path']
    # scs = parameters_dict['scs']
    # type_ = parameters_dict['type']
    logger.info('----------Plot Chart---------')
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception:
        raise FileNotFoundException(f'Cannot find {file_path}')

    if script in ['GENERAL', 'CSE', 'APT']:
        if tech == 'LTE':
            for ws_name in wb.sheetnames:
                if 'Raw_Data' in ws_name:
                    logger.info(f'========={ws_name}==========')
                    ws = wb[ws_name]
                    ws_dashboard = wb['Dashboard_' + ws_name[9:]]

                    if ws_dashboard._charts:  # if there is charts, delete it
                        ws_dashboard._charts.clear()

                    logger.info('----------Power---------')
                    chart = LineChart()
                    chart.title = 'Power'
                    chart.y_axis.title = 'Power(dBm)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A1")

                    logger.info('----------ACLR---------')
                    chart = LineChart()
                    chart.title = 'ACLR'
                    chart.y_axis.title = 'ACLR(dB)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'
                    chart.y_axis.scaling.min = -60
                    chart.y_axis.scaling.max = -20

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=7, min_row=1, max_col=12, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                    chart.series[0].marker.size = 10
                    chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[1].marker.size = 10
                    chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                    chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                    chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                    chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                    ws_dashboard.add_chart(chart, "A41")

                    logger.info('----------EVM---------')
                    chart = LineChart()
                    chart.title = 'EVM'
                    chart.y_axis.title = 'EVM(%)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=13, min_row=1, max_col=13, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A81")

                    logger.info('----------Current---------')
                    chart = LineChart()
                    chart.title = 'Current'
                    chart.y_axis.title = 'Current(mA)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=23, min_row=1, max_col=23, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A121")
                else:
                    pass

            wb.save(file_path)
            wb.close()

        elif tech == 'FR1':
            for ws_name in wb.sheetnames:
                if 'Raw_Data' in ws_name:
                    logger.info(f'========={ws_name}==========')
                    ws = wb[ws_name]
                    ws_dashboard = wb['Dashboard_' + ws_name[9:]]

                    if ws_dashboard._charts:  # if there is charts, delete it
                        ws_dashboard._charts.clear()

                    logger.info('----------Power---------')
                    chart = LineChart()
                    chart.title = 'Power'
                    chart.y_axis.title = 'Power(dBm)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A1")

                    logger.info('----------ACLR---------')
                    chart = LineChart()
                    chart.title = 'ACLR'
                    chart.y_axis.title = 'ACLR(dB)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'
                    chart.y_axis.scaling.min = -60
                    chart.y_axis.scaling.max = -20

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=7, min_row=1, max_col=12, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                    chart.series[0].marker.size = 10
                    chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[1].marker.size = 10
                    chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                    chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                    chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                    chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                    ws_dashboard.add_chart(chart, "A41")

                    logger.info('----------EVM---------')
                    chart = LineChart()
                    chart.title = 'EVM'
                    chart.y_axis.title = 'EVM(%)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=13, min_row=1, max_col=13, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A81")

                    logger.info('----------Current---------')
                    chart = LineChart()
                    chart.title = 'Current'
                    chart.y_axis.title = 'Current(mA)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=25, min_row=1, max_col=25, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A121")

            wb.save(file_path)
            wb.close()

        elif tech == 'WCDMA':
            for ws_name in wb.sheetnames:
                if 'Raw_Data' in ws_name:
                    logger.info(f'========={ws_name}==========')
                    ws = wb[ws_name]
                    ws_dashboard = wb['Dashboard']

                    if ws_dashboard._charts:  # if there is charts, delete it
                        ws_dashboard._charts.clear()

                    logger.info('----------Power---------')
                    chart = LineChart()
                    chart.title = 'Power'
                    chart.y_axis.title = 'Power(dBm)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A1")

                    logger.info('----------ACLR---------')
                    chart = LineChart()
                    chart.title = 'ACLR'
                    chart.y_axis.title = 'ACLR(dB)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'
                    chart.y_axis.scaling.min = -60
                    chart.y_axis.scaling.max = -20

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=7, min_row=1, max_col=10, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'triangle'  # for UTRA_-1
                    chart.series[0].marker.size = 10
                    chart.series[1].marker.symbol = 'circle'  # for UTRA_+1
                    chart.series[1].marker.size = 10
                    chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-2
                    chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+2

                    ws_dashboard.add_chart(chart, "A41")

                    logger.info('----------EVM---------')
                    chart = LineChart()
                    chart.title = 'EVM'
                    chart.y_axis.title = 'EVM(%)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A81")

                    logger.info('----------Current---------')
                    chart = LineChart()
                    chart.title = 'Current'
                    chart.y_axis.title = 'Current(mA)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=17, min_row=1, max_col=17, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A121")

            wb.save(file_path)
            wb.close()
        elif tech == 'GSM':
            for ws_name in wb.sheetnames:
                if 'Raw_Data' in ws_name:
                    logger.info(f'========={ws_name}==========')
                    ws = wb[ws_name]
                    ws_dashboard = wb['Dashboard_' + ws_name[9:]]

                    if ws_dashboard._charts:  # if there is charts, delete it
                        ws_dashboard._charts.clear()

                    logger.info('----------Power---------')
                    chart = LineChart()
                    chart.title = 'Power'
                    chart.y_axis.title = 'Power(dBm)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A1")

                    logger.info('----------ORFS_MOD---------')
                    chart = LineChart()
                    chart.title = 'ORFS_MOD'
                    chart.y_axis.title = 'ORFS_MOD(dB)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'
                    chart.y_axis.scaling.min = -80
                    chart.y_axis.scaling.max = -20

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=10, min_row=1, max_col=15, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                    chart.series[2].marker.size = 10
                    chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                    chart.series[3].marker.size = 10
                    chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-200
                    chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+200
                    chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-600
                    chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+600

                    ws_dashboard.add_chart(chart, "A41")

                    logger.info('----------ORFS_SW---------')
                    chart = LineChart()
                    chart.title = 'ORFS_SW'
                    chart.y_axis.title = 'ORFS_SW(dBm)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'
                    chart.y_axis.scaling.min = -80
                    chart.y_axis.scaling.max = -20

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=16, min_row=1, max_col=21, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                    chart.series[2].marker.size = 10
                    chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                    chart.series[3].marker.size = 10
                    chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-600
                    chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+600
                    chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-1200
                    chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+1200

                    ws_dashboard.add_chart(chart, "A81")

                    if 'GMSK' in ws_name:
                        logger.info('----------PHASE_RMS---------')
                        chart = LineChart()
                        chart.title = 'PHASE'
                        chart.y_axis.title = 'PHASE(degree)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=7, min_row=1, max_col=7, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws_dashboard.add_chart(chart, "A115")

                    elif 'EPSK' in ws_name:
                        logger.info('----------EVM_RMS---------')
                        chart = LineChart()
                        chart.title = 'EVM'
                        chart.y_axis.title = 'EVM(%)'
                        chart.x_axis.title = 'Band'
                        chart.x_axis.tickLblPos = 'low'

                        chart.height = 20
                        chart.width = 32

                        y_data = Reference(ws, min_col=8, min_row=1, max_col=8, max_row=ws.max_row)
                        x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                        chart.add_data(y_data, titles_from_data=True)
                        chart.set_categories(x_data)

                        chart.series[0].marker.symbol = 'circle'
                        chart.series[0].marker.size = 10

                        ws_dashboard.add_chart(chart, "A121")

                    logger.info('----------Current---------')
                    chart = LineChart()
                    chart.title = 'Current'
                    chart.y_axis.title = 'Current(mA)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=23, min_row=1, max_col=23, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A161")

            wb.save(file_path)
            wb.close()


def rx_power_relative_test_export_excel_ftm(data, parameters_dict):
    """
    data is dict like:
    rx_level: [ Power, Rx_level, [RSRP0-3], [CINR0-3], [AFC0-3]]
    """
    script = parameters_dict['script']
    tech = parameters_dict['tech']
    band = parameters_dict['band']
    bw = parameters_dict['bw']
    tx_level = parameters_dict['tx_level']
    mcs = parameters_dict['mcs']
    tx_path = parameters_dict['tx_path']
    rx_path = parameters_dict['rx_path']
    if tech in ['FR1', 'LTE']:
        rb_size = parameters_dict['rb_size']
        rb_start = parameters_dict['rb_start']

    logger.info('----------save to excel----------')
    if script == 'GENERAL':
        filename = select_file_name_rx_ftm(bw, tech)
        file_path = Path(excel_folder_path()) / Path(filename)

        if Path(file_path).exists() is False:  # if the file does not exist
            logger.info('----------file does not exist----------')
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])
            # create dashboard
            wb.create_sheet(f'Dashboard')

            # to create sheet
            if tech == 'LTE':
                # create the title and sheet for TxManx and -10dBm
                wb.create_sheet(f'Raw_Data_{mcs}_TxMax')
                wb.create_sheet(f'Raw_Data_{mcs}_-10dBm')
                for sheetname in wb.sheetnames:
                    if 'Raw_Data' in sheetname:
                        ws = wb[sheetname]
                        ws['A1'] = 'Band'
                        ws['B1'] = 'RX_Path'
                        ws['C1'] = 'Chan'
                        ws['D1'] = 'Tx_Freq'
                        ws['E1'] = 'Tx_level'
                        ws['F1'] = 'Power'
                        ws['G1'] = 'Rx_Level'
                        ws['H1'] = 'RSRP_RX0'
                        ws['I1'] = 'RSRP_RX1'
                        ws['J1'] = 'RSRP_RX2'
                        ws['K1'] = 'RSRP_RX3'
                        ws['L1'] = 'CINR_RX0'
                        ws['M1'] = 'CINR_RX1'
                        ws['N1'] = 'CINR_RX2'
                        ws['O1'] = 'CINR_RX3'
                        ws['P1'] = 'AGC_RX0'
                        ws['Q1'] = 'AGC_RX1'
                        ws['R1'] = 'AGC_RX2'
                        ws['S1'] = 'AGC_RX3'
                        ws['T1'] = 'TX_Path'
                        ws['U1'] = 'BW'
                        ws['V1'] = 'RB_num_UL'
                        ws['W1'] = 'RB_start_UL'
                        ws['X1'] = 'Condition'
                        ws['Y1'] = 'Temp0'
                        ws['Z1'] = 'Temp1'
                    else:
                        pass

                # create the title and sheet for Desense
                wb.create_sheet(f'Desens_{mcs}')
                ws = wb[f'Desens_{mcs}']
                ws['A1'] = 'Band'
                ws['B1'] = 'Rx_Path'
                ws['C1'] = 'Chan'
                ws['D1'] = 'Diff'
                ws['E1'] = 'TX_Path'

            elif tech == 'FR1':
                # create the title and sheet for TxManx and -10dBm
                wb.create_sheet(f'Raw_Data_{mcs}_TxMax')
                wb.create_sheet(f'Raw_Data_{mcs}_-10dBm')
                for sheetname in wb.sheetnames:
                    if 'Raw_Data' in sheetname:
                        ws = wb[sheetname]
                        ws['A1'] = 'Band'
                        ws['B1'] = 'RX_Path'
                        ws['C1'] = 'Chan'
                        ws['D1'] = 'Tx_Freq'
                        ws['E1'] = 'Tx_level'
                        ws['F1'] = 'Power'
                        ws['G1'] = 'Rx_Level'
                        ws['H1'] = 'RSRP_RX0'
                        ws['I1'] = 'RSRP_RX1'
                        ws['J1'] = 'RSRP_RX2'
                        ws['K1'] = 'RSRP_RX3'
                        ws['L1'] = 'CINR_RX0'
                        ws['M1'] = 'CINR_RX1'
                        ws['N1'] = 'CINR_RX2'
                        ws['O1'] = 'CINR_RX3'
                        ws['P1'] = 'AGC_RX0'
                        ws['Q1'] = 'AGC_RX1'
                        ws['R1'] = 'AGC_RX2'
                        ws['S1'] = 'AGC_RX3'
                        ws['T1'] = 'TX_Path'
                        ws['U1'] = 'BW'
                        ws['V1'] = 'RB_num_UL'
                        ws['W1'] = 'RB_start_UL'
                        ws['X1'] = 'Condition'
                        ws['Y1'] = 'Temp0'
                        ws['Z1'] = 'Temp1'
                    else:
                        pass

                # create the title and sheet for Desense
                wb.create_sheet(f'Desens_{mcs}')
                ws = wb[f'Desens_{mcs}']
                ws['A1'] = 'Band'
                ws['B1'] = 'Rx_Path'
                ws['C1'] = 'Chan'
                ws['D1'] = 'Diff'
                ws['E1'] = 'TX_Path'

            elif tech == 'WCDMA':
                # create the title and sheet for TxManx and -10dBm
                wb.create_sheet(f'Raw_Data')
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'RX_Path'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Rx_Channel'
                    ws['E1'] = 'Tx_Freq'
                    ws['F1'] = 'Rx_Level'

                # create the title and sheet for Desense
                wb.create_sheet(f'Desens_{mcs}')
                ws = wb[f'Desens_{mcs}']
                ws['A1'] = 'Band'
                ws['B1'] = 'Rx_Path'
                ws['C1'] = 'Chan'
                ws['D1'] = 'Diff'

            elif tech == 'GSM':
                wb.create_sheet(f'Raw_Data')
                for sheetname in wb.sheetnames:
                    # create the title and sheet for TxManx and -10dBm
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'RX_Path'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Channel'
                    ws['E1'] = 'Rx_Freq'
                    ws['F1'] = 'Rx_Level'
                    ws['G1'] = 'RSSI'

            # save and close file
            wb.save(file_path)
            wb.close()

        # if the file exist
        logger.info('----------file exist----------')
        wb = openpyxl.load_workbook(file_path)
        ws = None
        # to fetch the sheet name
        if tech == 'LTE':
            sheetname = f'Raw_Data_{mcs}_TxMax' if tx_level > 0 else f'Raw_Data_{mcs}_-10dBm'
            ws = wb[sheetname]
        elif tech == 'FR1':
            sheetname = f'Raw_Data_{mcs}_TxMax' if tx_level > 0 else f'Raw_Data_{mcs}_-10dBm'
            ws = wb[sheetname]
        elif tech == 'WCDMA':
            sheetname = 'Raw_Data'
            ws = wb[sheetname]
        elif tech == 'GSM':
            sheetname = 'Raw_Data'
            ws = wb[sheetname]

        if tech == 'LTE':
            max_row = ws.max_row
            row = max_row + 1  # skip title
            for tx_freq, measured_data in data.items():
                chan = chan_judge_lte(band, bw, tx_freq)
                ws.cell(row, 1).value = band
                ws.cell(row, 2).value = rx_path_lte_dict[rx_path]
                ws.cell(row, 3).value = chan  # LMH
                ws.cell(row, 4).value = tx_freq
                ws.cell(row, 5).value = tx_level  # this tx level
                ws.cell(row, 6).value = measured_data[0]  # measured power
                ws.cell(row, 7).value = measured_data[1]  # RX level
                ws.cell(row, 8).value = measured_data[2][0]  # RSRP_RX0
                ws.cell(row, 9).value = measured_data[2][1]  # RSRP_RX1
                ws.cell(row, 10).value = measured_data[2][2]  # RSRP_RX2
                ws.cell(row, 11).value = measured_data[2][3]  # RSRP_RX3
                ws.cell(row, 12).value = measured_data[3][0]  # CINR_RX0
                ws.cell(row, 13).value = measured_data[3][1]  # CINR_RX1
                ws.cell(row, 14).value = measured_data[3][2]  # CINR_RX2
                ws.cell(row, 15).value = measured_data[3][3]  # CINR_RX3
                ws.cell(row, 16).value = measured_data[4][0]  # AGC_RX0
                ws.cell(row, 17).value = measured_data[4][1]  # AGC_RX1
                ws.cell(row, 18).value = measured_data[4][2]  # AGC_RX2
                ws.cell(row, 19).value = measured_data[4][3]  # AGC_RX3
                ws.cell(row, 20).value = tx_path
                ws.cell(row, 21).value = bw
                ws.cell(row, 22).value = rb_size
                ws.cell(row, 23).value = rb_start
                ws.cell(row, 24).value = ext_pmt.condition
                ws.cell(row, 25).value = measured_data[5][0]  # thermister 0
                ws.cell(row, 26).value = measured_data[5][1]  # thermister 1

                row += 1

        elif tech == 'FR1':
            max_row = ws.max_row
            row = max_row + 1  # skip title
            for tx_freq, measured_data in data.items():
                chan = chan_judge_fr1(band, bw, tx_freq)
                ws.cell(row, 1).value = band
                ws.cell(row, 2).value = rx_path_fr1_dict[rx_path]
                ws.cell(row, 3).value = chan  # LMH
                ws.cell(row, 4).value = tx_freq
                ws.cell(row, 5).value = tx_level  # this tx level
                ws.cell(row, 6).value = measured_data[0]  # measured power
                ws.cell(row, 7).value = measured_data[1]  # RX level
                ws.cell(row, 8).value = measured_data[2][0]  # RSRP_RX0
                ws.cell(row, 9).value = measured_data[2][1]  # RSRP_RX1
                ws.cell(row, 10).value = measured_data[2][2]  # RSRP_RX2
                ws.cell(row, 11).value = measured_data[2][3]  # RSRP_RX3
                ws.cell(row, 12).value = measured_data[3][0]  # CINR_RX0
                ws.cell(row, 13).value = measured_data[3][1]  # CINR_RX1
                ws.cell(row, 14).value = measured_data[3][2]  # CINR_RX2
                ws.cell(row, 15).value = measured_data[3][3]  # CINR_RX3
                ws.cell(row, 16).value = measured_data[4][0]  # AGC_RX0
                ws.cell(row, 17).value = measured_data[4][1]  # AGC_RX1
                ws.cell(row, 18).value = measured_data[4][2]  # AGC_RX2
                ws.cell(row, 19).value = measured_data[4][3]  # AGC_RX3
                ws.cell(row, 20).value = tx_path
                ws.cell(row, 21).value = bw
                ws.cell(row, 22).value = rb_size
                ws.cell(row, 23).value = rb_start
                ws.cell(row, 24).value = ext_pmt.condition
                ws.cell(row, 25).value = measured_data[5][0]  # thermister 0
                ws.cell(row, 26).value = measured_data[5][1]  # thermister 1

                row += 1

        elif tech == 'WCDMA':
            max_row = ws.max_row
            row = max_row + 1  # skip title
            for tx_chan, rx_level in data.items():
                chan = chan_judge_wcdma(band, cm_pmt_ftm.transfer_chan2freq_wcdma(band, tx_chan))
                ws.cell(row, 1).value = band
                ws.cell(row, 2).value = rx_path_wcdma_dict[rx_path]
                ws.cell(row, 3).value = chan  # LMH
                ws.cell(row, 4).value = tx_chan  # channel
                ws.cell(row, 5).value = cm_pmt_ftm.transfer_chan2freq_wcdma(band, tx_chan, 'tx')  # freq_tx
                ws.cell(row, 6).value = rx_level
                row += 1

        elif tech == 'GSM':
            max_row = ws.max_row
            row = max_row + 1
            for rx_chan, rx_level_rssi in data.items():
                chan = chan_judge_gsm(band, cm_pmt_ftm.transfer_chan2freq_gsm(band, rx_chan))
                ws.cell(row, 1).value = band
                ws.cell(row, 2).value = rx_path_gsm_dict[rx_path]
                ws.cell(row, 3).value = chan  # LMH
                ws.cell(row, 4).value = rx_chan  # channel
                ws.cell(row, 5).value = cm_pmt_ftm.transfer_chan2freq_gsm(band, rx_chan, 'rx')  # freq_rx
                ws.cell(row, 6).value = rx_level_rssi[0]
                ws.cell(row, 7).value = rx_level_rssi[1]
                row += 1

        wb.save(file_path)
        wb.close()

        return file_path


def rx_power_endc_test_export_excel_ftm(data):
    """
    :param data:  data = [int(band_lte), int(band_fr1), power_endc_lte, power_endc_fr1,
    rxs_lte, rxs_fr1, bw_lte, bw_fr1, tx_freq_lte, tx_freq_fr1, tx_level_endc_lte,
    tx_level_endc_fr1, rb_size_lte, rb_start_lte, rb_size_fr1, rb_start_fr1]
    :return:
    """
    logger.info('----------save to excel----------')
    filename = select_file_name_endc_ftm()

    file_path = Path(excel_folder_path()) / Path(filename)

    if Path(file_path).exists() is False:
        logger.info('----------file does not exist, so create a new one----------')
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        # create dashboard
        wb.create_sheet(f'Dashboard')

        # create the title and sheet for TxManx and -10dBm
        wb.create_sheet(f'Raw_Data_ENDC_FR1_TxMax')
        wb.create_sheet(f'Raw_Data_ENDC_FR1_-10dBm')
        for sheetname in wb.sheetnames:
            if 'Raw_Data' in sheetname:
                ws = wb[sheetname]
                ws['A1'] = 'Band_LTE'
                ws['B1'] = 'Band_FR1'
                ws['C1'] = 'Power_LTE_measured'
                ws['D1'] = 'Power_FR1_measured'
                ws['E1'] = 'Sensitivity_LTE'
                ws['F1'] = 'Sensitivity_FR1_RX0'
                ws['G1'] = 'Sensitivity_FR1_RX1'
                ws['H1'] = 'Sensitivity_FR1_RX2'
                ws['I1'] = 'Sensitivity_FR1_RX3'
                ws['J1'] = 'BW_LTE'
                ws['K1'] = 'BW_FR1'
                ws['L1'] = 'Freq_tx_LTE'
                ws['M1'] = 'Freq_tx_FR1'
                ws['N1'] = 'Tx_level_LTE'
                ws['O1'] = 'Tx_level_FR1'
                ws['P1'] = 'rb_size_LTE'
                ws['Q1'] = 'rb_start_LTE'
                ws['R1'] = 'rb_size_FR1'
                ws['S1'] = 'rb_start_FR1'
                ws['T1'] = 'LTE_RX_PATH'
                # ws['U1'] = 'FR1_RX_PATH'

            else:
                pass

        # create the title and sheet for Desense
        wb.create_sheet(f'Desens_ENDC')
        ws = wb[f'Desens_ENDC']
        ws['A1'] = 'Band_LTE'
        ws['B1'] = 'Band_FR1'
        ws['C1'] = 'BW_LTE'
        ws['D1'] = 'BW_FR1'
        ws['E1'] = 'Freq_tx_LTE'
        ws['F1'] = 'Freq_tx_FR1'
        ws['G1'] = 'LTE_RX_PATH'
        # ws['K1'] = 'FR1_TX_PATH'
        ws['H1'] = 'Diff_LTE'
        ws['I1'] = 'Diff_FR1_RX0'
        ws['J1'] = 'Diff_FR1_RX1'
        ws['K1'] = 'Diff_FR1_RX2'
        ws['L1'] = 'Diff_FR1_RX3'

        wb.save(file_path)
        wb.close()

    logger.info('----------file exist----------')
    wb = openpyxl.load_workbook(file_path)

    sheetname = f'Raw_Data_ENDC_FR1_TxMax' if data[2] > 0 else f'Raw_Data_ENDC_FR1_-10dBm'
    ws = wb[sheetname]
    max_row = ws.max_row
    max_col = ws.max_column
    row = max_row + 1
    for col in range(max_col):
        ws.cell(row, col + 1).value = data[col]

    wb.save(file_path)
    wb.close()

    return file_path


def rx_desense_process_ftm(file_path, mcs):
    """
    cell column: Band	| Rx_Path | Chan | Diff | TX_Path
    """
    wb = openpyxl.load_workbook(file_path)
    ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
    ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']
    ws_desens = wb[f'Desens_{mcs}']
    for row in range(2, ws_txmax.max_row + 1):
        ws_desens.cell(row, 1).value = ws_txmax.cell(row, 1).value
        ws_desens.cell(row, 2).value = ws_txmax.cell(row, 2).value
        ws_desens.cell(row, 3).value = ws_txmax.cell(row, 3).value
        ws_desens.cell(row, 4).value = ws_txmax.cell(row, 7).value - ws_txmin.cell(row, 7).value
        ws_desens.cell(row, 5).value = ws_txmax.cell(row, 20).value

    wb.save(file_path)
    wb.close()


def rx_desense_endc_process_ftm(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws_txmax = wb[f'Raw_Data_ENDC_FR1_TxMax']
    ws_txmin = wb[f'Raw_Data_ENDC_FR1_-10dBm']
    ws_desens = wb[f'Desens_ENDC']
    for row in range(2, ws_txmax.max_row + 1):
        ws_desens.cell(row, 1).value = ws_txmax.cell(row, 1).value  # 'Band_LTE'
        ws_desens.cell(row, 2).value = ws_txmax.cell(row, 2).value  # 'Band_FR1'
        ws_desens.cell(row, 3).value = ws_txmax.cell(row, 10).value  # 'BW_LTE'
        ws_desens.cell(row, 4).value = ws_txmax.cell(row, 11).value  # 'BW_FR1'
        ws_desens.cell(row, 5).value = ws_txmax.cell(row, 12).value  # 'Freq_tx_LTE'
        ws_desens.cell(row, 6).value = ws_txmax.cell(row, 13).value  # 'Freq_tx_FR1'
        ws_desens.cell(row, 7).value = ws_txmax.cell(row, 20).value  # 'LTE_RX_PATH'
        # ws_desens.cell(row, 8).value = ws_txmax.cell(row, 18).value  # 'FR1_RX_PATH'
        ws_desens.cell(row, 8).value = ws_txmax.cell(row, 5).value - ws_txmin.cell(row, 5).value  # desens lte
        ws_desens.cell(row, 9).value = ws_txmax.cell(row, 6).value - ws_txmin.cell(row, 6).value  # desens fr1_rx0
        ws_desens.cell(row, 10).value = ws_txmax.cell(row, 7).value - ws_txmin.cell(row, 7).value  # desens fr1_rx1
        ws_desens.cell(row, 11).value = ws_txmax.cell(row, 8).value - ws_txmin.cell(row, 8).value  # desens fr1_rx2
        ws_desens.cell(row, 12).value = ws_txmax.cell(row, 9).value - ws_txmin.cell(row, 9).value  # desens fr1_rx3

    wb.save(file_path)
    wb.close()


def rxs_relative_plot_ftm(file_path, parameters_dict):
    logger.info('----------Plot Chart---------')
    # tech ='LTE'
    # mcs_lte = 'QPSK'
    script = parameters_dict['script']
    tech = parameters_dict['tech']
    mcs = parameters_dict['mcs']

    wb = openpyxl.load_workbook(file_path)
    if script == 'GENERAL':
        if tech == 'LTE':
            ws_dashboard = wb[f'Dashboard']
            ws_desens = wb[f'Desens_{mcs}']
            ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
            ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']

            if ws_dashboard._charts:  # if there is charts, delete it
                ws_dashboard._charts.clear()

            chart1 = LineChart()
            chart1.title = 'Sensitivity'
            chart1.y_axis.title = 'Rx_Level(dBm)'
            chart1.x_axis.title = 'Band'
            chart1.x_axis.tickLblPos = 'low'
            chart1.height = 20
            chart1.width = 32
            y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
            y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
            y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
            x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

            series_txmax = Series(y_data_txmax, title="Tx_Max")
            series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

            chart1.append(series_txmax)
            chart1.append(series_txmin)
            chart1.set_categories(x_data)
            chart1.y_axis.majorGridlines = None

            chart2 = BarChart()
            chart2.add_data(y_data_desens, titles_from_data=True)
            chart2.y_axis.axId = 200
            chart2.y_axis.title = 'Diff(dB)'

            chart1.y_axis.crosses = "max"
            chart1 += chart2
            # save at dashboard sheet
            ws_dashboard.add_chart(chart1, "A1")

            wb.save(file_path)
            wb.close()

        elif tech == 'FR1':
            ws_dashboard = wb[f'Dashboard']
            ws_desens = wb[f'Desens_{mcs}']
            ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
            ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']

            if ws_dashboard._charts:  # if there is charts, delete it
                ws_dashboard._charts.clear()

            chart1 = LineChart()
            chart1.title = 'Sensitivity'
            chart1.y_axis.title = 'Rx_Level(dBm)'
            chart1.x_axis.title = 'Band'
            chart1.x_axis.tickLblPos = 'low'
            chart1.height = 20
            chart1.width = 32
            y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
            y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
            y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
            x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

            series_txmax = Series(y_data_txmax, title="Tx_Max")
            series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

            chart1.append(series_txmax)
            chart1.append(series_txmin)
            chart1.set_categories(x_data)
            chart1.y_axis.majorGridlines = None

            chart2 = BarChart()
            chart2.add_data(y_data_desens, titles_from_data=True)
            chart2.y_axis.axId = 200
            chart2.y_axis.title = 'Diff(dB)'

            chart1.y_axis.crosses = "max"
            chart1 += chart2
            # save at dashboard sheet
            ws_dashboard.add_chart(chart1, "A1")

            wb.save(file_path)
            wb.close()

        elif tech == 'WCDMA':
            ws_dashboard = wb[f'Dashboard']
            ws = wb[f'Raw_Data']

            if ws_dashboard._charts:  # if there is charts, delete it
                ws_dashboard._charts.clear()

            chart1 = LineChart()
            chart1.title = 'Sensitivity'
            chart1.y_axis.title = 'Rx_Level(dBm)'
            chart1.x_axis.title = 'Band'
            chart1.x_axis.tickLblPos = 'low'
            chart1.height = 20
            chart1.width = 32
            y_data = Reference(ws, min_col=6, min_row=2, max_col=6, max_row=ws.max_row)
            x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
            # save at dashboard sheet
            series_pure_sens = Series(y_data, title="Pure_Sensititvity_FTM")

            chart1.append(series_pure_sens)
            chart1.set_categories(x_data)

            ws_dashboard.add_chart(chart1, "A1")

            wb.save(file_path)
            wb.close()

        elif tech == 'GSM':
            ws_dashboard = wb[f'Dashboard']
            ws = wb[f'Raw_Data']

            if ws_dashboard._charts:  # if there is charts, delete it
                ws_dashboard._charts.clear()

            chart1 = LineChart()
            chart1.title = 'Sensitivity'
            chart1.y_axis.title = 'Rx_Level(dBm)'
            chart1.x_axis.title = 'Band'
            chart1.x_axis.tickLblPos = 'low'
            chart1.height = 20
            chart1.width = 32
            y_data = Reference(ws, min_col=6, min_row=1, max_col=7, max_row=ws.max_row)
            x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)

            chart1.add_data(y_data, titles_from_data=True)

            chart1.set_categories(x_data)
            # save at dashboard sheet
            ws_dashboard.add_chart(chart1, "A1")

            wb.save(file_path)
            wb.close()


def rxs_endc_plot_ftm(file_path):
    logger.info('----------Plot Chart ENDC LTE ---------')
    wb = openpyxl.load_workbook(file_path)
    ws_dashboard = wb[f'Dashboard']
    ws_desens = wb[f'Desens_ENDC']
    ws_txmax = wb[f'Raw_Data_ENDC_FR1_TxMax']
    ws_txmin = wb[f'Raw_Data_ENDC_FR1_-10dBm']

    if ws_dashboard._charts:  # if there is charts, delete it
        ws_dashboard._charts.clear()

    chart1 = LineChart()
    chart1.title = 'Sensitivity_LTE'
    chart1.y_axis.title = 'Rx_Level(dBm)'
    chart1.x_axis.title = 'Band'
    chart1.x_axis.tickLblPos = 'low'
    chart1.height = 20
    chart1.width = 32
    y_data_txmax = Reference(ws_txmax, min_col=5, min_row=2, max_col=5, max_row=ws_txmax.max_row)
    y_data_txmin = Reference(ws_txmin, min_col=5, min_row=2, max_col=5, max_row=ws_txmin.max_row)
    y_data_desens = Reference(ws_desens, min_col=9, min_row=1, max_col=9, max_row=ws_desens.max_row)
    x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=8, max_row=ws_desens.max_row)

    series_txmax = Series(y_data_txmax, title="Tx_Max")
    series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

    chart1.append(series_txmax)
    chart1.append(series_txmin)
    chart1.set_categories(x_data)
    chart1.y_axis.majorGridlines = None

    chart2 = BarChart()
    chart2.add_data(y_data_desens, titles_from_data=True)
    chart2.y_axis.axId = 200
    chart2.y_axis.title = 'Diff(dB)'

    chart1.y_axis.crosses = "max"
    chart1 += chart2

    ws_dashboard.add_chart(chart1, "A40")

    logger.info('----------Plot Chart ENDC FR1 ---------')
    # wb = openpyxl.load_workbook(file_path)
    # # ws_dashboard = wb[f'Dashboard']
    # ws_desens = wb[f'Desens_ENDC']
    # ws_txmax = wb[f'Raw_Data_ENDC_FR1_TxMax']
    # ws_txmin = wb[f'Raw_Data_ENDC_FR1_-10dBm']

    # if ws_dashboard._charts:  # if there is charts, delete it
    #     ws_dashboard._charts.clear()

    chart1 = LineChart()
    chart1.title = 'Sensitivity_FR1'
    chart1.y_axis.title = 'Rx_Level(dBm)'
    chart1.x_axis.title = 'Band'
    chart1.x_axis.tickLblPos = 'low'
    chart1.height = 20
    chart1.width = 32
    y_data_txmax = Reference(ws_txmax, min_col=6, min_row=2, max_col=6, max_row=ws_txmax.max_row)
    y_data_txmin = Reference(ws_txmin, min_col=6, min_row=2, max_col=6, max_row=ws_txmin.max_row)
    y_data_desens = Reference(ws_desens, min_col=10, min_row=1, max_col=10, max_row=ws_desens.max_row)
    x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=8, max_row=ws_desens.max_row)

    series_txmax = Series(y_data_txmax, title="Tx_Max")
    series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

    chart1.append(series_txmax)
    chart1.append(series_txmin)
    chart1.set_categories(x_data)
    chart1.y_axis.majorGridlines = None

    chart2 = BarChart()
    chart2.add_data(y_data_desens, titles_from_data=True)
    chart2.y_axis.axId = 200
    chart2.y_axis.title = 'Diff(dB)'

    chart1.y_axis.crosses = "max"
    chart1 += chart2

    ws_dashboard.add_chart(chart1, "A1")

    wb.save(file_path)
    wb.close()


# =================== below is for anritsu ===================
def rx_desense_process_sig(tech, file_path, mcs='QPSK'):
    """
    cell column: Band	| Rx_Path | Chan | Diff | TX_Path
    """
    ws_txmax = None
    ws_txmin = None
    ws_desens = None
    wb = openpyxl.load_workbook(file_path)
    if tech == 'LTE':
        ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
        ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']
        ws_desens = wb[f'Desens_{mcs}']
    elif tech == 'WCDMA' or tech == 'GSM':
        wb = openpyxl.load_workbook(file_path)
        ws_txmax = wb[f'Raw_Data_TxMax']
        ws_txmin = wb[f'Raw_Data_-10dBm']
        ws_desens = wb[f'Desens']

    for row in range(2, ws_txmax.max_row + 1):
        ws_desens.cell(row, 1).value = ws_txmax.cell(row, 1).value
        ws_desens.cell(row, 2).value = ws_txmax.cell(row, 2).value
        ws_desens.cell(row, 3).value = ws_txmax.cell(row, 3).value
        sens_Txmax = ws_txmax.cell(row, 7).value if ws_txmax.cell(row, 7).value is not None else 0
        sens_Txmin = ws_txmin.cell(row, 7).value if ws_txmin.cell(row, 7).value is not None else 0
        ws_desens.cell(row, 4).value = sens_Txmax - sens_Txmin
        ws_desens.cell(row, 5).value = ws_txmax.cell(row, 20).value

    wb.save(file_path)
    wb.close()


def rx_power_relative_test_export_excel_sig(data, parameters_dict):
    """
    data = sens_list = [power, sensitivity, per]

    """
    tech = parameters_dict['tech']
    band = parameters_dict['band']
    bw = parameters_dict['bw']
    tx_level = parameters_dict['tx_level']
    mcs = parameters_dict['mcs']
    # tx_path = parameters_dict['tx_path']
    rx_path = parameters_dict['rx_path']
    rb_size = parameters_dict['rb_size']
    rb_start = parameters_dict['rb_start']
    thermal = parameters_dict['thermal']
    dl_ch = parameters_dict['dl_ch']
    tx_freq = parameters_dict['tx_freq']

    logger.info('----------save to excel----------')
    filename = select_file_name_rx_sig(bw, tech)
    file_path = Path(excel_folder_path()) / Path(filename)

    if Path(file_path).exists() is False:  # if the file does not exist
        logger.info('----------file does not exist----------')
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        # create dashboard
        wb.create_sheet(f'Dashboard')

        # to create sheet
        if tech == 'LTE':
            # create the title and sheet for TxManx and -10dBm
            wb.create_sheet(f'Raw_Data_{mcs}_TxMax')
            wb.create_sheet(f'Raw_Data_{mcs}_-10dBm')
            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'RX_Path'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Tx_Freq'
                    ws['E1'] = 'Tx_level'
                    ws['F1'] = 'Measured Power'
                    ws['G1'] = 'Rx level'
                    ws['H1'] = 'TX_Path'
                    ws['I1'] = 'BW'
                    ws['J1'] = 'RB_num_UL'
                    ws['K1'] = 'RB_start_UL'
                    ws['L1'] = 'Condition'
                    ws['M1'] = 'Temp0'
                    ws['N1'] = 'Temp1'

                else:  # to skip dashboard
                    pass

            # create the title and sheet for Desense
            wb.create_sheet(f'Desens_{mcs}')
            ws = wb[f'Desens_{mcs}']
            ws['A1'] = 'Band'
            ws['B1'] = 'Rx_Path'
            ws['C1'] = 'Chan'
            ws['D1'] = 'Diff'
            ws['E1'] = 'TX_Path'

        elif tech == 'WCDMA' or tech == 'GSM':
            # create the title and sheet for TxManx and -10dBm
            wb.create_sheet(f'Raw_Data_TxMax')
            wb.create_sheet(f'Raw_Data_-10dBm')

            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'RX_Path'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Channel'
                    ws['E1'] = 'Tx_level'
                    ws['F1'] = 'Measured Power'
                    ws['G1'] = 'Rx level'
                    ws['H1'] = 'Condition'
                    ws['I1'] = 'Temp0'
                    ws['J1'] = 'Temp1'
                else:  # to skip dashboard
                    pass

            # create the title and sheet for Desense
            wb.create_sheet(f'Desens')
            ws = wb[f'Desens']
            ws['A1'] = 'Band'
            ws['B1'] = 'Rx_Path'
            ws['C1'] = 'Chan'
            ws['D1'] = 'Diff'
            ws['E1'] = 'TX_Path'

        # save and close file
        wb.save(file_path)
        wb.close()

    # if the file exist
    logger.info('----------file exist----------')
    wb = openpyxl.load_workbook(file_path)
    ws = None
    # to fetch the sheet name
    if tech == 'LTE':
        sheetname = f'Raw_Data_{mcs}_TxMax' if tx_level > 0 else f'Raw_Data_{mcs}_-10dBm'
        ws = wb[sheetname]
    elif tech == 'WCDMA' or tech == 'GSM':
        sheetname = f'Raw_Data_TxMax' if tx_level > 0 else f'Raw_Data_-10dBm'
        ws = wb[sheetname]

    if tech == 'LTE':
        max_row = ws.max_row
        row = max_row + 1  # skip title

        chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch, bw)
        ws.cell(row, 1).value = band
        ws.cell(row, 2).value = rx_path_lte_dict[rx_path] if isinstance(rx_path, int) else rx_path
        ws.cell(row, 3).value = chan  # LMH
        ws.cell(row, 4).value = tx_freq
        ws.cell(row, 5).value = tx_level  # this tx level
        ws.cell(row, 6).value = data[0]  # measured power
        ws.cell(row, 7).value = data[1]  # RX level
        ws.cell(row, 8).value = None  # tx_path
        ws.cell(row, 9).value = bw
        ws.cell(row, 10).value = rb_size
        ws.cell(row, 11).value = rb_start
        ws.cell(row, 12).value = ext_pmt.condition
        ws.cell(row, 13).value = thermal[0]  # thermister 0
        ws.cell(row, 14).value = thermal[1]  # thermister 1

        row += 1

    elif tech == 'WCDMA' or tech == 'GSM':
        tx_chan = cm_pmt_ftm.transfer_chan_rx2tx_wcdma(band, dl_ch)

        max_row = ws.max_row
        row = max_row + 1  # skip title

        chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch)
        ws.cell(row, 1).value = band
        ws.cell(row, 2).value = rx_path_wcdma_dict[rx_path] if isinstance(rx_path, int) else rx_path
        ws.cell(row, 3).value = chan  # LMH
        ws.cell(row, 4).value = tx_chan  # channel
        ws.cell(row, 5).value = tx_level  # freq_tx
        ws.cell(row, 6).value = data[0]  # measured Power
        ws.cell(row, 7).value = data[1]  # Rx level
        ws.cell(row, 8).value = ext_pmt.condition
        ws.cell(row, 9).value = thermal[0]  # thermister 0
        ws.cell(row, 10).value = thermal[1]  # thermister 1
        row += 1

    wb.save(file_path)
    wb.close()

    return file_path


def rx_freq_sweep_power_relative_test_export_excel_sig(data, parameters_dict):
    """
    data = sens_list = [power, sensitivity, per]

    """
    tech = parameters_dict['tech']
    band = parameters_dict['band']
    bw = parameters_dict['bw']
    tx_level = parameters_dict['tx_level']
    mcs = parameters_dict['mcs']
    # tx_path = parameters_dict['tx_path']
    rx_path = parameters_dict['rx_path']
    rb_size = parameters_dict['rb_size']
    rb_start = parameters_dict['rb_start']
    thermal = parameters_dict['thermal']
    dl_ch = parameters_dict['dl_ch']
    tx_freq = parameters_dict['tx_freq']

    logger.info('----------save to excel----------')
    filename = select_file_name_rx_freq_sweep_sig(bw, tech)
    file_path = Path(excel_folder_path()) / Path(filename)

    if Path(file_path).exists() is False:  # if the file does not exist
        logger.info('----------file does not exist----------')
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        # create dashboard
        wb.create_sheet(f'Dashboard')

        # to create sheet
        if tech == 'LTE':
            # create the title and sheet for TxManx and -10dBm
            wb.create_sheet(f'Raw_Data_{mcs}_TxMax')
            wb.create_sheet(f'Raw_Data_{mcs}_-10dBm')
            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'RX_Path'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Tx_Freq'
                    ws['E1'] = 'Tx_level'
                    ws['F1'] = 'Measured Power'
                    ws['G1'] = 'Rx level'
                    ws['H1'] = 'TX_Path'
                    ws['I1'] = 'BW'
                    ws['J1'] = 'RB_num_UL'
                    ws['K1'] = 'RB_start_UL'
                    ws['L1'] = 'Condition'
                    ws['M1'] = 'Temp0'
                    ws['N1'] = 'Temp1'

                else:  # to skip dashboard
                    pass

            # create the title and sheet for Desense
            wb.create_sheet(f'Desens_{mcs}')
            ws = wb[f'Desens_{mcs}']
            ws['A1'] = 'Band'
            ws['B1'] = 'Rx_Path'
            ws['C1'] = 'Chan'
            ws['D1'] = 'Diff'
            ws['E1'] = 'TX_Path'

        elif tech == 'WCDMA' or tech == 'GSM':
            # create the title and sheet for TxManx and -10dBm
            wb.create_sheet(f'Raw_Data_TxMax')
            wb.create_sheet(f'Raw_Data_-10dBm')

            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'RX_Path'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Channel'
                    ws['E1'] = 'Tx_level'
                    ws['F1'] = 'Measured Power'
                    ws['G1'] = 'Rx level'
                    ws['H1'] = 'Condition'
                    ws['I1'] = 'Temp0'
                    ws['J1'] = 'Temp1'
                else:  # to skip dashboard
                    pass

            # create the title and sheet for Desense
            wb.create_sheet(f'Desens')
            ws = wb[f'Desens']
            ws['A1'] = 'Band'
            ws['B1'] = 'Rx_Path'
            ws['C1'] = 'Chan'
            ws['D1'] = 'Diff'
            ws['E1'] = 'TX_Path'

        # save and close file
        wb.save(file_path)
        wb.close()

    # if the file exist
    logger.info('----------file exist----------')
    wb = openpyxl.load_workbook(file_path)
    ws = None
    # to fetch the sheet name
    if tech == 'LTE':
        sheetname = f'Raw_Data_{mcs}_TxMax' if tx_level > 0 else f'Raw_Data_{mcs}_-10dBm'
        ws = wb[sheetname]
    elif tech == 'WCDMA' or tech == 'GSM':
        sheetname = f'Raw_Data_TxMax' if tx_level > 0 else f'Raw_Data_-10dBm'
        ws = wb[sheetname]

    if tech == 'LTE':
        max_row = ws.max_row
        row = max_row + 1  # skip title

        # chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch, bw)
        ws.cell(row, 1).value = band
        ws.cell(row, 2).value = rx_path_lte_dict[rx_path] if isinstance(rx_path, int) else rx_path
        ws.cell(row, 3).value = None  # LMH
        ws.cell(row, 4).value = tx_freq
        ws.cell(row, 5).value = tx_level  # this tx level
        ws.cell(row, 6).value = data[0]  # measured power
        ws.cell(row, 7).value = data[1]  # RX level
        ws.cell(row, 8).value = None  # tx_path
        ws.cell(row, 9).value = bw
        ws.cell(row, 10).value = rb_size
        ws.cell(row, 11).value = rb_start
        ws.cell(row, 12).value = ext_pmt.condition
        ws.cell(row, 13).value = thermal[0]  # thermister 0
        ws.cell(row, 14).value = thermal[1]  # thermister 1

        row += 1

    elif tech == 'WCDMA' or tech == 'GSM':
        tx_chan = cm_pmt_ftm.transfer_chan_rx2tx_wcdma(band, dl_ch)

        max_row = ws.max_row
        row = max_row + 1  # skip title

        # chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch)
        ws.cell(row, 1).value = band
        ws.cell(row, 2).value = rx_path_wcdma_dict[rx_path] if isinstance(rx_path, int) else rx_path
        ws.cell(row, 3).value = None  # LMH
        ws.cell(row, 4).value = tx_chan  # channel
        ws.cell(row, 5).value = tx_level  # freq_tx
        ws.cell(row, 6).value = data[0]  # measured Power
        ws.cell(row, 7).value = data[1]  # Rx level
        ws.cell(row, 8).value = ext_pmt.condition
        ws.cell(row, 9).value = thermal[0]  # thermister 0
        ws.cell(row, 10).value = thermal[1]  # thermister 1
        row += 1

    wb.save(file_path)
    wb.close()

    return file_path


def tx_power_relative_test_export_excel_sig(data, parameters_dict):
    """
    LTE data:
    Only measure RB@min
    The format in dictionary is {Q_1: [power, (rb_size, rb_start)],
    Q_P: [power, aclr, evm, (rb_size, rb_start)], Q_F: [power, aclr, evm, (rb_size, rb_start)],
    16_P: [power, aclr, evm, (rb_size, rb_start)], 16_F: [power, aclr, evm, (rb_size, rb_start)], ...}
    and ACLR format is [EUTRA-1, EUTRA+1, UTRA-1, URTA+1, UTRA-2, URTA+2,]

    WCDMA data:
    [power, aclr, evm]
    and ACLR format is [TRA-1, URTA+1, UTRA-2, URTA+2,]
    """
    tech = parameters_dict['standard']
    band = parameters_dict['band']
    bw = parameters_dict['bw']
    dl_ch = parameters_dict['dl_ch']
    tx_freq = parameters_dict['tx_freq']
    # tx_freq_level = parameters_dict['tx_freq_level']
    # mcs = parameters_dict['mcs']
    # tx_path = parameters_dict['tx_path']
    # mod = parameters_dict['mod']
    # rb_size = parameters_dict['rb_size']
    # rb_start = parameters_dict['rb_start']
    # sync_path = parameters_dict['sync_path']
    # asw_srs_path = parameters_dict['asw_srs_path']
    # scs = parameters_dict['scs']
    # type_ = parameters_dict['type']
    # test_item = parameters_dict['test_item']
    thermal = parameters_dict['thermal']
    chcoding = parameters_dict['chcoding']
    logger.info('----------save to excel----------')

    filename = select_file_name_genre_tx_sig(bw, tech, chcoding)

    file_path = Path(excel_folder_path()) / Path(filename)

    if Path(file_path).exists() is False:
        logger.info('----------file does not exist----------')
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        # to create sheet
        if tech == 'LTE':
            # create dashboard
            for _ in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw100 might not have licesnse of Q256
                wb.create_sheet(f'Dashboard_{_}')
                # wb.create_sheet(f'Dashboard_{_}_PRB')
                # wb.create_sheet(f'Dashboard_{_}_FRB')

            # create the Raw data sheets
            for _ in ['QPSK', 'Q16', 'Q64', 'Q256']:  # some cmw100 might not have licesnse of Q256
                wb.create_sheet(f'Raw_Data_{_}')
                # wb.create_sheet(f'Raw_Data_{_}_PRB')
                # wb.create_sheet(f'Raw_Data_{_}_FRB')

            # create the title for every sheets
            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'BW'
                    ws['C1'] = 'Tx_Freq'
                    ws['D1'] = 'Chan'
                    ws['E1'] = 'Tx_level'
                    ws['F1'] = 'Measured_Power'
                    ws['G1'] = 'E_-1'
                    ws['H1'] = 'E_+1'
                    ws['I1'] = 'U_-1'
                    ws['J1'] = 'U_+1'
                    ws['K1'] = 'U_-2'
                    ws['L1'] = 'U_+2'
                    ws['M1'] = 'EVM'
                    ws['N1'] = 'Freq_Err'
                    ws['O1'] = 'IQ_OFFSET'
                    ws['P1'] = 'RB_num'
                    ws['Q1'] = 'RB_start'
                    ws['R1'] = 'MCS'
                    ws['S1'] = 'Tx_Path'
                    ws['T1'] = 'RB_STATE'
                    ws['U1'] = 'Sync_Path'
                    ws['V1'] = 'AS_Path'
                    ws['W1'] = 'Current(mA)'
                    ws['X1'] = 'Condition'
                    ws['Y1'] = 'Temp0'
                    ws['Z1'] = 'Temp1'
                else:  # to pass the dashboard
                    pass

        elif tech == 'WCDMA':
            # create dashboard
            wb.create_sheet(f'Dashboard')

            # create the Raw data sheets
            wb.create_sheet(f'Raw_Data')

            # create the title for every sheets
            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'Channel'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Tx_Freq'
                    ws['E1'] = 'Tx_level'
                    ws['F1'] = 'Measured_Power'
                    ws['G1'] = 'U_-1'
                    ws['H1'] = 'U_+1'
                    ws['I1'] = 'U_-2'
                    ws['J1'] = 'U_+2'
                    ws['K1'] = 'OBW'
                    ws['L1'] = 'EVM'
                    ws['M1'] = 'Freq_Err'
                    ws['N1'] = 'IQ_OFFSET'
                    ws['O1'] = 'Tx_Path'
                    ws['P1'] = 'AS_Path'
                    ws['Q1'] = 'Current(mA)'
                    ws['R1'] = 'Condition'
                    ws['S1'] = 'Temp0'
                    ws['T1'] = 'Temp1'
                    ws['U1'] = 'Subtest'
                else:  # to pass the dashboard
                    pass

        elif tech == 'GSM':  # this is not yet done, just placeholder
            # create dashboard
            wb.create_sheet(f'Dashboard_GMSK')
            wb.create_sheet(f'Dashboard_EPSK')

            # create the Raw data sheets
            wb.create_sheet(f'Raw_Data_GMSK')
            wb.create_sheet(f'Raw_Data_EPSK')

            # creat the title for every sheets
            for sheetname in wb.sheetnames:
                if 'Raw_Data' in sheetname:
                    ws = wb[sheetname]
                    ws['A1'] = 'Band'
                    ws['B1'] = 'Channel'
                    ws['C1'] = 'Chan'
                    ws['D1'] = 'Rx_Freq'
                    ws['E1'] = 'Tx_PCL'
                    ws['F1'] = 'Measured_Power'
                    ws['G1'] = 'Phase_rms'
                    ws['H1'] = 'EVM_rms'
                    ws['I1'] = 'Ferr'
                    ws['J1'] = 'ORFS_MOD_-200'
                    ws['K1'] = 'ORFS_MOD_+200'
                    ws['L1'] = 'ORFS_MOD_-400'
                    ws['M1'] = 'ORFS_MOD_+400'
                    ws['N1'] = 'ORFS_MOD_-600'
                    ws['O1'] = 'ORFS_MOD_+600'
                    ws['P1'] = 'ORFS_SW_-400'
                    ws['Q1'] = 'ORFS_SW_+400'
                    ws['R1'] = 'ORFS_SW_-600'
                    ws['S1'] = 'ORFS_SW_+600'
                    ws['T1'] = 'ORFS_SW_-1200'
                    ws['U1'] = 'ORFS_SW_+1200'
                    ws['V1'] = 'AS_Path'
                    ws['W1'] = 'Current(mA)'
                    ws['X1'] = 'Condition'
                    ws['Y1'] = 'Temp0'
                    ws['Z1'] = 'Temp1'
                else:  # to pass the dashboard
                    pass

        # save and close file
        wb.save(file_path)
        wb.close()

    logger.info('----------file exist----------')
    wb = openpyxl.load_workbook(file_path)
    ws = None

    if tech == 'LTE':
        mod = None
        rb_state = None
        for mcs, measured_items in data.items():  # to seperate the data to mcs and sub-data
            mcs_list = mcs.split('_')
            if mcs_list[0] == 'Q':
                mod = 'QPSK'
            elif mcs_list[0] == '16':
                mod = 'Q16'
            elif mcs_list[0] == '64':
                mod = 'Q64'
            elif mcs_list[0] == '256':
                mod = 'Q256'

            if mcs_list[1] == 'P':
                rb_state = 'PRB'
            elif mcs_list[1] == 'F':
                rb_state = 'FRB'
            elif mcs_list[1] == '1':
                rb_state = '1RB'

            ws = wb[f'Raw_Data_{mod}']

            max_row = ws.max_row
            row = max_row + 1

            chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch, bw)
            ws.cell(row, 1).value = band
            ws.cell(row, 2).value = bw
            ws.cell(row, 3).value = tx_freq  # this freq_lte
            ws.cell(row, 4).value = chan  # LMH
            ws.cell(row, 5).value = ext_pmt.tx_level
            ws.cell(row, 6).value = measured_items[0]  # measured power
            ws.cell(row, 7).value = measured_items[1][0] if mcs != 'Q_1' else None  # 'E_-1'
            ws.cell(row, 8).value = measured_items[1][1] if mcs != 'Q_1' else None  # 'E_+1'
            ws.cell(row, 9).value = measured_items[1][2] if mcs != 'Q_1' else None  # 'U_-1'
            ws.cell(row, 10).value = measured_items[1][3] if mcs != 'Q_1' else None  # 'U_+1'
            ws.cell(row, 11).value = measured_items[1][4] if mcs != 'Q_1' else None  # 'U_-2'
            ws.cell(row, 12).value = measured_items[1][5] if mcs != 'Q_1' else None  # 'U_+2'
            ws.cell(row, 13).value = measured_items[2] if mcs != 'Q_1' else None  # 'EVM'
            ws.cell(row, 14).value = None  # 'Freq Err'
            ws.cell(row, 15).value = None  # 'IQ OFFSET'
            ws.cell(row, 16).value = measured_items[3][0] if mcs != 'Q_1' else measured_items[1][0]  # rb_size
            ws.cell(row, 17).value = measured_items[3][1] if mcs != 'Q_1' else measured_items[1][1]  # rb_start
            ws.cell(row, 18).value = mod  # mcs or modulation
            ws.cell(row, 19).value = None  # tx_path
            ws.cell(row, 20).value = rb_state  # rb_state
            ws.cell(row, 21).value = None  # sync_path
            ws.cell(row, 22).value = None  # asw_srs_path
            ws.cell(row, 23).value = None  # current
            ws.cell(row, 24).value = ext_pmt.condition
            ws.cell(row, 25).value = thermal[0]  # 'Temp0'
            ws.cell(row, 26).value = thermal[1]  # 'Temp1'
            row += 1

    elif tech == 'WCDMA':
        if chcoding == 'REFMEASCH':
            ws = wb[f'Raw_Data']
            max_row = ws.max_row
            row = max_row + 1

            chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch)
            ws.cell(row, 1).value = band
            ws.cell(row, 2).value = cm_pmt_ftm.transfer_chan_rx2tx_wcdma(band, dl_ch)  # this channel
            ws.cell(row, 3).value = chan  # LMH
            ws.cell(row, 4).value = tx_freq
            ws.cell(row, 5).value = ext_pmt.tx_level  # this tx_level
            ws.cell(row, 6).value = data[0]  # 'Measured_Power'
            ws.cell(row, 7).value = data[1][0]  # U_-1
            ws.cell(row, 8).value = data[1][1]  # U_+1
            ws.cell(row, 9).value = data[1][2]  # U_-2
            ws.cell(row, 10).value = data[1][3]  # U_+2
            ws.cell(row, 11).value = None  # 'OBW'
            ws.cell(row, 12).value = data[2]  # 'EVM'
            ws.cell(row, 13).value = None  # 'Freq_Err'
            ws.cell(row, 14).value = None  # 'IQ_OFFSET'
            ws.cell(row, 15).value = None  # 'Tx_Path'
            ws.cell(row, 16).value = None  # 'AS_Path'
            ws.cell(row, 17).value = None  # 'Current(mA)'
            ws.cell(row, 18).value = ext_pmt.condition  # 'Condition'
            ws.cell(row, 19).value = thermal[0]  # 'Temp0'
            ws.cell(row, 20).value = thermal[1]  # 'Temp1'
            ws.cell(row, 21).value = None  # subtest
            row += 1
        else:
            for subtest, data in data.items():
                ws = wb[f'Raw_Data']
                max_row = ws.max_row
                row = max_row + 1

                chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch)
                ws.cell(row, 1).value = band
                ws.cell(row, 2).value = cm_pmt_ftm.transfer_chan_rx2tx_wcdma(band, dl_ch)  # this channel
                ws.cell(row, 3).value = chan  # LMH
                ws.cell(row, 4).value = tx_freq
                ws.cell(row, 5).value = ext_pmt.tx_level  # this tx_level
                ws.cell(row, 6).value = data[0]  # 'Measured_Power'
                ws.cell(row, 7).value = data[1][0]  # U_-1
                ws.cell(row, 8).value = data[1][1]  # U_+1
                ws.cell(row, 9).value = data[1][2]  # U_-2
                ws.cell(row, 10).value = data[1][3]  # U_+2
                ws.cell(row, 11).value = None  # 'OBW'
                ws.cell(row, 12).value = data[2] if chcoding != 'EDCHTEST' and subtest == 3 else None  # 'EVM'
                ws.cell(row, 13).value = None  # 'Freq_Err'
                ws.cell(row, 14).value = None  # 'IQ_OFFSET'
                ws.cell(row, 15).value = None  # 'Tx_Path'
                ws.cell(row, 16).value = None  # 'AS_Path'
                ws.cell(row, 17).value = None  # 'Current(mA)'
                ws.cell(row, 18).value = ext_pmt.condition  # 'Condition'
                ws.cell(row, 19).value = thermal[0]  # 'Temp0'
                ws.cell(row, 20).value = thermal[1]  # 'Temp1'
                ws.cell(row, 21).value = subtest  # subtest
                row += 1

    elif tech == 'GSM':  # this is template, not verify and not implement
        for mcs, measured_items in data:  # to seperate the data to mcs and sub-data
            mcs_list = mcs.split('_')
            if mcs_list[0] == 'Q':
                mod = 'QPSK'
            elif mcs_list[0] == '16':
                mod = 'Q16'
            elif mcs_list[0] == '64':
                mod = 'Q64'
            elif mcs_list[0] == '256':
                mod = 'Q256'

            if mcs_list[1] == 'P':
                rb_state = 'PRB'
            elif mcs_list[1] == 'F':
                rb_state = 'FRB'

            ws = wb[f'Raw_Data_{mod}']
            max_row = ws.max_row
            row = max_row + 1

            chan = cm_pmt_anritsu.sig_ch_judge(tech, band, dl_ch)
            ws.cell(row, 1).value = band
            ws.cell(row, 2).value = cm_pmt_ftm.transfer_chan_rx2tx_wcdma(band, dl_ch)  # this channel
            ws.cell(row, 3).value = chan  # LMH
            ws.cell(row, 4).value = tx_freq
            ws.cell(row, 5).value = ext_pmt.tx_level  # this tx_level
            ws.cell(row, 6).value = measured_items[0]  # 'Measured_Power'
            ws.cell(row, 7).value = measured_items[1][0]  # U_-1
            ws.cell(row, 8).value = measured_items[1][1]  # U_+1
            ws.cell(row, 9).value = measured_items[1][2]  # U_-2
            ws.cell(row, 10).value = measured_items[1][3]  # U_+2
            ws.cell(row, 11).value = None  # 'OBW'
            ws.cell(row, 12).value = measured_items[2]  # 'EVM'
            ws.cell(row, 13).value = None  # 'Freq_Err'
            ws.cell(row, 14).value = None  # 'IQ_OFFSET'
            ws.cell(row, 15).value = None  # 'Tx_Path'
            ws.cell(row, 16).value = None  # 'AS_Path'
            ws.cell(row, 17).value = None  # 'Current(mA)'
            ws.cell(row, 18).value = ext_pmt.condition  # 'Condition'
            ws.cell(row, 19).value = thermal[0]  # 'Temp0'
            ws.cell(row, 20).value = thermal[1]  # 'Temp1'
            row += 1

    wb.save(file_path)
    wb.close()

    return file_path


def txp_aclr_evm_current_plot_sig(standard, file_path):
    # script = parameters_dict['script']
    tech = standard
    # band = parameters_dict['band']
    # bw = parameters_dict['bw']
    # tx_freq_level = parameters_dict['tx_freq_level']
    # mcs = parameters_dict['mcs']
    # tx_path = parameters_dict['tx_path']
    # mod = parameters_dict['mod']
    # rb_state = parameters_dict['rb_state']
    # rb_size = parameters_dict['rb_size']
    # rb_start = parameters_dict['rb_start']
    # sync_path = parameters_dict['sync_path']
    # asw_srs_path = parameters_dict['asw_srs_path']
    # scs = parameters_dict['scs']
    # type_ = parameters_dict['type']
    logger.info('----------Plot Chart---------')
    wb = openpyxl.load_workbook(file_path)
    if tech == 'LTE':
        for ws_name in wb.sheetnames:
            if 'Raw_Data' in ws_name:
                logger.info(f'========={ws_name}==========')
                ws = wb[ws_name]
                ws_dashboard = wb['Dashboard_' + ws_name[9:]]

                if ws_dashboard._charts:  # if there is charts, delete it
                    ws_dashboard._charts.clear()

                logger.info('----------Power---------')
                chart = LineChart()
                chart.title = 'Power'
                chart.y_axis.title = 'Power(dBm)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A1")

                logger.info('----------ACLR---------')
                chart = LineChart()
                chart.title = 'ACLR'
                chart.y_axis.title = 'ACLR(dB)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'
                chart.y_axis.scaling.min = -60
                chart.y_axis.scaling.max = -20

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=7, min_row=1, max_col=12, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                chart.series[0].marker.size = 10
                chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[1].marker.size = 10
                chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                ws_dashboard.add_chart(chart, "A41")

                logger.info('----------EVM---------')
                chart = LineChart()
                chart.title = 'EVM'
                chart.y_axis.title = 'EVM(%)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=13, min_row=1, max_col=13, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A81")

                logger.info('----------Current---------')
                chart = LineChart()
                chart.title = 'Current'
                chart.y_axis.title = 'Current(mA)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=23, min_row=1, max_col=23, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A121")
            else:
                pass

        wb.save(file_path)
        wb.close()

    elif tech == 'FR1':
        for ws_name in wb.sheetnames:
            if 'Raw_Data' in ws_name:
                logger.info(f'========={ws_name}==========')
                ws = wb[ws_name]
                ws_dashboard = wb['Dashboard_' + ws_name[9:]]

                if ws_dashboard._charts:  # if there is charts, delete it
                    ws_dashboard._charts.clear()

                logger.info('----------Power---------')
                chart = LineChart()
                chart.title = 'Power'
                chart.y_axis.title = 'Power(dBm)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A1")

                logger.info('----------ACLR---------')
                chart = LineChart()
                chart.title = 'ACLR'
                chart.y_axis.title = 'ACLR(dB)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'
                chart.y_axis.scaling.min = -60
                chart.y_axis.scaling.max = -20

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=7, min_row=1, max_col=12, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'triangle'  # for EUTRA_-1
                chart.series[0].marker.size = 10
                chart.series[1].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[1].marker.size = 10
                chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-1
                chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+1
                chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_-2
                chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for UTRA_+2

                ws_dashboard.add_chart(chart, "A41")

                logger.info('----------EVM---------')
                chart = LineChart()
                chart.title = 'EVM'
                chart.y_axis.title = 'EVM(%)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=13, min_row=1, max_col=13, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A81")

                logger.info('----------Current---------')
                chart = LineChart()
                chart.title = 'Current'
                chart.y_axis.title = 'Current(mA)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=24, min_row=1, max_col=24, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A121")

        wb.save(file_path)
        wb.close()

    elif tech == 'WCDMA':
        for ws_name in wb.sheetnames:
            if 'Raw_Data' in ws_name:
                logger.info(f'========={ws_name}==========')
                ws = wb[ws_name]
                ws_dashboard = wb['Dashboard']

                if ws_dashboard._charts:  # if there is charts, delete it
                    ws_dashboard._charts.clear()

                logger.info('----------Power---------')
                chart = LineChart()
                chart.title = 'Power'
                chart.y_axis.title = 'Power(dBm)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A1")

                logger.info('----------ACLR---------')
                chart = LineChart()
                chart.title = 'ACLR'
                chart.y_axis.title = 'ACLR(dB)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'
                chart.y_axis.scaling.min = -60
                chart.y_axis.scaling.max = -20

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=7, min_row=1, max_col=10, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'triangle'  # for UTRA_-1
                chart.series[0].marker.size = 10
                chart.series[1].marker.symbol = 'circle'  # for UTRA_+1
                chart.series[1].marker.size = 10
                chart.series[2].graphicalProperties.line.width = 50000  # for UTRA_-2
                chart.series[3].graphicalProperties.line.width = 50000  # for UTRA_+2

                ws_dashboard.add_chart(chart, "A41")

                logger.info('----------EVM---------')
                chart = LineChart()
                chart.title = 'EVM'
                chart.y_axis.title = 'EVM(%)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=12, min_row=1, max_col=12, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A81")

                logger.info('----------Current---------')
                chart = LineChart()
                chart.title = 'Current'
                chart.y_axis.title = 'Current(mA)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=17, min_row=1, max_col=17, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A121")

        wb.save(file_path)
        wb.close()
    elif tech == 'GSM':
        for ws_name in wb.sheetnames:
            if 'Raw_Data' in ws_name:
                logger.info(f'========={ws_name}==========')
                ws = wb[ws_name]
                ws_dashboard = wb['Dashboard_' + ws_name[9:]]

                if ws_dashboard._charts:  # if there is charts, delete it
                    ws_dashboard._charts.clear()

                logger.info('----------Power---------')
                chart = LineChart()
                chart.title = 'Power'
                chart.y_axis.title = 'Power(dBm)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'  # for EUTRA_+1
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A1")

                logger.info('----------ORFS_MOD---------')
                chart = LineChart()
                chart.title = 'ORFS_MOD'
                chart.y_axis.title = 'ORFS_MOD(dB)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'
                chart.y_axis.scaling.min = -80
                chart.y_axis.scaling.max = -20

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=10, min_row=1, max_col=15, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                chart.series[2].marker.size = 10
                chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                chart.series[3].marker.size = 10
                chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-200
                chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+200
                chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-600
                chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+600

                ws_dashboard.add_chart(chart, "A41")

                logger.info('----------ORFS_SW---------')
                chart = LineChart()
                chart.title = 'ORFS_SW'
                chart.y_axis.title = 'ORFS_SW(dBm)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'
                chart.y_axis.scaling.min = -80
                chart.y_axis.scaling.max = -20

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=16, min_row=1, max_col=21, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[2].marker.symbol = 'triangle'  # for ORFS_-400
                chart.series[2].marker.size = 10
                chart.series[3].marker.symbol = 'circle'  # for ORFS_+400
                chart.series[3].marker.size = 10
                chart.series[0].graphicalProperties.line.width = 50000  # for ORFS_-600
                chart.series[1].graphicalProperties.line.width = 50000  # for ORFS_+600
                chart.series[4].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_-1200
                chart.series[5].graphicalProperties.line.dashStyle = 'dash'  # for ORFS_+1200

                ws_dashboard.add_chart(chart, "A81")

                if 'GMSK' in ws_name:
                    logger.info('----------PHASE_RMS---------')
                    chart = LineChart()
                    chart.title = 'PHASE'
                    chart.y_axis.title = 'PHASE(degree)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=7, min_row=1, max_col=7, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A115")

                elif 'EPSK' in ws_name:
                    logger.info('----------EVM_RMS---------')
                    chart = LineChart()
                    chart.title = 'EVM'
                    chart.y_axis.title = 'EVM(%)'
                    chart.x_axis.title = 'Band'
                    chart.x_axis.tickLblPos = 'low'

                    chart.height = 20
                    chart.width = 32

                    y_data = Reference(ws, min_col=8, min_row=1, max_col=8, max_row=ws.max_row)
                    x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                    chart.add_data(y_data, titles_from_data=True)
                    chart.set_categories(x_data)

                    chart.series[0].marker.symbol = 'circle'
                    chart.series[0].marker.size = 10

                    ws_dashboard.add_chart(chart, "A121")

                logger.info('----------Current---------')
                chart = LineChart()
                chart.title = 'Current'
                chart.y_axis.title = 'Current(mA)'
                chart.x_axis.title = 'Band'
                chart.x_axis.tickLblPos = 'low'

                chart.height = 20
                chart.width = 32

                y_data = Reference(ws, min_col=23, min_row=1, max_col=23, max_row=ws.max_row)
                x_data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=ws.max_row)
                chart.add_data(y_data, titles_from_data=True)
                chart.set_categories(x_data)

                chart.series[0].marker.symbol = 'circle'
                chart.series[0].marker.size = 10

                ws_dashboard.add_chart(chart, "A161")

        wb.save(file_path)
        wb.close()


def rxs_relative_plot_sig(file_path, parameters_dict):
    logger.info('----------Plot Chart---------')
    # tech ='LTE'
    # mcs_lte = 'QPSK'
    # script = parameters_dict['script']
    tech = parameters_dict['tech']
    mcs = parameters_dict['mcs']

    wb = openpyxl.load_workbook(file_path)
    if tech == 'LTE':
        ws_dashboard = wb[f'Dashboard']
        ws_desens = wb[f'Desens_{mcs}']
        ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
        ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        chart1 = LineChart()
        chart1.title = 'Sensitivity'
        chart1.y_axis.title = 'Rx_Level(dBm)'
        chart1.x_axis.title = 'Band'
        chart1.x_axis.tickLblPos = 'low'
        chart1.height = 20
        chart1.width = 32
        y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
        y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
        y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
        x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

        series_txmax = Series(y_data_txmax, title="Tx_Max")
        series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

        chart1.append(series_txmax)
        chart1.append(series_txmin)
        chart1.set_categories(x_data)
        chart1.y_axis.majorGridlines = None

        chart2 = BarChart()
        chart2.add_data(y_data_desens, titles_from_data=True)
        chart2.y_axis.axId = 200
        chart2.y_axis.title = 'Diff(dB)'

        chart1.y_axis.crosses = "max"
        chart1 += chart2
        # save at dashboard sheet
        ws_dashboard.add_chart(chart1, "A1")

        wb.save(file_path)
        wb.close()

    elif tech == 'FR1':  # not yet have this function
        ws_dashboard = wb[f'Dashboard']
        ws_desens = wb[f'Desens_{mcs}']
        ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
        ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        chart1 = LineChart()
        chart1.title = 'Sensitivity'
        chart1.y_axis.title = 'Rx_Level(dBm)'
        chart1.x_axis.title = 'Band'
        chart1.x_axis.tickLblPos = 'low'
        chart1.height = 20
        chart1.width = 32
        y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
        y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
        y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
        x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

        series_txmax = Series(y_data_txmax, title="Tx_Max")
        series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

        chart1.append(series_txmax)
        chart1.append(series_txmin)
        chart1.set_categories(x_data)
        chart1.y_axis.majorGridlines = None

        chart2 = BarChart()
        chart2.add_data(y_data_desens, titles_from_data=True)
        chart2.y_axis.axId = 200
        chart2.y_axis.title = 'Diff(dB)'

        chart1.y_axis.crosses = "max"
        chart1 += chart2
        # save at dashboard sheet
        ws_dashboard.add_chart(chart1, "A1")

        wb.save(file_path)
        wb.close()

    elif tech == 'WCDMA':
        ws_dashboard = wb[f'Dashboard']
        ws_desens = wb[f'Desens']
        ws_txmax = wb[f'Raw_Data_TxMax']
        ws_txmin = wb[f'Raw_Data_-10dBm']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        chart1 = LineChart()
        chart1.title = 'Sensitivity'
        chart1.y_axis.title = 'Rx_Level(dBm)'
        chart1.x_axis.title = 'Band'
        chart1.x_axis.tickLblPos = 'low'
        chart1.height = 20
        chart1.width = 32
        y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
        y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
        y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
        x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

        series_txmax = Series(y_data_txmax, title="Tx_Max")
        series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

        chart1.append(series_txmax)
        chart1.append(series_txmin)
        chart1.set_categories(x_data)
        chart1.y_axis.majorGridlines = None

        chart2 = BarChart()
        chart2.add_data(y_data_desens, titles_from_data=True)
        chart2.y_axis.axId = 200
        chart2.y_axis.title = 'Diff(dB)'

        chart1.y_axis.crosses = "max"
        chart1 += chart2
        # save at dashboard sheet
        ws_dashboard.add_chart(chart1, "A1")

        wb.save(file_path)
        wb.close()

    elif tech == 'GSM':
        ws_dashboard = wb[f'Dashboard']
        ws = wb[f'Raw_Data']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        chart1 = LineChart()
        chart1.title = 'Sensitivity'
        chart1.y_axis.title = 'Rx_Level(dBm)'
        chart1.x_axis.title = 'Band'
        chart1.x_axis.tickLblPos = 'low'
        chart1.height = 20
        chart1.width = 32
        y_data = Reference(ws, min_col=6, min_row=1, max_col=7, max_row=ws.max_row)
        x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)

        chart1.add_data(y_data, titles_from_data=True)

        chart1.set_categories(x_data)
        # save at dashboard sheet
        ws_dashboard.add_chart(chart1, "A1")

        wb.save(file_path)
        wb.close()


def rxs_freq_relative_plot_sig(file_path, parameters_dict):
    logger.info('----------Plot Chart---------')
    # tech ='LTE'
    # mcs_lte = 'QPSK'
    # script = parameters_dict['script']
    tech = parameters_dict['tech']
    mcs = parameters_dict['mcs']

    wb = openpyxl.load_workbook(file_path)
    if tech == 'LTE':
        ws_dashboard = wb[f'Dashboard']
        ws_desens = wb[f'Desens_{mcs}']
        ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
        ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        for i, ws in enumerate([ws_txmax, ws_txmin]):
            chart_sens = LineChart()
            chart_sens.title = 'Rx_Freq_Chan_Sweep_TxMax' if i == 0 else 'Rx_Freq_Chan_Sweep_Sweep_-10dBm'
            chart_sens.y_axis.title = 'Sensitivity'
            chart_sens.x_axis.title = 'Rx_chan'
            chart_sens.x_axis.tickLblPos = 'low'
            # chart_sens.y_axis.scaling.min = -60
            # chart_sens.y_axis.scaling.max = -20

            chart_sens.height = 20
            chart_sens.width = 40

            y_data_sens = Reference(ws, min_col=7, min_row=1, max_col=7,
                                    max_row=ws.max_row)  # colume 7 is snesitivity

            x_data = Reference(ws, min_col=1, min_row=2, max_col=4, max_row=ws.max_row)
            chart_sens.add_data(y_data_sens, titles_from_data=True)
            chart_sens.set_categories(x_data)

            # chart_sens.y_axis.majorGridlines = None

            chart_sens.series[0].marker.symbol = 'circle'  # for sensitivity
            chart_sens.series[0].marker.size = 3

            chart_pwr = LineChart()  # create a second chart

            y_data_pwr = Reference(ws, min_col=6, min_row=1, max_col=6,
                                   max_row=ws.max_row)
            chart_pwr.add_data(y_data_pwr, titles_from_data=True)

            chart_pwr.series[0].graphicalProperties.line.dashStyle = 'dash'  # for power
            chart_pwr.y_axis.title = 'Power(dBm)'
            chart_pwr.y_axis.axId = 200
            chart_pwr.y_axis.majorGridlines = None

            chart_sens.y_axis.crosses = 'max'
            chart_sens += chart_pwr

            if i == 0:
                ws_dashboard.add_chart(chart_sens, "A1")
            elif i == 1:
                ws_dashboard.add_chart(chart_sens, "A42")

        wb.save(file_path)
        wb.close()

    elif tech == 'FR1':  # not yet has this function
        ws_dashboard = wb[f'Dashboard']
        ws_desens = wb[f'Desens_{mcs}']
        ws_txmax = wb[f'Raw_Data_{mcs}_TxMax']
        ws_txmin = wb[f'Raw_Data_{mcs}_-10dBm']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        chart1 = LineChart()
        chart1.title = 'Sensitivity'
        chart1.y_axis.title = 'Rx_Level(dBm)'
        chart1.x_axis.title = 'Band'
        chart1.x_axis.tickLblPos = 'low'
        chart1.height = 20
        chart1.width = 32
        y_data_txmax = Reference(ws_txmax, min_col=7, min_row=2, max_col=7, max_row=ws_txmax.max_row)
        y_data_txmin = Reference(ws_txmin, min_col=7, min_row=2, max_col=7, max_row=ws_txmin.max_row)
        y_data_desens = Reference(ws_desens, min_col=4, min_row=1, max_col=4, max_row=ws_desens.max_row)
        x_data = Reference(ws_desens, min_col=1, min_row=2, max_col=3, max_row=ws_desens.max_row)

        series_txmax = Series(y_data_txmax, title="Tx_Max")
        series_txmin = Series(y_data_txmin, title="Tx_-10dBm")

        chart1.append(series_txmax)
        chart1.append(series_txmin)
        chart1.set_categories(x_data)
        chart1.y_axis.majorGridlines = None

        chart2 = BarChart()
        chart2.add_data(y_data_desens, titles_from_data=True)
        chart2.y_axis.axId = 200
        chart2.y_axis.title = 'Diff(dB)'

        chart1.y_axis.crosses = "max"
        chart1 += chart2
        # save at dashboard sheet
        ws_dashboard.add_chart(chart1, "A1")

        wb.save(file_path)
        wb.close()

    elif tech == 'WCDMA':
        ws_dashboard = wb[f'Dashboard']
        ws_desens = wb[f'Desens']
        ws_txmax = wb[f'Raw_Data_TxMax']
        ws_txmin = wb[f'Raw_Data_-10dBm']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        for i, ws in enumerate([ws_txmax, ws_txmin]):
            chart_sens = LineChart()
            chart_sens.title = 'Rx_Freq_Chan_Sweep_TxMax' if i == 0 else 'Rx_Freq_Chan_Sweep_Sweep_-10dBm'
            chart_sens.y_axis.title = 'Sensitivity'
            chart_sens.x_axis.title = 'Rx_chan'
            chart_sens.x_axis.tickLblPos = 'low'
            # chart_sens.y_axis.scaling.min = -60
            # chart_sens.y_axis.scaling.max = -20

            chart_sens.height = 20
            chart_sens.width = 40

            y_data_sens = Reference(ws, min_col=7, min_row=1, max_col=7,
                                    max_row=ws.max_row)  # colume 7 is snesitivity

            x_data = Reference(ws, min_col=1, min_row=2, max_col=4, max_row=ws.max_row)
            chart_sens.add_data(y_data_sens, titles_from_data=True)
            chart_sens.set_categories(x_data)

            # chart_sens.y_axis.majorGridlines = None

            chart_sens.series[0].marker.symbol = 'circle'  # for sensitivity
            chart_sens.series[0].marker.size = 3

            chart_pwr = LineChart()  # create a second chart

            y_data_pwr = Reference(ws, min_col=6, min_row=1, max_col=6,
                                   max_row=ws.max_row)
            chart_pwr.add_data(y_data_pwr, titles_from_data=True)

            chart_pwr.series[0].graphicalProperties.line.dashStyle = 'dash'  # for power
            chart_pwr.y_axis.title = 'Power(dBm)'
            chart_pwr.y_axis.axId = 200
            chart_pwr.y_axis.majorGridlines = None

            chart_sens.y_axis.crosses = 'max'
            chart_sens += chart_pwr

            if i == 0:
                ws_dashboard.add_chart(chart_sens, "A1")
            elif i == 1:
                ws_dashboard.add_chart(chart_sens, "A42")

        wb.save(file_path)
        wb.close()

    elif tech == 'GSM':  # not yet has this function
        ws_dashboard = wb[f'Dashboard']
        ws = wb[f'Raw_Data']

        if ws_dashboard._charts:  # if there is charts, delete it
            ws_dashboard._charts.clear()

        chart1 = LineChart()
        chart1.title = 'Sensitivity'
        chart1.y_axis.title = 'Rx_Level(dBm)'
        chart1.x_axis.title = 'Band'
        chart1.x_axis.tickLblPos = 'low'
        chart1.height = 20
        chart1.width = 32
        y_data = Reference(ws, min_col=6, min_row=1, max_col=7, max_row=ws.max_row)
        x_data = Reference(ws, min_col=1, min_row=2, max_col=3, max_row=ws.max_row)

        chart1.add_data(y_data, titles_from_data=True)

        chart1.set_categories(x_data)
        # save at dashboard sheet
        ws_dashboard.add_chart(chart1, "A1")

        wb.save(file_path)
        wb.close()


def color_format_clear(file_path):  # clear all condition pattern
    logger.info('========== Clear Condition Pattern ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue

        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format

    wb.save(file_path)

def color_format_fr1_aclr_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # import yaml file
    color_codes = import_aclr_limits()
    aclr_red_usl = color_codes['FR1']['e_utra_color_red_usl']
    aclr_yellow_usl = color_codes['FR1']['e_utra_color_yellow_usl']

    # define the color of fill and font
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C')
    font_red = Font(color='9C0006')
    font_yellow = Font(color='9C6500')

    # define the rule
    rule_red = CellIsRule(operator='greaterThan', formula=[aclr_red_usl], stopIfTrue=True, fill=fill_red, font=font_red)
    rule_yellow = CellIsRule(operator='between', formula=[aclr_yellow_usl, aclr_red_usl], stopIfTrue=True,
                             fill=fill_yellow, font=font_yellow)
    ## rule = FormulaRule(formula=['AND($B2>0, $B2<=0.4)'], stopIfTrue=True)

    # load workbook and sheets for ACLR EUTRA
    logger.info('========== Color code judge for ACLR EUTRA ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'G2:H{ws.max_row}', rule_red)  # Apply to range G:H
            ws.conditional_formatting.add(f'G2:H{ws.max_row}', rule_yellow)  # Apply to range G:H
        else:
            pass

    wb.save(file_path)


def color_format_lte_aclr_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # import yaml file
    color_codes = import_aclr_limits()
    aclr_red_usl = color_codes['LTE']['e_utra_color_red_usl']
    aclr_yellow_usl = color_codes['LTE']['e_utra_color_yellow_usl']

    # define the color of fill and font
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C')
    font_red = Font(color='9C0006')
    font_yellow = Font(color='9C6500')

    # define the rule
    rule_red = CellIsRule(operator='greaterThan', formula=[aclr_red_usl], stopIfTrue=True, fill=fill_red, font=font_red)
    rule_yellow = CellIsRule(operator='between', formula=[aclr_yellow_usl, aclr_red_usl], stopIfTrue=True,
                             fill=fill_yellow, font=font_yellow)
    ## rule = FormulaRule(formula=['AND($B2>0, $B2<=0.4)'], stopIfTrue=True)

    # load workbook and sheets for ACLR EUTRA
    logger.info('========== Color code judge for ACLR EUTRA ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'G2:H{ws.max_row}', rule_red)  # Apply to range G:H
            ws.conditional_formatting.add(f'G2:H{ws.max_row}', rule_yellow)  # Apply to range G:H
        else:
            pass

    wb.save(file_path)


def color_format_wcdma_aclr_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # import yaml file
    color_codes = import_aclr_limits()
    aclr_red_usl = color_codes['WCDMA']['utra_color_red_usl']
    aclr_yellow_usl = color_codes['WCDMA']['utra_color_yellow_usl']

    # define the color of fill and font
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C')
    font_red = Font(color='9C0006')
    font_yellow = Font(color='9C6500')

    # define the rule
    rule_red = CellIsRule(operator='greaterThan', formula=[aclr_red_usl], stopIfTrue=True, fill=fill_red, font=font_red)
    rule_yellow = CellIsRule(operator='between', formula=[aclr_yellow_usl, aclr_red_usl], stopIfTrue=True,
                             fill=fill_yellow, font=font_yellow)
    ## rule = FormulaRule(formula=['AND($B2>0, $B2<=0.4)'], stopIfTrue=True)

    # load workbook and sheets for ACLR EUTRA
    logger.info('========== Color code judge for ACLR UTRA ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'G2:H{ws.max_row}', rule_red)  # Apply to range G:H
            ws.conditional_formatting.add(f'G2:H{ws.max_row}', rule_yellow)  # Apply to range G:H
        else:
            pass

    wb.save(file_path)


def color_format_gsm_orfs_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # import yaml file
    color_codes = import_aclr_limits()
    md_orfs_red_usl = color_codes['GSM']['md_orfs_color_red_usl']
    md_orfs_yellow_usl = color_codes['GSM']['md_orfs_color_yellow_usl']
    sw_orfs_red_usl = color_codes['GSM']['sw_orfs_color_red_usl']
    sw_orfs_yellow_usl = color_codes['GSM']['sw_orfs_color_yellow_usl']

    # define the color of fill and font
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C')
    font_red = Font(color='9C0006')
    font_yellow = Font(color='9C6500')

    # define the rule
    md_rule_red = CellIsRule(operator='greaterThan', formula=[md_orfs_red_usl], stopIfTrue=True, fill=fill_red,
                             font=font_red)
    md_rule_yellow = CellIsRule(operator='between', formula=[md_orfs_yellow_usl, md_orfs_red_usl], stopIfTrue=True,
                                fill=fill_yellow, font=font_yellow)
    sw_rule_red = CellIsRule(operator='greaterThan', formula=[sw_orfs_red_usl], stopIfTrue=True, fill=fill_red,
                             font=font_red)
    sw_rule_yellow = CellIsRule(operator='between', formula=[sw_orfs_yellow_usl, sw_orfs_red_usl], stopIfTrue=True,
                                fill=fill_yellow, font=font_yellow)
    ## rule = FormulaRule(formula=['AND($B2>0, $B2<=0.4)'], stopIfTrue=True)

    # load workbook and sheets for ORFS
    logger.info('========== Color code judge for ORFS ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'L2:M{ws.max_row}', md_rule_red)  # Apply to range L:M for md
            ws.conditional_formatting.add(f'L2:M{ws.max_row}', md_rule_yellow)  # Apply to range L:M for md
            ws.conditional_formatting.add(f'P2:Q{ws.max_row}', sw_rule_red)  # Apply to range P:Q for sw
            ws.conditional_formatting.add(f'P2:Q{ws.max_row}', sw_rule_yellow)  # Apply to range P:Q for sw
        else:
            pass

    wb.save(file_path)


def color_rule_evm(tech, mod):
    # import yaml file
    color_codes = import_evm_limits()
    evm_red_usl = color_codes[tech][mod]['evm_color_red_usl']
    evm_yellow_usl = color_codes[tech][mod]['evm_color_yellow_usl']

    # define the color of fill and font
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C')
    font_red = Font(color='9C0006')
    font_yellow = Font(color='9C6500')

    # define the rule
    rule_red = CellIsRule(operator='greaterThan', formula=[evm_red_usl], stopIfTrue=True, fill=fill_red, font=font_red)
    rule_yellow = CellIsRule(operator='between', formula=[evm_yellow_usl, evm_red_usl], stopIfTrue=True,
                             fill=fill_yellow, font=font_yellow)
    ## rule = FormulaRule(formula=['AND($B2>0, $B2<=0.4)'], stopIfTrue=True)
    return rule_red, rule_yellow


def color_format_fr1_evm_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # load workbook and sheets for ACLR EUTRA
    logger.info('========== Color code judge for EVM ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            mod = sheetname.split('_')[-1]  # split to get modulation
            rule_red, rule_yellow = color_rule_evm('FR1', mod)
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'M2:M{ws.max_row}', rule_red)  # Apply to range M:M
            ws.conditional_formatting.add(f'M2:M{ws.max_row}', rule_yellow)  # Apply to range M:M
        else:
            pass

    wb.save(file_path)


def color_format_lte_evm_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # load workbook and sheets for ACLR EUTRA
    logger.info('========== Color code judge for EVM ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            mod = sheetname.split('_')[-1]  # split to get modulation
            rule_red, rule_yellow = color_rule_evm('LTE', mod)
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'M2:M{ws.max_row}', rule_red)  # Apply to range M:M
            ws.conditional_formatting.add(f'M2:M{ws.max_row}', rule_yellow)  # Apply to range M:M
        else:
            pass

    wb.save(file_path)


def color_format_wcdma_evm_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # load workbook and sheets for ACLR EUTRA
    logger.info('========== Color code judge for EVM ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            mod = 'QPSK'
            rule_red, rule_yellow = color_rule_evm('WCDMA', mod)
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'L2:L{ws.max_row}', rule_red)  # Apply to range L:L
            ws.conditional_formatting.add(f'L2:L{ws.max_row}', rule_yellow)  # Apply to range L:L
        else:
            pass

    wb.save(file_path)


def color_format_gsm_evm_ftm(file_path):
    """
    CellIsRule:
        operator: (string) Specifies the comparison operator to use. Options include:
            'between'
            'equal'
            'greaterThan'
            'greaterThanOrEqual'
            'lessThan'
            'lessThanOrEqual'
            'notBetween'
            'notEqual'
        formula: (list) Contains the value or formula to compare against.
        stopIfTrue: (bool) Determines whether to stop evaluating other conditional formatting rules if this rule is True.
        fill: (PatternFill object) Specifies the fill formatting to apply if the condition is met.
        font: (Font object) Specifies the font formatting to apply if the condition is met.

    PatternFill:
        patternType: (string) Specifies the fill pattern. Common options:

            'solid': Solid fill with a single color.
            'lightGray': Light gray pattern.
            'mediumGray': Medium gray pattern.
            'darkGray': Dark gray pattern.
            'darkHorizontal': Dark horizontal lines.
            'darkVertical': Dark vertical lines.
            'darkDown': Dark diagonal lines sloping down.
            'darkUp': Dark diagonal lines sloping up.
            'darkGrid': Dark grid pattern.
            'lightGrid': Light grid pattern.
            'lightHorizontal': Light horizontal lines.
            'lightVertical': Light vertical lines.
            'lightDown': Light diagonal lines sloping down.
            'lightUp': Light diagonal lines sloping up.
            start_color: (string or Color object) The starting color for gradient fills.

        end_color: (string or Color object) The ending color for gradient fills.

        bgColor: (string or Color object) An alias for start_color in older openpyxl versions.

        fgColor: (string or Color object) An alias for end_color in older openpyxl versions.

    """
    # load workbook and sheets for ACLR EUTRA
    logger.info('========== Color code judge for EVM ==========')
    wb = openpyxl.load_workbook(file_path)
    for sheetname in wb.sheetnames:
        if 'Dashboard' in sheetname:
            continue
        ws = wb[sheetname]
        if ws.max_row > 1:  # if not only the header, this step can continue to be activated
            mod = sheetname.split('_')[-1]
            rule_red, rule_yellow = color_rule_evm('GSM', mod)
            # ws.conditional_formatting = ConditionalFormattingList()  # clear pattern from entire format
            ws.conditional_formatting.add(f'H2:H{ws.max_row}', rule_red)  # Apply to range H:H
            ws.conditional_formatting.add(f'H2:H{ws.max_row}', rule_yellow)  # Apply to range H:H
        else:
            pass

    wb.save(file_path)



if __name__ == '__main__':
    file_path = r'C:\Users\pricewu\Documents\meta_tool\output\1FDG65013956000B3CM0001N\Tx_Pwr_ACLR_EVM_10MHZ_FR1_LMH_color.xlsx'
    color_format_fr1_aclr_ftm(file_path)
