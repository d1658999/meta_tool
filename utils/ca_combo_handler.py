import openpyxl
from pathlib import Path


FILE_NAME = 'CA_36508.xlsx'
FILE_PATH = Path.cwd() / Path('utils') / Path('parameters') / Path(FILE_NAME)  # formal use
# FILE_PATH = Path.cwd().parents[1] / Path('utils') / Path('parameters') / Path(FILE_NAME)  # test use


def ca_combo_load_excel(band, file_path=FILE_PATH):
    """
    # return [(chan, combo_rb, cc1_rb_size, cc2_rb_size, cc1_chan, cc2_chan), ...]
    return {chan: combo_rb: (cc1_rb_size, cc2_rb_size, cc1_chan, cc2_chan), ...}
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[f'{band}']
    max_row = ws.max_row

    chan_ca_combo_dict = {}
    chan_ca_combo_dict.setdefault('L', {})
    chan_ca_combo_dict.setdefault('M', {})
    chan_ca_combo_dict.setdefault('H', {})
    if band in ['38C', '39C', '40C', '41C', '42C', '48C', ]:
        for row in range(3, max_row):
            chan = ws.cell(row, 1).value
            combo = ws.cell(row, 2).value
            bw_cc1 = ws.cell(row, 3).value
            chan_ul_cc1 = ws.cell(row, 4).value
            bw_cc2 = ws.cell(row, 6).value
            chan_ul_cc2 = ws.cell(row, 7).value
            chan_ca_combo_dict[chan][combo] = (bw_cc1, bw_cc2, chan_ul_cc1, chan_ul_cc2)
    else:
        for row in range(3, max_row):
            chan = ws.cell(row, 1).value
            combo = ws.cell(row, 2).value
            bw_cc1 = ws.cell(row, 3).value
            chan_ul_cc1 = ws.cell(row, 4).value
            bw_cc2 = ws.cell(row, 8).value
            chan_ul_cc2 = ws.cell(row, 9).value
            chan_ca_combo_dict[chan][combo] = (bw_cc1, bw_cc2, chan_ul_cc1, chan_ul_cc2)

    return chan_ca_combo_dict


def main():
    test_list = ca_combo_load_excel('7C')
    print(test_list['L']['50+100'])



if __name__ == '__main__':
    main()
